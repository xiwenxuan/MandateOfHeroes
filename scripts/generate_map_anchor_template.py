#!/usr/bin/env python3
"""生成《02_地图坐标与对象锚点表.xlsx》空白正式模板。"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "deliverables" / "02_地图坐标与对象锚点表.xlsx"
VERSION = "V0.1"
RESERVED_ROWS = 500

SHEETS = [
    "00_使用说明", "01_世界设置", "02_固定地理", "03_地点锚点", "04_路径网络",
    "05_层级表现LOD", "06_建筑锚点", "07_动态状态", "08_资产清单",
    "09_连续性验收", "10_数据字典", "11_校验结果",
]

COLORS = {
    "dark": "1F4E78", "header": "17365D", "title": "0F243E", "input": "E2F0D9",
    "formula": "E7E6E6", "warning": "FFF2CC", "error": "F4CCCC", "success": "D9EAD3",
    "white": "FFFFFF", "light_blue": "DDEBF7", "border": "B4C6E7",
}

THIN = Side(style="thin", color=COLORS["border"])
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color=COLORS["white"])
BODY_FONT = Font(name="微软雅黑", size=10, color="222222")
TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color=COLORS["white"])
INPUT_FILL = PatternFill("solid", fgColor=COLORS["input"])
FORMULA_FILL = PatternFill("solid", fgColor=COLORS["formula"])
WARNING_FILL = PatternFill("solid", fgColor=COLORS["warning"])
ERROR_FILL = PatternFill("solid", fgColor=COLORS["error"])
SUCCESS_FILL = PatternFill("solid", fgColor=COLORS["success"])
HEADER_FILL = PatternFill("solid", fgColor=COLORS["header"])

DICT = {
    "地图层级": ["L01_天下战略图", "L02_州郡区域图", "L03_县域乡野图", "L04_城镇空间图", "L05_村庄庄园图", "L06_建筑院落图", "L07_战场地图"],
    "对象类型": ["海域", "海岸线", "河流", "湖泊", "山系", "山脉", "山峰", "平原", "盆地", "丘陵", "高原", "河谷", "峡谷", "湿地", "荒地", "森林", "天然关口", "岛屿", "地点", "路径", "建筑", "其他"],
    "地点类型": ["都城", "州治", "郡治", "国治", "县城", "重要聚落", "村庄", "庄园", "坞堡", "关隘", "港口", "渡口", "驿站", "军营", "市场", "仓储节点", "寺观", "废墟", "其他"],
    "路径类型": ["一级官道", "二级官道", "县乡道路", "村路", "山路", "小径", "桥梁连接", "渡河路线", "主航线", "支航线", "运河", "临时军路", "补给线", "其他"],
    "建筑类型": ["官署", "县衙", "官仓", "武库", "驿馆", "军营", "市场", "店铺", "商号", "仓库", "客舍", "酒肆", "码头", "车马场", "医馆", "私塾", "住宅", "豪族宅院", "祠堂", "寺观", "铁匠铺", "木工作坊", "织坊", "染坊", "皮革作坊", "磨坊", "酿造场", "屠宰场", "窑场", "冶炼场", "农舍", "牲畜棚", "其他"],
    "状态编码": ["NORMAL", "CLOSED", "BUILDING", "EXPANDED", "DAMAGED", "BURNING", "ABANDONED", "OCCUPIED", "BESIEGED", "FLOODED", "LOOTED", "UNKNOWN"],
    "审核状态": ["未审核", "待审核", "审核中", "已通过", "有条件通过", "已驳回", "已废弃"],
    "制作状态": ["未开始", "草图", "确认稿", "制作中", "待审核", "返修", "已通过", "已废弃"],
    "方向": ["北", "东北", "东", "东南", "南", "西南", "西", "西北", "无"],
    "几何类型": ["点", "折线", "多边形", "复合几何"],
    "布尔值": ["是", "否"],
    "显示规则": ["是", "否", "条件显示", "情报决定"],
    "锚点规则": ["底部中心", "几何中心", "入口中心", "道路连接点", "水线中心", "自定义"],
    "通行状态": ["正常", "拥堵", "受损", "阻断", "封锁", "施工", "废弃", "未知"],
    "容量等级": ["极低", "低", "中", "高", "极高"],
    "文件格式": ["PSB", "PSD", "AI", "SVG", "BLEND", "PNG", "WEBP", "JPG", "其他"],
    "资产类别": ["主体", "覆盖层", "阴影", "文字", "旗帜", "所有权颜色", "图标", "轮廓", "室内图", "其他"],
    "检查结果": ["未检查", "通过", "有条件通过", "不通过", "不适用"],
    "检查类型": ["河流连续", "道路连续", "城市位置", "城门方向", "桥梁连接", "港口临水", "建筑位置", "入口方向", "战场继承", "损毁反馈", "其他"],
    "道路等级": ["一级", "二级", "三级", "四级", "临时", "未知"],
    "建设状态": ["规划", "建设中", "已完成", "受损", "废弃", "未知"],
    "状态类别": ["常态", "运营", "建设", "损毁", "灾害", "占领", "情报", "其他"],
    "制作格式": ["PSB", "PSD", "AI", "SVG", "BLEND", "其他"],
    "导出格式": ["PNG", "WEBP", "JPG", "其他"],
}

DICT_NAMES = {
    "地图层级": "dict_map_level", "对象类型": "dict_object_type", "地点类型": "dict_location_type",
    "路径类型": "dict_path_type", "建筑类型": "dict_building_type", "状态编码": "dict_state_code",
    "审核状态": "dict_review_status", "制作状态": "dict_production_status", "方向": "dict_direction",
    "几何类型": "dict_geometry_type", "布尔值": "dict_boolean", "显示规则": "dict_display_rule",
    "锚点规则": "dict_anchor_rule", "通行状态": "dict_passage_status", "容量等级": "dict_capacity",
    "文件格式": "dict_file_format", "资产类别": "dict_asset_type", "检查结果": "dict_check_result",
    "检查类型": "dict_check_type", "道路等级": "dict_road_level", "建设状态": "dict_build_status",
    "状态类别": "dict_state_category", "制作格式": "dict_source_format", "导出格式": "dict_export_format",
}

BUSINESS = {
    "02_固定地理": ("tbl_geography", ["对象ID", "中文名称", "备用名称", "对象类型", "父级对象ID", "世界中心X", "世界中心Y", "最小X", "最小Y", "最大X", "最大Y", "几何类型", "路径或多边形文件", "最低显示层级", "最高显示层级", "是否固定", "通行影响", "水运影响", "农业影响", "军事影响", "美术资产ID", "资料来源", "资料版本", "审核状态", "备注"]),
    "03_地点锚点": ("tbl_locations", ["地点ID", "中文名称", "备用名称", "地点类型", "父级行政对象ID", "父级地点ID", "世界X", "世界Y", "海拔或地形等级", "朝向", "最低显示层级", "最高显示层级", "天下图表现", "州郡图表现", "县域图表现", "城镇图表现", "固定位置", "允许迁移", "当前资产ID", "历史资料来源", "资料版本", "审核状态", "备注"]),
    "04_路径网络": ("tbl_paths", ["路径ID", "路径名称", "路径类型", "起点对象ID", "终点对象ID", "父级路径ID", "路径文件", "节点数量", "道路等级", "基础宽度", "基础速度系数", "容量等级", "允许军团", "允许商队", "允许平民", "允许车辆", "允许船只", "季节影响", "天气影响", "控制对象ID", "建设状态", "通行状态", "最低显示层级", "最高显示层级", "资料来源", "审核状态", "备注"]),
    "05_层级表现LOD": ("tbl_lod", ["对象ID", "对象类别", "天下战略图是否显示", "天下战略图表现", "州郡区域图是否显示", "州郡区域图表现", "县域乡野图是否显示", "县域乡野图表现", "城镇空间图是否显示", "城镇空间图表现", "村庄庄园图是否显示", "村庄庄园图表现", "建筑院落图是否显示", "建筑院落图表现", "战场图是否显示", "战场图表现", "聚合规则", "替换资产规则", "标签显示规则", "情报影响规则", "备注"]),
    "06_建筑锚点": ("tbl_buildings", ["建筑ID", "建筑名称", "建筑类型", "所属地点ID", "所属街区或村落ID", "世界X", "世界Y", "城镇局部X", "城镇局部Y", "院落局部X", "院落局部Y", "朝向", "主入口方向", "主入口锚点X", "主入口锚点Y", "建筑占地宽度", "建筑占地高度", "锚点规则", "所有权对象ID", "运营状态", "当前资产ID", "关联室内图ID", "可进入", "可破坏", "可扩建", "最低显示层级", "资料来源", "审核状态", "备注"]),
    "07_动态状态": ("tbl_states", ["状态编码", "状态名称", "状态类别", "适用对象类型", "是否替换主体资产", "主体资产后缀", "覆盖层资产后缀", "是否影响碰撞", "是否影响通行", "是否影响功能", "是否影响人口", "是否影响库存", "是否影响所有权", "进入条件", "退出条件", "显示优先级", "备注"]),
    "08_资产清单": ("tbl_assets", ["资产ID", "资产名称", "对象ID", "资产类别", "状态编码", "LOD层级", "文件名", "源文件路径", "导出文件路径", "源文件格式", "导出格式", "画布宽度px", "画布高度px", "锚点X", "锚点Y", "阴影是否独立", "文字是否独立", "旗帜是否独立", "所有权颜色是否独立", "版本", "制作人", "审核人", "制作状态", "审核状态", "备注"]),
    "09_连续性验收": ("tbl_continuity", ["验收ID", "对象ID", "检查类型", "上级地图", "下级地图", "上级地图坐标X", "上级地图坐标Y", "下级换算坐标X", "下级换算坐标Y", "坐标偏差", "允许偏差", "方向一致", "连接关系一致", "外形继承", "状态继承", "情报继承", "损毁继承", "检查结果", "问题说明", "责任人", "检查日期", "复核人", "复核日期", "备注"]),
}

STATE_ROWS = [
    ("NORMAL", "正常", "常态"), ("CLOSED", "关闭", "运营"), ("BUILDING", "建设中", "建设"),
    ("EXPANDED", "扩建", "建设"), ("DAMAGED", "损坏", "损毁"), ("BURNING", "火灾", "灾害"),
    ("ABANDONED", "废弃", "运营"), ("OCCUPIED", "被占用", "占领"), ("BESIEGED", "被围困", "占领"),
    ("FLOODED", "被淹", "灾害"), ("LOOTED", "被劫掠", "损毁"), ("UNKNOWN", "未知", "情报"),
]


def col_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(index)


def add_name(wb: Workbook, name: str, target: str) -> None:
    wb.defined_names.add(DefinedName(name, attr_text=target))


def setup_page(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddHeader.center.text = "02_地图坐标与对象锚点表.xlsx"
    ws.oddHeader.center.size = 9
    ws.oddFooter.center.text = "第 &P 页 / 共 &N 页"
    ws.oddFooter.right.text = VERSION
    ws.sheet_view.zoomScale = 85


def style_header(ws, headers: list[str]) -> None:
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(1, idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
        hint = "按字段语义填写；正式对象必须使用稳定ID并保持可追溯。"
        if header.endswith("ID") or header in {"对象ID", "地点ID", "路径ID", "建筑ID", "资产ID", "验收ID"}:
            hint = "稳定唯一ID。示例只见说明页；投入使用后不得随意修改或重用。"
        elif "世界" in header and (header.endswith("X") or header.endswith("Y")):
            hint = "世界逻辑坐标，不是图片像素坐标；必须位于01_世界设置定义的范围内。"
        elif "局部" in header or "锚点" in header:
            hint = "对象所属局部空间的坐标/规则，不得替代世界逻辑坐标。"
        elif "资料来源" in header:
            hint = "记录可复核的历史、地理或制作来源；未知时留空，不得编造。"
        cell.comment = Comment(hint, "Codex")
    ws.row_dimensions[1].height = 34


def configure_columns(ws, headers: list[str]) -> None:
    for idx, header in enumerate(headers, 1):
        width = 13
        if header.endswith("ID") or "对象ID" in header or "地点ID" in header:
            width = 21
        if any(k in header for k in ["名称", "说明", "规则", "条件", "来源", "路径", "备注", "表现"]):
            width = 20 if header not in {"备注", "问题说明"} else 28
        if len(header) >= 10:
            width = max(width, 18)
        ws.column_dimensions[col_letter(idx)].width = min(width, 30)


def add_table(ws, table_name: str, headers: list[str], rows: int = RESERVED_ROWS) -> None:
    style_header(ws, headers)
    configure_columns(ws, headers)
    end_row = rows + 1
    for row in ws.iter_rows(min_row=2, max_row=end_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = BODY_FONT
            cell.fill = INPUT_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=False)
    tab = Table(displayName=table_name, ref=f"A1:{col_letter(len(headers))}{end_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = tab.ref
    ws.print_title_rows = "1:1"
    setup_page(ws)


def add_list_validation(ws, column: int, dict_key: str, start: int = 2, end: int = RESERVED_ROWS + 1) -> None:
    dv = DataValidation(type="list", formula1=f"={DICT_NAMES[dict_key]}", allow_blank=True)
    dv.errorTitle = "输入无效"
    dv.error = f"请从“{dict_key}”下拉字典中选择，不要自行输入未登记值。"
    dv.promptTitle = "请选择标准值"
    dv.prompt = f"数据来源：10_数据字典 / {dict_key}"
    dv.showErrorMessage = True
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{col_letter(column)}{start}:{col_letter(column)}{end}")


def add_decimal_validation(ws, column: int, maximum_name: str) -> None:
    dv = DataValidation(type="decimal", operator="between", formula1="0", formula2=f"={maximum_name}", allow_blank=True)
    dv.errorTitle = "世界坐标越界"
    dv.error = "请输入世界设置范围内的数值；空白表示尚未录入，不代表零。"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{col_letter(column)}2:{col_letter(column)}{RESERVED_ROWS + 1}")


def add_positive_validation(ws, column: int, whole: bool = False) -> None:
    dv = DataValidation(type="whole" if whole else "decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv.errorTitle = "数值无效"
    dv.error = "请输入大于或等于0的数值，或保持空白。"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{col_letter(column)}2:{col_letter(column)}{RESERVED_ROWS + 1}")


def add_date_validation(ws, column: int) -> None:
    letter = col_letter(column)
    dv = DataValidation(type="date", operator="between", formula1="DATE(1900,1,1)", formula2="DATE(9999,12,31)", allow_blank=True)
    dv.errorTitle = "日期格式错误"
    dv.error = "请输入Excel日期，建议格式：yyyy-mm-dd。"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{letter}2:{letter}{RESERVED_ROWS + 1}")
    for cell in ws[letter][1:RESERVED_ROWS + 1]:
        cell.number_format = "yyyy-mm-dd"


def add_duplicate_format(ws, column: int) -> None:
    letter = col_letter(column)
    rng = f"{letter}2:{letter}{RESERVED_ROWS + 1}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND({letter}2<>"",COUNTIF(${letter}$2:${letter}${RESERVED_ROWS + 1},{letter}2)>1)'], fill=ERROR_FILL))


def create_instructions(ws) -> None:
    ws.merge_cells("A1:H1")
    ws["A1"] = "02_地图坐标与对象锚点表｜使用说明"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["title"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32
    sections = [
        ("文件用途", "统一管理世界逻辑坐标、固定地理、地点、路径、LOD、建筑、状态、美术资产和跨层级验收。"),
        ("核心地图原则", "所有地图层级表现的是同一个世界。放大只增加细节，不重新生成另一套城市、道路、河流和建筑布局。"),
        ("坐标区别", "世界逻辑坐标：全世界统一；地图局部坐标：某层/区域内部；图片像素坐标：PSB/PNG画布；建筑院落局部坐标：建筑内部或院落空间。四者不得混用。"),
        ("对象ID规则", "重要对象必须使用唯一稳定ID；PSB、PNG、程序和其他数据文件均通过对象ID关联。投入使用后不得随意改名或重用。"),
        ("ID格式", "GEO-类型-流水号；LOC-类型-流水号；PATH-类型-流水号；BLD-地点简称-类型-流水号；ASSET-对象ID-状态-LOD；CHK-流水号。示例：GEO-RIV-0001、LOC-CITY-0001、PATH-ROAD-0001、BLD-ZS-SHOP-0001、ASSET-BLD-ZS-SHOP-0001-NORMAL-L04、CHK-0001。"),
        ("推荐填写顺序", "01_世界设置 → 02_固定地理 → 03_地点锚点 → 04_路径网络 → 05_层级表现LOD → 06_建筑锚点 → 08_资产清单 → 09_连续性验收。"),
        ("美术与程序关系", "PSB/PNG只提供表现资产；程序、数据和资产以对象ID关联。动态设施、所有权、状态、库存和情报不得永久画死在背景图中。"),
        ("严禁事项", "不得凭肉眼在下一层重新放置城市；不得把世界坐标直接等同图片像素；不得在缩放或进入场景时生成第二套地点、路线或设施。"),
        ("常见错误", "零坐标与空白混淆、ID重复、对象引用不存在、路径起终点相同、坐标越界、最小值大于最大值、LOD改变世界事实、资产文件名重复。"),
        ("数据维护责任", "录入者负责来源和字段完整性；审核者负责坐标、引用、连续性和许可证；程序与美术只消费已审核的稳定对象ID。"),
        ("版本信息", f"模板版本 {VERSION}；生成脚本可重复执行并覆盖同名文件。正式录入后重新生成前必须备份或编写迁移脚本。"),
        ("修改记录", "V0.1：建立空白正式模板、系统字典、坐标公式、数据验证、条件格式和校验汇总。"),
    ]
    row = 3
    for title, text in sections:
        ws.cell(row, 1, title)
        ws.cell(row, 1).font = Font(name="微软雅黑", size=11, bold=True, color=COLORS["dark"])
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=COLORS["light_blue"])
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.cell(row, 2, text)
        ws.cell(row, 2).font = BODY_FONT
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 36 if len(text) < 90 else 54
        row += 1
    ws.column_dimensions["A"].width = 18
    for col in range(2, 9):
        ws.column_dimensions[col_letter(col)].width = 16
    ws.freeze_panes = "A3"
    setup_page(ws)


def create_world_settings(ws, wb) -> None:
    headers = ["参数编码", "参数名称", "参数值", "单位", "是否必填", "允许修改", "说明"]
    rows = [
        ("PROJECT_NAME", "项目名称", "统一世界地图系统", "文本", "是", "是", "项目显示名称"),
        ("DOC_VERSION", "文档版本", VERSION, "文本", "是", "是", "模板与数据版本"),
        ("GAME_START_YEAR", "游戏起始年份", None, "年", "是", "是", "确认历史开局后填写"),
        ("SOURCE_BASE_DATE", "历史资料基准日期", None, "日期", "是", "是", "资料冻结或审核基准日期"),
        ("WORLD_WIDTH", "世界逻辑宽度", 100000, "逻辑单位", "是", "谨慎", "世界X有效范围为0至该值"),
        ("WORLD_HEIGHT", "世界逻辑高度", 60000, "逻辑单位", "是", "谨慎", "世界Y有效范围为0至该值"),
        ("WORLD_ORIGIN", "世界原点", "左上角", "文本", "是", "否", "统一坐标原点"),
        ("X_DIRECTION", "X轴方向", "向右增加", "文本", "是", "否", "全层级一致"),
        ("Y_DIRECTION", "Y轴方向", "向下增加", "文本", "是", "否", "全层级一致"),
        ("MASTER_WIDTH_PX", "天下母版宽度px", 15360, "px", "是", "谨慎", "PSB母版画布宽度"),
        ("MASTER_HEIGHT_PX", "天下母版高度px", 8640, "px", "是", "谨慎", "PSB母版画布高度"),
        ("PREVIEW_WIDTH_PX", "预览图宽度px", 3840, "px", "是", "是", "预览PNG宽度"),
        ("PREVIEW_HEIGHT_PX", "预览图高度px", 2160, "px", "是", "是", "预览PNG高度"),
        ("COLOR_MODE", "色彩模式", "RGB", "文本", "是", "否", "美术母版色彩模式"),
        ("COLOR_DEPTH", "色彩深度", "16位", "文本", "是", "谨慎", "美术母版色彩深度"),
        ("COLOR_SPACE", "色彩空间", "sRGB", "文本", "是", "谨慎", "统一色彩空间"),
        ("LIGHT_DIRECTION", "默认光照方向", "左上至右下", "文本", "是", "是", "地图和资产统一光照"),
        ("MASTER_FORMAT", "主母版格式", "PSB", "文本", "是", "否", "主母版文件格式"),
        ("DYNAMIC_FORMAT", "动态对象格式", "PNG", "文本", "是", "否", "动态对象导出格式"),
        ("COORD_DECIMALS", "坐标小数位", 2, "位", "是", "是", "坐标显示与交换精度"),
    ]
    style_header(ws, headers)
    for r, values in enumerate(rows, 2):
        for c, value in enumerate(values, 1):
            ws.cell(r, c, value)
            ws.cell(r, c).font = BODY_FONT
            ws.cell(r, c).fill = INPUT_FILL if c == 3 else PatternFill("solid", fgColor="FFFFFF")
            ws.cell(r, c).alignment = Alignment(vertical="center", wrap_text=True)
    tab = Table(displayName="tbl_world_settings", ref=f"A1:G{len(rows)+1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    configure_columns(ws, headers)
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["A"].width = 22
    # 核心参数命名区域
    row_by_code = {code: idx for idx, (code, *_rest) in enumerate(rows, 2)}
    names = {
        "WorldWidth": "WORLD_WIDTH", "WorldHeight": "WORLD_HEIGHT", "MasterWidthPx": "MASTER_WIDTH_PX",
        "MasterHeightPx": "MASTER_HEIGHT_PX", "PreviewWidthPx": "PREVIEW_WIDTH_PX", "PreviewHeightPx": "PREVIEW_HEIGHT_PX",
        "CoordinateDecimals": "COORD_DECIMALS",
    }
    for name, code in names.items():
        add_name(wb, name, f"'01_世界设置'!$C${row_by_code[code]}")
    # 坐标换算示例：仅系统示例，不是业务对象。
    ws["I1"] = "坐标换算示例（输入世界X/Y）"
    ws["I1"].font = HEADER_FONT
    ws["I1"].fill = HEADER_FILL
    ws.merge_cells("I1:N1")
    sample_headers = ["世界X", "世界Y", "母版像素X", "母版像素Y", "预览像素X", "预览像素Y"]
    for c, h in enumerate(sample_headers, 9):
        ws.cell(2, c, h)
        ws.cell(2, c).font = HEADER_FONT
        ws.cell(2, c).fill = HEADER_FILL
    ws["I3"].fill = WARNING_FILL
    ws["J3"].fill = WARNING_FILL
    ws["K3"] = '=IF(OR(I3="",J3=""),"",ROUND(I3/WorldWidth*MasterWidthPx,CoordinateDecimals))'
    ws["L3"] = '=IF(OR(I3="",J3=""),"",ROUND(J3/WorldHeight*MasterHeightPx,CoordinateDecimals))'
    ws["M3"] = '=IF(OR(I3="",J3=""),"",ROUND(I3/WorldWidth*PreviewWidthPx,CoordinateDecimals))'
    ws["N3"] = '=IF(OR(I3="",J3=""),"",ROUND(J3/WorldHeight*PreviewHeightPx,CoordinateDecimals))'
    for cell in ws[3][10:14]:
        cell.fill = FORMULA_FILL
        cell.number_format = "0.00"
    ws["I5"] = "公式"
    ws["J5"] = "母版像素X = 世界X ÷ 世界逻辑宽度 × 母版画布宽度；Y轴及预览图同理。空白不按0计算。"
    ws.merge_cells("J5:N6")
    ws["J5"].alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(9, 15):
        ws.column_dimensions[col_letter(col)].width = 16
    setup_page(ws)


def create_dictionary(ws, wb) -> None:
    ws["A1"] = "字典分类"
    ws["B1"] = "字典值"
    ws["C1"] = "说明"
    style_header(ws, ["字典分类", "字典值", "说明"])
    row = 2
    ranges = {}
    for key, values in DICT.items():
        start = row
        for value in values:
            ws.cell(row, 1, key)
            ws.cell(row, 2, value)
            ws.cell(row, 3, "由本页统一维护；业务表下拉不得硬编码。")
            row += 1
        ranges[key] = (start, row - 1)
    tab = Table(displayName="tbl_dictionary", ref=f"A1:C{row-1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    for key, (start, end) in ranges.items():
        add_name(wb, DICT_NAMES[key], f"'10_数据字典'!$B${start}:$B${end}")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 44
    setup_page(ws)


def validations_for_sheet(ws, sheet_name: str, headers: list[str]) -> None:
    idx = {name: pos + 1 for pos, name in enumerate(headers)}
    maps = {
        "02_固定地理": {"对象类型": "对象类型", "几何类型": "几何类型", "最低显示层级": "地图层级", "最高显示层级": "地图层级", "是否固定": "布尔值", "审核状态": "审核状态"},
        "03_地点锚点": {"地点类型": "地点类型", "朝向": "方向", "最低显示层级": "地图层级", "最高显示层级": "地图层级", "固定位置": "布尔值", "允许迁移": "布尔值", "审核状态": "审核状态"},
        "04_路径网络": {"路径类型": "路径类型", "道路等级": "道路等级", "容量等级": "容量等级", "允许军团": "布尔值", "允许商队": "布尔值", "允许平民": "布尔值", "允许车辆": "布尔值", "允许船只": "布尔值", "建设状态": "建设状态", "通行状态": "通行状态", "最低显示层级": "地图层级", "最高显示层级": "地图层级", "审核状态": "审核状态"},
        "05_层级表现LOD": {"对象类别": "对象类型", "天下战略图是否显示": "显示规则", "州郡区域图是否显示": "显示规则", "县域乡野图是否显示": "显示规则", "城镇空间图是否显示": "显示规则", "村庄庄园图是否显示": "显示规则", "建筑院落图是否显示": "显示规则", "战场图是否显示": "显示规则"},
        "06_建筑锚点": {"建筑类型": "建筑类型", "朝向": "方向", "主入口方向": "方向", "锚点规则": "锚点规则", "运营状态": "状态编码", "可进入": "布尔值", "可破坏": "布尔值", "可扩建": "布尔值", "最低显示层级": "地图层级", "审核状态": "审核状态"},
        "07_动态状态": {"状态编码": "状态编码", "状态类别": "状态类别", "适用对象类型": "对象类型", "是否替换主体资产": "布尔值", "是否影响碰撞": "布尔值", "是否影响通行": "布尔值", "是否影响功能": "布尔值", "是否影响人口": "布尔值", "是否影响库存": "布尔值", "是否影响所有权": "布尔值"},
        "08_资产清单": {"资产类别": "资产类别", "状态编码": "状态编码", "LOD层级": "地图层级", "源文件格式": "制作格式", "导出格式": "导出格式", "阴影是否独立": "布尔值", "文字是否独立": "布尔值", "旗帜是否独立": "布尔值", "所有权颜色是否独立": "布尔值", "制作状态": "制作状态", "审核状态": "审核状态"},
        "09_连续性验收": {"检查类型": "检查类型", "上级地图": "地图层级", "下级地图": "地图层级", "方向一致": "布尔值", "连接关系一致": "布尔值", "外形继承": "布尔值", "状态继承": "布尔值", "情报继承": "布尔值", "损毁继承": "布尔值", "检查结果": "检查结果"},
    }
    for field, key in maps.get(sheet_name, {}).items():
        add_list_validation(ws, idx[field], key)
    id_field = headers[0]
    add_duplicate_format(ws, idx[id_field])
    # 坐标与数字验证
    for field in [f for f in headers if f in {"世界中心X", "最小X", "最大X", "世界X"}]:
        add_decimal_validation(ws, idx[field], "WorldWidth")
    for field in [f for f in headers if f in {"世界中心Y", "最小Y", "最大Y", "世界Y"}]:
        add_decimal_validation(ws, idx[field], "WorldHeight")
    for field in ["节点数量", "基础宽度", "基础速度系数", "建筑占地宽度", "建筑占地高度", "画布宽度px", "画布高度px", "锚点X", "锚点Y", "显示优先级", "允许偏差"]:
        if field in idx:
            add_positive_validation(ws, idx[field], whole=field in {"节点数量", "画布宽度px", "画布高度px", "显示优先级"})
    if sheet_name == "09_连续性验收":
        add_date_validation(ws, idx["检查日期"])
        add_date_validation(ws, idx["复核日期"])


def create_business_sheets(wb) -> None:
    for sheet_name, (table_name, headers) in BUSINESS.items():
        ws = wb[sheet_name]
        add_table(ws, table_name, headers)
        validations_for_sheet(ws, sheet_name, headers)
        # 允许预置的状态字典。
        if sheet_name == "07_动态状态":
            for r, (code, name, category) in enumerate(STATE_ROWS, 2):
                ws.cell(r, 1, code)
                ws.cell(r, 2, name)
                ws.cell(r, 3, category)
                ws.cell(r, 4, "其他")
                ws.cell(r, 5, "否")
                for c in range(8, 14):
                    ws.cell(r, c, "否")
        # 连续性偏差公式，输入不完整时保持空白。
        if sheet_name == "09_连续性验收":
            for r in range(2, RESERVED_ROWS + 2):
                ws.cell(r, 10, f'=IF(COUNTA(F{r}:I{r})<4,"",SQRT((F{r}-H{r})^2+(G{r}-I{r})^2))')
                ws.cell(r, 10).fill = FORMULA_FILL
                ws.cell(r, 10).number_format = "0.00"
            ws.conditional_formatting.add(f"J2:J{RESERVED_ROWS+1}", FormulaRule(formula=['AND(J2<>"",K2<>"",J2>K2)'], fill=ERROR_FILL))
            ws.conditional_formatting.add(f"R2:R{RESERVED_ROWS+1}", FormulaRule(formula=['R2="通过"'], fill=SUCCESS_FILL))
            ws.conditional_formatting.add(f"R2:R{RESERVED_ROWS+1}", FormulaRule(formula=['R2="不通过"'], fill=ERROR_FILL))
        # 必填空白提示：只有一行开始录入后才标黄。
        required = {
            "02_固定地理": [1, 2, 4], "03_地点锚点": [1, 2, 4, 7, 8], "04_路径网络": [1, 3, 4, 5],
            "05_层级表现LOD": [1, 2], "06_建筑锚点": [1, 2, 3, 4, 6, 7], "07_动态状态": [1, 2, 3],
            "08_资产清单": [1, 2, 3, 4, 7], "09_连续性验收": [1, 2, 3],
        }[sheet_name]
        last = col_letter(len(headers))
        for col in required:
            letter = col_letter(col)
            ws.conditional_formatting.add(f"{letter}2:{letter}{RESERVED_ROWS+1}", FormulaRule(formula=[f'AND(COUNTA($A2:${last}2)>0,{letter}2="")'], fill=WARNING_FILL))


def create_validation_results(ws) -> None:
    headers = ["问题编号", "问题级别", "工作表", "对象ID", "字段", "问题说明", "建议处理", "校验状态"]
    style_header(ws, headers)
    source_sheets = list(BUSINESS.keys())
    out_row = 2
    for source in source_sheets:
        headers_src = BUSINESS[source][1]
        last = col_letter(len(headers_src))
        for src_row in range(2, RESERVED_ROWS + 2):
            id_cell = f"'{source}'!A{src_row}"
            checks = []
            required = {
                "02_固定地理": ["对象ID", "中文名称", "对象类型"],
                "03_地点锚点": ["地点ID", "中文名称", "地点类型", "世界X", "世界Y"],
                "04_路径网络": ["路径ID", "路径类型", "起点对象ID", "终点对象ID"],
                "05_层级表现LOD": ["对象ID", "对象类别"],
                "06_建筑锚点": ["建筑ID", "建筑名称", "建筑类型", "所属地点ID", "世界X", "世界Y"],
                "07_动态状态": ["状态编码", "状态名称", "状态类别"],
                "08_资产清单": ["资产ID", "资产名称", "对象ID", "资产类别", "文件名"],
                "09_连续性验收": ["验收ID", "对象ID", "检查类型"],
            }[source]
            indices = {h: i + 1 for i, h in enumerate(headers_src)}
            for field in required:
                cell = f"'{source}'!{col_letter(indices[field])}{src_row}"
                checks.append(f'IF({cell}="","缺少{field}；","")')
            if source == "02_固定地理":
                checks += [
                    f'IF(AND(\'02_固定地理\'!F{src_row}<>"",OR(\'02_固定地理\'!F{src_row}<0,\'02_固定地理\'!F{src_row}>WorldWidth)),"世界中心X越界；","")',
                    f'IF(AND(\'02_固定地理\'!G{src_row}<>"",OR(\'02_固定地理\'!G{src_row}<0,\'02_固定地理\'!G{src_row}>WorldHeight)),"世界中心Y越界；","")',
                    f'IF(AND(\'02_固定地理\'!H{src_row}<>"",\'02_固定地理\'!J{src_row}<>"",\'02_固定地理\'!H{src_row}>\'02_固定地理\'!J{src_row}),"最小X大于最大X；","")',
                    f'IF(AND(\'02_固定地理\'!I{src_row}<>"",\'02_固定地理\'!K{src_row}<>"",\'02_固定地理\'!I{src_row}>\'02_固定地理\'!K{src_row}),"最小Y大于最大Y；","")',
                    f'IF(AND(\'02_固定地理\'!E{src_row}<>"",COUNTIF(\'02_固定地理\'!$A$2:$A${RESERVED_ROWS+1},\'02_固定地理\'!E{src_row})=0),"父级对象不存在；","")',
                ]
            if source == "03_地点锚点":
                checks += [
                    f'IF(AND(\'03_地点锚点\'!G{src_row}<>"",OR(\'03_地点锚点\'!G{src_row}<0,\'03_地点锚点\'!G{src_row}>WorldWidth)),"世界X越界；","")',
                    f'IF(AND(\'03_地点锚点\'!H{src_row}<>"",OR(\'03_地点锚点\'!H{src_row}<0,\'03_地点锚点\'!H{src_row}>WorldHeight)),"世界Y越界；","")',
                    f'IF(AND(\'03_地点锚点\'!E{src_row}<>"",COUNTIF(\'02_固定地理\'!$A$2:$A${RESERVED_ROWS+1},\'03_地点锚点\'!E{src_row})=0),"父级行政对象不存在；","")',
                    f'IF(AND(\'03_地点锚点\'!F{src_row}<>"",COUNTIF(\'03_地点锚点\'!$A$2:$A${RESERVED_ROWS+1},\'03_地点锚点\'!F{src_row})=0),"父级地点不存在；","")',
                ]
            if source == "04_路径网络":
                checks += [
                    f'IF(AND(\'04_路径网络\'!D{src_row}<>"",\'04_路径网络\'!D{src_row}=\'04_路径网络\'!E{src_row}),"起点与终点相同；","")',
                    f'IF(AND(\'04_路径网络\'!D{src_row}<>"",COUNTIF(\'02_固定地理\'!$A$2:$A${RESERVED_ROWS+1},\'04_路径网络\'!D{src_row})+COUNTIF(\'03_地点锚点\'!$A$2:$A${RESERVED_ROWS+1},\'04_路径网络\'!D{src_row})+COUNTIF(\'06_建筑锚点\'!$A$2:$A${RESERVED_ROWS+1},\'04_路径网络\'!D{src_row})=0),"起点引用对象不存在；","")',
                    f'IF(AND(\'04_路径网络\'!E{src_row}<>"",COUNTIF(\'02_固定地理\'!$A$2:$A${RESERVED_ROWS+1},\'04_路径网络\'!E{src_row})+COUNTIF(\'03_地点锚点\'!$A$2:$A${RESERVED_ROWS+1},\'04_路径网络\'!E{src_row})+COUNTIF(\'06_建筑锚点\'!$A$2:$A${RESERVED_ROWS+1},\'04_路径网络\'!E{src_row})=0),"终点引用对象不存在；","")',
                    f'IF(AND(\'04_路径网络\'!F{src_row}<>"",COUNTIF(\'04_路径网络\'!$A$2:$A${RESERVED_ROWS+1},\'04_路径网络\'!F{src_row})=0),"父级路径不存在；","")',
                ]
            if source == "05_层级表现LOD":
                checks.append(f'IF(AND({id_cell}<>"",COUNTIF(\'02_固定地理\'!$A$2:$A${RESERVED_ROWS+1},{id_cell})+COUNTIF(\'03_地点锚点\'!$A$2:$A${RESERVED_ROWS+1},{id_cell})+COUNTIF(\'04_路径网络\'!$A$2:$A${RESERVED_ROWS+1},{id_cell})+COUNTIF(\'06_建筑锚点\'!$A$2:$A${RESERVED_ROWS+1},{id_cell})=0),"LOD引用对象不存在；","")')
            if source == "06_建筑锚点":
                checks += [
                    f'IF(AND(\'06_建筑锚点\'!F{src_row}<>"",OR(\'06_建筑锚点\'!F{src_row}<0,\'06_建筑锚点\'!F{src_row}>WorldWidth)),"世界X越界；","")',
                    f'IF(AND(\'06_建筑锚点\'!G{src_row}<>"",OR(\'06_建筑锚点\'!G{src_row}<0,\'06_建筑锚点\'!G{src_row}>WorldHeight)),"世界Y越界；","")',
                    f'IF(AND(\'06_建筑锚点\'!D{src_row}<>"",COUNTIF(\'03_地点锚点\'!$A$2:$A${RESERVED_ROWS+1},\'06_建筑锚点\'!D{src_row})=0),"建筑所属地点不存在；","")',
                    f'IF(AND(\'06_建筑锚点\'!T{src_row}<>"",COUNTIF(\'07_动态状态\'!$A$2:$A${RESERVED_ROWS+1},\'06_建筑锚点\'!T{src_row})=0),"运营状态编码不存在；","")',
                ]
            if source == "08_资产清单":
                checks += [
                    f'IF(AND(\'08_资产清单\'!C{src_row}<>"",COUNTIF(\'02_固定地理\'!$A$2:$A${RESERVED_ROWS+1},\'08_资产清单\'!C{src_row})+COUNTIF(\'03_地点锚点\'!$A$2:$A${RESERVED_ROWS+1},\'08_资产清单\'!C{src_row})+COUNTIF(\'04_路径网络\'!$A$2:$A${RESERVED_ROWS+1},\'08_资产清单\'!C{src_row})+COUNTIF(\'06_建筑锚点\'!$A$2:$A${RESERVED_ROWS+1},\'08_资产清单\'!C{src_row})=0),"资产引用对象不存在；","")',
                    f'IF(AND(\'08_资产清单\'!E{src_row}<>"",COUNTIF(\'07_动态状态\'!$A$2:$A${RESERVED_ROWS+1},\'08_资产清单\'!E{src_row})=0),"状态编码不存在；","")',
                    f'IF(AND(\'08_资产清单\'!G{src_row}<>"",COUNTIF(\'08_资产清单\'!$G$2:$G${RESERVED_ROWS+1},\'08_资产清单\'!G{src_row})>1),"文件名重复；","")',
                ]
            if source == "09_连续性验收":
                checks += [
                    f'IF(AND(\'09_连续性验收\'!J{src_row}<>"",\'09_连续性验收\'!K{src_row}<>"",\'09_连续性验收\'!J{src_row}>\'09_连续性验收\'!K{src_row}),"坐标偏差超标；","")',
                    f'IF(AND(\'09_连续性验收\'!B{src_row}<>"",COUNTIF(\'02_固定地理\'!$A$2:$A${RESERVED_ROWS+1},\'09_连续性验收\'!B{src_row})+COUNTIF(\'03_地点锚点\'!$A$2:$A${RESERVED_ROWS+1},\'09_连续性验收\'!B{src_row})+COUNTIF(\'04_路径网络\'!$A$2:$A${RESERVED_ROWS+1},\'09_连续性验收\'!B{src_row})+COUNTIF(\'06_建筑锚点\'!$A$2:$A${RESERVED_ROWS+1},\'09_连续性验收\'!B{src_row})=0),"验收对象不存在；","")',
                ]
            # 重复ID。
            checks.append(f'IF(AND({id_cell}<>"",COUNTIF(\'{source}\'!$A$2:$A${RESERVED_ROWS+1},{id_cell})>1),"ID重复；","")')
            issue_formula = "=" + "&".join(checks)
            # 业务行只在稳定ID已填写后启用；预置公式不应把空白行误判为正式数据。
            active = f"{id_cell}<>\"\""
            ws.cell(out_row, 1, f'=IF(F{out_row}<>"","ISSUE-"&TEXT(ROW()-1,"00000"),"")')
            ws.cell(out_row, 2, f'=IF(F{out_row}<>"","错误","")')
            ws.cell(out_row, 3, f'=IF({active},"{source}","")')
            ws.cell(out_row, 4, f'=IF({active},{id_cell},"")')
            ws.cell(out_row, 5, f'=IF(F{out_row}<>"","综合校验","")')
            ws.cell(out_row, 6, f'=IF({active},{issue_formula[1:]},"")')
            ws.cell(out_row, 7, f'=IF(F{out_row}<>"","按字段批注、数据字典和来源记录修正后复核","")')
            ws.cell(out_row, 8, f'=IF(F{out_row}<>"","待处理","")')
            out_row += 1
    for row in ws.iter_rows(min_row=2, max_row=out_row - 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = BODY_FONT
            cell.fill = FORMULA_FILL
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    tab = Table(displayName="tbl_validation", ref=f"A1:H{out_row-1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = tab.ref
    widths = [16, 10, 18, 24, 14, 42, 38, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[col_letter(i)].width = width
    ws.conditional_formatting.add(f"F2:F{out_row-1}", FormulaRule(formula=['F2<>""'], fill=ERROR_FILL))
    setup_page(ws)


def build_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEETS:
        wb.create_sheet(name)
    create_instructions(wb["00_使用说明"])
    create_world_settings(wb["01_世界设置"], wb)
    create_dictionary(wb["10_数据字典"], wb)
    create_business_sheets(wb)
    create_validation_results(wb["11_校验结果"])
    wb["10_数据字典"].sheet_state = "hidden"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    return wb


def quality_check(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    errors = []
    if wb.sheetnames != SHEETS:
        errors.append(f"工作表顺序错误：{wb.sheetnames}")
    expected_tables = {name: table for name, (table, _headers) in BUSINESS.items()}
    for sheet, table_name in expected_tables.items():
        ws = wb[sheet]
        if table_name not in ws.tables:
            errors.append(f"{sheet}缺少表格{table_name}")
        if not ws.freeze_panes:
            errors.append(f"{sheet}缺少冻结窗格")
        if not ws.auto_filter.ref:
            errors.append(f"{sheet}缺少自动筛选")
        if not ws.data_validations.dataValidation:
            errors.append(f"{sheet}缺少数据验证")
    for name in DICT_NAMES.values():
        if name not in wb.defined_names:
            errors.append(f"缺少命名区域{name}")
    for name in ["WorldWidth", "WorldHeight", "MasterWidthPx", "MasterHeightPx", "PreviewWidthPx", "PreviewHeightPx", "CoordinateDecimals"]:
        if name not in wb.defined_names:
            errors.append(f"缺少核心命名区域{name}")
    if not str(wb["01_世界设置"]["K3"].value).startswith("=IF("):
        errors.append("缺少母版像素X换算公式")
    if not str(wb["09_连续性验收"]["J2"].value).startswith("=IF("):
        errors.append("缺少连续性坐标偏差公式")
    if not str(wb["11_校验结果"]["F2"].value).startswith("=IF("):
        errors.append("缺少校验汇总公式")
    if getattr(wb, "_external_links", []):
        errors.append("检测到外部链接")
    if wb.vba_archive is not None:
        errors.append("检测到宏")
    # 正式数据表不得误填业务示例；状态表是任务书允许的标准字典。
    for sheet in ["02_固定地理", "03_地点锚点", "04_路径网络", "05_层级表现LOD", "06_建筑锚点", "08_资产清单"]:
        ws = wb[sheet]
        if any(ws.cell(r, c).value not in (None, "") for r in range(2, RESERVED_ROWS + 2) for c in range(1, ws.max_column + 1)):
            errors.append(f"{sheet}存在预填业务数据")
    # 往返保存与再次读取。
    wb.save(path)
    load_workbook(path, data_only=False).close()
    wb.close()
    if errors:
        print("QUALITY CHECK FAILED")
        for error in errors:
            print(f"- {error}")
        raise SystemExit(1)
    print("QUALITY CHECK PASSED")
    print(f"- sheets={len(SHEETS)} order=ok")
    print(f"- business_tables={len(expected_tables)} freeze=ok filters=ok")
    print(f"- validations=ok named_ranges={len(DICT_NAMES) + 7}")
    print("- formulas=coordinate_conversion, continuity_offset, validation_summary")
    print("- external_links=0 macros=0 business_rows=blank")
    print(f"- output={path}")


def main() -> int:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(OUTPUT)
    wb.close()
    quality_check(OUTPUT)
    return 0


if __name__ == "__main__":
    sys.exit(main())
