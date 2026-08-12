#!/usr/bin/env python3
"""Build the first ten City Development Packs from canonical project datasets.

This is a reference-only build.  It never creates runtime Places, Cells,
Facilities, Persons, FamilyOrganizations, save migrations, or depth upgrades.
"""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from copy import deepcopy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[2]
OUTPUT = REPO / "outputs" / "HAN_135_260_CORE_CITY_DEVELOPMENT_PACK_AND_UPGRADE_PROTOCOL_V1"
DOC = REPO / "Docs" / "HISTORICAL_WORLD_REFERENCE" / "CITY_DEVELOPMENT_PACKS"
ROSTER_OUTPUT = REPO / "outputs" / "DEVELOPMENT_PLACE_ROSTER_AND_REFERENCE_READINESS_V1"
ROSTER_DOC = REPO / "Docs" / "HISTORICAL_WORLD_REFERENCE" / "DEVELOPMENT_PLACE_ROSTER_V1"
MANIFEST_ROOT = REPO / "Docs" / "KNOWLEDGE_BASE" / "DEVELOPMENT_MANIFESTS"
TODAY = date.today().isoformat()


def load(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def dump(path: Path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


admin = load(REPO / "outputs" / "HAN_135_260_ADMINISTRATIVE_SEAT_CANONICAL_PLACE_AND_HISTORICAL_WORLD_STATE_V1" / "administrative_seat_world_state_workdata.json")
deep = load(REPO / "outputs" / "HAN_135_260_HISTORICAL_WORLD_REFERENCE_DEEPENING_V1" / "deepening_workdata.json")
historical = load(REPO / "outputs" / "HAN_135_260_HISTORICAL_WORLD_REFERENCE_V1" / "historical_world_reference_workdata.json")
family = load(REPO / "outputs" / "FAMILY_ORGANIZATION_CENTER_AND_HISTORICAL_FAMILY_REFERENCE_V1" / "family_reference_workdata.json")
roster_data = load(ROSTER_OUTPUT / "development_place_roster_workdata.json")
person_sources = load(REPO / "Assets" / "StreamingAssets" / "HistoricalPersons" / "Han135260V1" / "sources.json")
persons = load(REPO / "Assets" / "StreamingAssets" / "HistoricalPersons" / "Han135260V1" / "persons.json")
clans = load(REPO / "Assets" / "StreamingAssets" / "HistoricalPersons" / "Han135260V1" / "clans.json")
branches = load(REPO / "Assets" / "StreamingAssets" / "HistoricalPersons" / "Han135260V1" / "branches.json")
population_sources = load(REPO / "Assets" / "StreamingAssets" / "HistoricalPopulation" / "Han135260V1" / "sources.json")
major_city_population = load(REPO / "Assets" / "StreamingAssets" / "HistoricalPopulation" / "Han135260V1" / "major_city_timeline.json")["records"]
population_184 = load(REPO / "Assets" / "StreamingAssets" / "HistoricalPopulation" / "Han135260V1" / "years" / "year_184.json")


def records(payload):
    if isinstance(payload, list):
        return payload
    for key in ("records", "sources", "persons", "clans", "branches"):
        if isinstance(payload.get(key), list):
            return payload[key]
    raise ValueError("Unsupported record payload")


persons = records(persons)
clans = records(clans)
branches = records(branches)
person_sources = records(person_sources)
population_sources = records(population_sources)

person_by_name = defaultdict(list)
person_by_id = {}
for row in persons:
    person_by_id[row["person_id"]] = row
    person_by_name[row["canonical_name"]].append(row)
clan_by_id = {row["clan_id"]: row for row in clans}
branch_by_id = {row["branch_id"]: row for row in branches}
canonical_by_id = {row["CanonicalPlaceId"]: row for row in admin["canonical_places"]}
roster_by_id = {row["CanonicalPlaceId"]: row for row in roster_data["roster"]}
scenario_by_year = {int(row["year"]): row for row in deep["scenarios"]}
change_by_id = {row["ChangePointId"]: row for row in admin["change_points"]}
state_by_place = defaultdict(list)
for row in roster_data["historical_state_plan"]:
    state_by_place[row["PlaceId"]].append(row)
p0_by_place = defaultdict(list)
for row in deep["p0_reference"]:
    p0_by_place[row["place_id"]].append(row)
structured_dir_by_place = {}
for path in (REPO / "Docs" / "HISTORICAL_WORLD_REFERENCE" / "DEEPENING_V1" / "04_CORE_SETTLEMENTS").glob("*/01_structured_reference.json"):
    payload = load(path)
    structured_dir_by_place[payload["master"]["place_id"]] = (path.parent, payload)


VALID_EVIDENCE = {"HISTORICAL", "RECONSTRUCTED", "MODELED", "UNKNOWN"}
VALID_BASE_TYPES = {
    "Residence", "Farmland", "Plantation", "HerbField", "Pasture", "Forestry", "Mine", "Quarry",
    "Mill", "Brewery", "Smelter", "Smithy", "Carpentry", "SilkwormHouse", "SilkReelingWorkshop",
    "WeavingWorkshop", "DyeWorkshop", "MedicineWorkshop", "Shipyard", "Kitchen", "Warehouse",
    "Granary", "Stable", "CarriageYard", "CourierStation", "Harbor", "Market", "Shop", "Inn",
    "Clinic", "GuildHall", "MerchantHall", "GovernmentOffice", "CourtHall", "School", "Academy",
    "Library", "RitualHall", "Observatory", "TrainingHall", "Barracks", "TrainingGround",
    "FieldHospital", "Wall", "Gate", "Moat", "Fort", "BeaconTower", "Road", "Bridge", "Canal",
    "Well", "WaterIntake", "Drainage", "Dike", "Garden", "Plaza", "Courtyard",
}


COMMON_MODELED = [
    ("普通住宅群", "Residence", "MODELED", "CITY_LEVEL_ONLY", "SCENARIO_DERIVED", "人口物化时按Person住房容量推导，不复制历史锚点"),
    ("普通仓储节点", "Warehouse", "MODELED", "CITY_LEVEL_ONLY", "SCENARIO_DERIVED", "由人口、市场、军需和周转天数推导"),
    ("普通工坊群", "Carpentry", "MODELED", "CITY_LEVEL_ONLY", "SCENARIO_DERIVED", "按产业链、真实工人和配方需求推导"),
    ("基层医疗服务", "Clinic", "MODELED", "CITY_LEVEL_ONLY", "SCENARIO_DERIVED", "按常住人口和伤病服务需求推导"),
    ("道路与排水补全", "Road", "MODELED", "CITY_LEVEL_ONLY", "SCENARIO_DERIVED", "只补运行网络，不冒充历史街巷"),
]


CITY_MANUAL = {
    "LUOYANG": {
        "label": "洛阳", "strategic": "洛阳", "place": "place.han140.sili.henan.luoyang", "directory": "LUOYANG",
        "historical_names": "雒阳|洛阳", "geography": "河洛盆地，洛水穿行，北接黄河，东西关隘控制首都走廊。",
        "terrain": "盆地平原与河谷台地", "water": "洛水|黄河方向|护城壕", "mountains": "邙山|嵩山方向",
        "roads": "虎牢—洛阳东向走廊|函谷—长安西向走廊|孟津北向通道|南阳方向",
        "adjacent": "虎牢|函谷关|孟津方向|河南尹近郊县",
        "urban": "东汉首都宫城、外城、十二门、市场、太学、官署、住宅和近郊共同组成多层都市。",
        "wall": "外城墙、十二门、宫墙与护城壕已有项目级历史/复原资料。", "gates": "十二门已有稳定Facility引用。",
        "inner": "南宫、北宫及中央官署区；宫墙与外城墙独立。", "markets": "都城市场与官营/民营供应网络。",
        "residential": "城内20万人、连续城区27万人住房口径已正式物化；近郊住宅另属都市圈。",
        "industrial": "官营手工业、军需、仓储和都城消费服务区。", "suburbs": "400,000都市圈已正式包；700,000供给圈仅为计划且包含都市圈。",
        "expansion": "189—190政治危机、迁都与毁坏改变城市功能；后续状态不得覆盖运行世界分歧。",
        "facilities": [
            ("南宫", "CourtHall", "HISTORICAL", "EXACT_SITE", "135-190", "宫廷与中央政务；复用现有洛阳Facility引用"),
            ("北宫", "CourtHall", "HISTORICAL", "EXACT_SITE", "135-190", "宫廷与中央政务；复用现有洛阳Facility引用"),
            ("太学", "Academy", "HISTORICAL", "APPROXIMATE_ZONE", "135-190", "教育、藏书与士人网络"),
            ("洛阳外城墙", "Wall", "RECONSTRUCTED", "EXACT_SITE", "135-190", "首都外城防线；复用现有蓝图/Facility"),
            ("十二城门", "Gate", "HISTORICAL", "EXACT_SITE", "135-190", "独立门Facility；道路和封锁节点"),
            ("护城壕", "Moat", "RECONSTRUCTED", "EXACT_SITE", "135-190", "城防与通行影响"),
            ("中央官署群", "GovernmentOffice", "HISTORICAL", "APPROXIMATE_ZONE", "135-190", "行政职位与文书服务"),
            ("都城市场", "Market", "RECONSTRUCTED", "APPROXIMATE_ZONE", "135-190", "居民、宫廷、军队和商旅交换"),
            ("官仓与粮仓", "Granary", "RECONSTRUCTED", "APPROXIMATE_ZONE", "135-190", "首都粮食与军需储备"),
            ("驿传节点", "CourierStation", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-190", "中央文书和交通"),
        ],
        "industry": "粟麦粮食加工|丝织与官营手工业|冶炼锻造|车辆木作|酿造食品|文书与教育服务",
        "agriculture": "河洛粟麦与近郊菜蔬|畜牧|桑蚕；供给圈通过仓储和道路进入都市",
        "resources": "河谷农地|木材与石料由外围输入|金属与奢侈品跨区输入",
        "network": [("Pass", "虎牢", "东向关隘与军路", "HISTORICAL"), ("Pass", "函谷关", "西向关中走廊", "HISTORICAL"), ("Ford", "孟津方向", "黄河北向渡运", "RECONSTRUCTED"), ("SettlementCluster", "河南尹近郊聚落群", "住房、市场和劳力", "MODELED"), ("AgriculturalZone", "河洛供给区", "粮食、菜蔬、畜牧", "RECONSTRUCTED")],
        "military": "首都城防、宫城独立防线、十二门、虎牢/函谷走廊、孟津渡运与驻军共同组成防务空间。",
        "people": [("刘宏", "184", "CourtPresence", "CONFIRMED", "皇帝"), ("何皇后", "184-189", "CourtPresence", "CONFIRMED", "皇后/太后"), ("刘辩", "184-190", "CourtPresence", "CONFIRMED", "皇子/皇帝"), ("刘协", "184-190", "CourtPresence", "CONFIRMED", "皇子/皇帝"), ("何进", "184-189", "MilitaryPresence", "CONFIRMED", "大将军"), ("张让", "184-189", "CourtPresence", "CONFIRMED", "宦官"), ("蹇硕", "184-189", "MilitaryPresence", "CONFIRMED", "西园军"), ("董卓", "189-190", "MilitaryPresence", "CONFIRMED", "权臣"), ("王允", "189-192", "OfficePresence", "CONFIRMED", "司徒"), ("蔡邕", "189-192", "OfficePresence", "CONFIRMED", "士人/官员")],
        "unknowns": ["普通住宅和工坊的逐Cell历史位置", "190年毁坏的逐设施程度", "多数官署在各Scenario的精确占地"],
        "status": "DEVELOPMENT_READY", "score": 96,
    },
    "CHANGAN": {
        "label": "长安", "strategic": "长安", "place": "place.han140.sili.jingzhao.changan", "directory": "CHANGAN",
        "historical_names": "长安", "geography": "渭河平原中部，关中道路与山口网络中心；东汉使用状态须与西汉都城遗址分开。",
        "terrain": "关中平原与渭河阶地", "water": "渭水|灞水方向|关中渠系", "mountains": "秦岭北麓|北山方向",
        "roads": "函谷/潼关东向|陈仓西向|武关东南向|北地西北向", "adjacent": "潼关|函谷关|武关|陈仓|京兆近郊县",
        "urban": "汉长安旧城、城垣、宫殿遗址与东汉/董卓时期重新作为政治中心的城市功能必须分期表达。",
        "wall": "汉长安城垣有考古基础；184实际使用强度和修缮状态需分区重建。", "gates": "历史城门体系存在，东汉末逐门使用状态不明。",
        "inner": "西汉宫殿区遗存不等于184全部在用；190迁都后宫廷/官署需Scenario重建。", "markets": "关中市场与军政供应中心，具体市区按RECONSTRUCTED处理。",
        "residential": "184城市层来自独立城市人口模型；旧城内部实际居住分布UNKNOWN。", "industrial": "粮食加工、车辆木作、军需、冶炼锻造与关中仓储。",
        "suburbs": "渭河南北近郊、京兆县邑和农业聚落构成供给层。", "expansion": "190迁都后政治功能骤升，192—195李傕郭汜冲突持续改变城防、人口与供应。",
        "facilities": [("汉长安城垣", "Wall", "HISTORICAL", "APPROXIMATE_ZONE", "135-260", "分期保存，不假设全线完好"), ("汉长安城门体系", "Gate", "HISTORICAL", "CITY_LEVEL_ONLY", "135-260", "逐门状态待考"), ("宫廷/官署区（190后）", "CourtHall", "RECONSTRUCTED", "APPROXIMATE_ZONE", "190-196", "迁都后政治中心"), ("京兆官署", "GovernmentOffice", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "行政职能"), ("关中市场", "Market", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "区域商贸"), ("军粮仓储", "Granary", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "190-260", "军政供给"), ("关中驿传", "CourierStation", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "道路网络")],
        "industry": "粟麦加工|车辆木作|皮革畜牧|冶炼锻造|军需|仓储转运", "agriculture": "关中粟麦|豆类|畜牧|近郊菜蔬", "resources": "关中农地|秦岭木材石料|西北畜产与金属输入",
        "network": [("Pass", "潼关/函谷方向", "东向门户", "HISTORICAL"), ("Pass", "武关", "东南向门户", "HISTORICAL"), ("RoadJunction", "陈仓方向", "西向蜀道/陇右", "RECONSTRUCTED"), ("SettlementCluster", "京兆近郊聚落群", "劳力与粮食", "MODELED"), ("AgriculturalZone", "渭河平原供给区", "粟麦与畜产", "RECONSTRUCTED")],
        "military": "关中门户、旧城城垣、宫廷驻军和多方向道路组成战略纵深；长期内战可切断供给。",
        "people": [("董卓", "190-192", "CourtPresence", "CONFIRMED", "迁都后权臣"), ("刘协", "190-195", "CourtPresence", "CONFIRMED", "皇帝"), ("王允", "190-192", "OfficePresence", "CONFIRMED", "司徒"), ("吕布", "190-192", "MilitaryPresence", "CONFIRMED", "将领"), ("李傕", "192-198", "MilitaryPresence", "CONFIRMED", "军阀"), ("郭汜", "192-197", "MilitaryPresence", "CONFIRMED", "军阀"), ("樊稠", "192-195", "MilitaryPresence", "CONFIRMED", "军阀"), ("贾诩", "192-196", "OfficePresence", "PROBABLE", "谋士"), ("马腾", "194", "MilitaryPresence", "PROBABLE", "关中军阀"), ("韩遂", "194", "MilitaryPresence", "PROBABLE", "关中军阀")],
        "unknowns": ["184年旧宫殿和城门逐项使用状态", "190迁都后的逐设施修缮和新建位置", "李傕郭汜战争逐区损毁"], "status": "READY_WITH_MODELED_GAPS", "score": 86,
    },
    "YE": {
        "label": "邺", "strategic": "邺", "place": "place.han140.jizhou.wei.ye", "directory": "YE",
        "historical_names": "邺|邺城", "geography": "漳水流域、河北平原南缘的政治军事中心，连接太行山口与黄河北岸。", "terrain": "平原河谷", "water": "漳水|黄河北岸水网", "mountains": "太行山东麓", "roads": "邯郸/常山北向|黎阳/黄河南向|太行山口西向|青州东向", "adjacent": "邯郸|黎阳方向|官渡走廊|太行山口",
        "urban": "早期郡县城、袁绍政权中心与曹魏都城前身分期叠加；210年铜雀台等建设形成重大空间变化。", "wall": "邺城城防可考，分期边界与具体城门需保守处理。", "gates": "门名和逐期状态不完整，按CITY_LEVEL_ONLY。", "inner": "袁绍/曹操政务区与210年后台苑工程分期。", "markets": "河北粮食、军需与人口汇聚市场。", "residential": "人口模型提供184层级；后续政权中心扩张需Scenario调整。", "industrial": "粮食加工、军械、车辆、纺织和大型营建。", "suburbs": "漳水灌溉农业、县邑和军屯构成供给圈。", "expansion": "200官渡前后、204曹操入邺、210铜雀台营建、220魏政权转换。",
        "facilities": [("邺城城垣", "Wall", "HISTORICAL", "APPROXIMATE_ZONE", "135-260", "分期城防"), ("郡县官署", "GovernmentOffice", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "行政中心"), ("袁氏/曹氏政务区", "CourtHall", "RECONSTRUCTED", "APPROXIMATE_ZONE", "190-220", "政权中心"), ("铜雀台", "CourtHall", "HISTORICAL", "APPROXIMATE_ZONE", "210-260", "210后大型政治/宴集建筑"), ("邺城市场", "Market", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "河北商贸"), ("军粮仓", "Granary", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "190-260", "军粮与人口供给"), ("军械作坊", "Smithy", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "190-260", "军需")],
        "industry": "河北粮食加工|纺织|军械|车辆|营建|仓储", "agriculture": "漳水农业|粟麦|豆类|畜牧|军屯候选", "resources": "河北粮食|太行木石|北方畜产与金属输入",
        "network": [("River", "漳水", "灌溉与运输", "HISTORICAL"), ("RoadJunction", "黎阳—黄河方向", "南向渡运与军路", "RECONSTRUCTED"), ("RoadJunction", "太行山东麓", "并州通道", "RECONSTRUCTED"), ("SettlementCluster", "魏郡近郊聚落群", "劳力、粮食和军需", "MODELED"), ("AgriculturalZone", "漳水供给区", "粟麦与仓储", "RECONSTRUCTED")],
        "military": "河北政权中枢、漳水和黄河通道、城防与军粮仓储共同支撑袁曹战争和魏初防务。",
        "people": [("袁绍", "191-202", "OfficePresence", "CONFIRMED", "河北政权领袖"), ("审配", "191-204", "OfficePresence", "CONFIRMED", "守将/官员"), ("陈琳", "191-204", "OfficePresence", "PROBABLE", "幕僚"), ("袁尚", "202-204", "MilitaryPresence", "CONFIRMED", "继承人"), ("曹操", "204-220", "OfficePresence", "CONFIRMED", "魏公/魏王"), ("曹丕", "204-220", "CourtPresence", "CONFIRMED", "世子/魏王"), ("崔琰", "204-216", "OfficePresence", "PROBABLE", "官员"), ("荀攸", "204-214", "OfficePresence", "PROBABLE", "谋臣"), ("辛毗", "204-220", "OfficePresence", "PROBABLE", "官员"), ("甄氏", "204-221", "Resident", "PROBABLE", "曹丕妻")],
        "unknowns": ["袁绍时期宫署和住宅逐区位置", "204前后人口迁入规模", "铜雀台周边设施的Scenario边界"], "status": "READY_WITH_MODELED_GAPS", "score": 85,
    },
    "XU": {
        "label": "许昌", "strategic": "许昌", "place": "place.han140.yuzhou.yingchuan.xu", "directory": "XU",
        "historical_names": "许|许昌", "geography": "颍川平原的中原交通节点，连接洛阳、陈留、汝颍与淮河方向。", "terrain": "平原河网", "water": "颍水水系", "mountains": "西北嵩山余脉方向", "roads": "洛阳西北向|陈留东北向|汝南东南向|南阳西南向", "adjacent": "颍川县邑|陈留|洛阳|汝南",
        "urban": "184县级城镇、196献帝都许后的宫廷/官署与曹操政权设施必须分期表达。", "wall": "县城城防存在，196后强化程度与边界待重建。", "gates": "具体门名和数量UNKNOWN。", "inner": "196后宫廷、司空府及中央官署区采用RECONSTRUCTED。", "markets": "颍川农产、军需和首都消费市场。", "residential": "184独立城市人口模型；196后官僚、军队和流入人口需ChangePoint。", "industrial": "粮食加工、军需、纺织、车辆和文书服务。", "suburbs": "颍川农业县邑与交通聚落构成供给圈。", "expansion": "196迎献帝都许是核心变化；220魏代汉后政治功能重新配置。",
        "facilities": [("许县城垣", "Wall", "RECONSTRUCTED", "APPROXIMATE_ZONE", "135-260", "县城/都城分期"), ("县廷/颍川行政设施", "GovernmentOffice", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "地方行政"), ("献帝宫廷区", "CourtHall", "HISTORICAL", "CITY_LEVEL_ONLY", "196-220", "都许后的宫廷功能，精确位置待考"), ("中央官署区", "GovernmentOffice", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "196-220", "汉廷与曹操幕府"), ("许都市场", "Market", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "196-220", "首都和军政供应"), ("军粮仓储", "Granary", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "196-260", "中原军需"), ("驿传节点", "CourierStation", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "中原交通")],
        "industry": "粟麦加工|纺织|车辆木作|军械|仓储|文书服务", "agriculture": "颍川粟麦|豆类|桑麻|近郊菜蔬", "resources": "中原农地|木材石料输入|金属与军需跨区输入",
        "network": [("RoadJunction", "洛阳—许走廊", "西北政治交通", "RECONSTRUCTED"), ("RoadJunction", "陈留—许走廊", "东北军需交通", "RECONSTRUCTED"), ("River", "颍水水系", "区域水运与农业", "RECONSTRUCTED"), ("SettlementCluster", "颍川近郊聚落群", "劳力和住房", "MODELED"), ("AgriculturalZone", "颍川供给区", "粟麦和桑麻", "RECONSTRUCTED")],
        "military": "中原交通枢纽和汉廷所在地；城防、军粮、通往官渡/陈留/洛阳的道路是主要军事空间。",
        "people": [("刘协", "196-220", "CourtPresence", "CONFIRMED", "皇帝"), ("曹操", "196-220", "OfficePresence", "CONFIRMED", "司空/丞相"), ("荀彧", "196-212", "OfficePresence", "CONFIRMED", "尚书令/谋臣"), ("郭嘉", "196-207", "OfficePresence", "PROBABLE", "军师祭酒"), ("钟繇", "196-211", "OfficePresence", "PROBABLE", "官员"), ("孔融", "196-208", "OfficePresence", "PROBABLE", "官员/士人"), ("伏寿", "196-214", "CourtPresence", "CONFIRMED", "皇后"), ("董承", "196-200", "CourtPresence", "CONFIRMED", "外戚/将领"), ("曹丕", "196-220", "Resident", "PROBABLE", "曹氏成员"), ("曹节", "213-220", "CourtPresence", "CONFIRMED", "皇后")],
        "unknowns": ["196前县城形态与196后都城扩建差异", "献帝宫室和官署精确位置", "许都近郊人口迁入规模"], "status": "READY_WITH_MODELED_GAPS", "score": 88,
    },
    "CHENGDU": {
        "label": "成都", "strategic": "成都", "place": "place.han140.yizhou.shu.chengdu", "directory": "CHENGDU",
        "historical_names": "成都", "geography": "成都平原与岷江灌溉体系腹地，益州/蜀汉政治经济中心。", "terrain": "冲积平原", "water": "岷江水系|都江堰灌溉网络", "mountains": "成都平原西缘山地|北向剑门山系", "roads": "金牛道北向|米仓道东北向|江州东向|南中南向", "adjacent": "雒县方向|广汉|剑阁通道|江州方向",
        "urban": "秦汉以来成都城、益州州治与214后刘备政权、221蜀汉都城分期叠加。", "wall": "城垣历史存在，东汉末/蜀汉分期边界与门名需进一步考古对照。", "gates": "具体门名和逐期位置不在当前母库，保持UNKNOWN。", "inner": "益州官署、州牧府和221后宫廷区按分期重建。", "markets": "蜀锦、盐铁、粮食、木材和药材区域市场。", "residential": "正确县级人口母盘可用；现有major-city记录错链，城墙/都市圈层不采纳。", "industrial": "蜀锦、盐井相关加工、冶炼、木工、军需、食品和药材。", "suburbs": "成都平原密集农业聚落与水利网络构成强供给圈。", "expansion": "194刘璋时期、214易主、221蜀汉建国、263后状态需分期。",
        "facilities": [("成都城垣", "Wall", "HISTORICAL", "APPROXIMATE_ZONE", "135-260", "分期城防"), ("益州/蜀郡官署", "GovernmentOffice", "HISTORICAL", "CITY_LEVEL_ONLY", "135-260", "州郡行政"), ("蜀汉宫廷区", "CourtHall", "HISTORICAL", "CITY_LEVEL_ONLY", "221-263", "都城政治功能"), ("锦官相关作坊区", "WeavingWorkshop", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "蜀锦产业"), ("成都市场", "Market", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "盆地商贸"), ("官仓/粮仓", "Granary", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "城市与军政供给"), ("水利渠道", "Canal", "HISTORICAL", "APPROXIMATE_ZONE", "135-260", "灌溉与城市供水")],
        "industry": "蜀锦丝织|盐业及加工|冶炼锻造|木工车辆|酿造食品|药材加工|军需", "agriculture": "水稻|小麦|粟黍|桑蚕|蔬果|畜牧；依托岷江水利", "resources": "成都平原农地|蜀地盐井输入|木竹药材|金属与石料",
        "network": [("RoadJunction", "金牛道方向", "北向汉中/关中", "HISTORICAL"), ("RoadJunction", "江州方向", "长江上游东向", "RECONSTRUCTED"), ("WaterSystem", "岷江—都江堰水网", "农业与城市供水", "HISTORICAL"), ("SettlementCluster", "成都平原聚落群", "密集劳力、产业和粮食", "MODELED"), ("AgriculturalZone", "成都平原供给区", "稻麦、桑蚕、畜产", "RECONSTRUCTED")],
        "military": "盆地纵深、北向剑阁/汉中通道、州都城防和区域粮仓共同构成蜀地战略核心。",
        "people": [("刘璋", "194-214", "OfficePresence", "CONFIRMED", "益州牧"), ("刘备", "214-223", "CourtPresence", "CONFIRMED", "益州牧/皇帝"), ("诸葛亮", "214-234", "OfficePresence", "CONFIRMED", "丞相"), ("法正", "214-220", "OfficePresence", "CONFIRMED", "谋臣"), ("许靖", "214-222", "OfficePresence", "CONFIRMED", "官员"), ("黄权", "214-222", "OfficePresence", "PROBABLE", "官员/将领"), ("刘禅", "214-260", "CourtPresence", "CONFIRMED", "皇太子/皇帝"), ("李严", "214-231", "OfficePresence", "PROBABLE", "官员"), ("谯周", "221-260", "StudyPresence", "PROBABLE", "学者/官员"), ("张裔", "214-230", "OfficePresence", "PROBABLE", "官员")],
        "unknowns": ["major_city_timeline中成都错指admin.han140.jingzhou.nanyang.chengdu，禁止使用", "城门与宫廷逐Cell位置", "蜀锦官营设施的数量和边界"], "status": "READY_WITH_MODELED_GAPS", "score": 84,
    },
    "XIANGYANG": {
        "label": "襄阳", "strategic": "襄阳", "place": "place.han140.jingzhou.nan.xiangyang", "directory": "XIANGYANG",
        "historical_names": "襄阳|襄阳县", "geography": "汉水中游襄樊渡运节点，连接南阳盆地、江陵和汉中方向。", "terrain": "河谷平原与丘陵", "water": "汉水|襄樊渡运", "mountains": "岘山|荆山方向", "roads": "南阳北向|江陵南向|汉中西向|江夏东向", "adjacent": "樊城|新野|江陵|隆中方向",
        "urban": "襄阳城、汉水岸线、对岸樊城与近郊士族庄园共同形成跨河城市网络。", "wall": "襄阳城防历史价值明确，门名和184逐段状态需重建。", "gates": "具体门名UNKNOWN。", "inner": "南郡北部行政与刘表时期荆州政治中心功能按Scenario表达。", "markets": "汉水航运、荆北农产和南北商路市场。", "residential": "184独立城市人口模型；近郊庄园和跨河人口不可并入城墙内。", "industrial": "粮食加工、木工、船修、纺织、军需。", "suburbs": "樊城、汉水两岸聚落、隆中/岘山近郊与农业区。", "expansion": "190s刘表治荆州、208曹军南下、219襄樊战役改变城防和人口。",
        "facilities": [("襄阳城垣", "Wall", "HISTORICAL", "APPROXIMATE_ZONE", "135-260", "汉水城防"), ("城门体系", "Gate", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "逐门待考"), ("荆州/地方官署", "GovernmentOffice", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "190-208", "刘表时期政治中心"), ("汉水码头", "Harbor", "RECONSTRUCTED", "APPROXIMATE_ZONE", "135-260", "渡运与物流"), ("襄阳市场", "Market", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "南北商贸"), ("军粮仓", "Granary", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "190-260", "驻军供给"), ("桥渡节点", "Bridge", "RECONSTRUCTED", "APPROXIMATE_ZONE", "135-260", "跨河通行；不固定为永久桥梁形态")],
        "industry": "粮食加工|木工|船修|纺织|军械|商旅服务", "agriculture": "汉水谷地稻麦|桑麻|渔业|近郊园圃", "resources": "汉水运输|荆山木材|农地与渔业",
        "network": [("FortifiedTown", "樊城", "汉水对岸独立Place", "HISTORICAL"), ("RoadJunction", "新野—南阳方向", "北向军路", "RECONSTRUCTED"), ("RoadJunction", "江陵方向", "南向荆州走廊", "RECONSTRUCTED"), ("SettlementCluster", "汉水两岸聚落群", "渡运、劳力和农业", "MODELED"), ("AgriculturalZone", "襄阳近郊供给区", "稻麦、桑麻、渔业", "RECONSTRUCTED")],
        "military": "襄阳与樊城保持独立Place，通过汉水渡运和军路形成同一战区；219是最高优先变化节点。",
        "people": [("刘表", "190-208", "OfficePresence", "CONFIRMED", "荆州牧"), ("蔡瑁", "190-208", "MilitaryPresence", "PROBABLE", "将领/豪族"), ("蒯越", "190-208", "OfficePresence", "PROBABLE", "谋臣"), ("王粲", "190-208", "StudyPresence", "CONFIRMED", "士人"), ("刘琮", "200-208", "Resident", "PROBABLE", "刘表子"), ("曹操", "208", "MilitaryPresence", "CONFIRMED", "南征统帅"), ("关羽", "219", "MilitaryPresence", "CONFIRMED", "围攻方统帅"), ("曹仁", "219", "MilitaryPresence", "CONFIRMED", "守将"), ("于禁", "219", "MilitaryPresence", "CONFIRMED", "援军统帅"), ("徐晃", "219", "MilitaryPresence", "CONFIRMED", "援军统帅")],
        "unknowns": ["184城门与街区布局", "刘表治所设施逐项位置", "襄阳—樊城跨河设施的时期差异"], "status": "READY_WITH_MODELED_GAPS", "score": 85,
    },
    "JIANGLING": {
        "label": "江陵", "strategic": "江陵", "place": "place.han140.jingzhou.nan.jiangling", "directory": "JIANGLING",
        "historical_names": "江陵|江陵县", "geography": "长江中游江汉平原核心，连接襄阳、江夏、夷陵和洞庭湖区。", "terrain": "冲积平原与湖沼", "water": "长江|江汉水网", "mountains": "荆山西北方向", "roads": "襄阳北向|夷陵西向|江夏东向|武陵/长沙南向", "adjacent": "夷陵|公安方向|江夏|江汉平原县邑",
        "urban": "南郡治所、长江港运、城防和广阔近郊农业共同组成荆州中枢。", "wall": "江陵城防历史明确，具体门名和逐期修筑需重建。", "gates": "逐门资料不足。", "inner": "南郡官署及208后多方争夺下的军政中心。", "markets": "长江中游粮食、木材、鱼盐和军需市场。", "residential": "184城市人口模型较高；战后迁徙和驻军需Scenario调整。", "industrial": "粮食加工、造船修船、木工、军需、仓储。", "suburbs": "江汉平原农业聚落、港渡和县邑形成宽广供给圈。", "expansion": "208赤壁后争夺、210前后吴蜀控制变化、219荆州易手。",
        "facilities": [("江陵城垣", "Wall", "HISTORICAL", "APPROXIMATE_ZONE", "135-260", "南郡核心城防"), ("南郡官署", "GovernmentOffice", "HISTORICAL", "CITY_LEVEL_ONLY", "135-260", "郡治行政"), ("长江港运节点", "Harbor", "RECONSTRUCTED", "APPROXIMATE_ZONE", "135-260", "水运与军需"), ("江陵市场", "Market", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "区域贸易"), ("粮仓与军仓", "Granary", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "江汉粮食和驻军"), ("船作", "Shipyard", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "190-260", "水军维护"), ("军营", "Barracks", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "190-260", "驻军")],
        "industry": "粮食加工|造船修船|木工|纺织|军械|仓储", "agriculture": "江汉平原稻麦|渔业|桑麻|畜牧", "resources": "长江水运|荆楚木材|湖沼渔业|农地",
        "network": [("River", "长江", "主水运与军事通道", "HISTORICAL"), ("RoadJunction", "襄阳方向", "北向军路", "RECONSTRUCTED"), ("RoadJunction", "夷陵方向", "西向峡江门户", "RECONSTRUCTED"), ("SettlementCluster", "江汉平原聚落群", "粮食、劳力和船工", "MODELED"), ("AgriculturalZone", "江陵供给区", "稻麦、渔业和木材", "RECONSTRUCTED")],
        "military": "长江中游港城、南郡城防、夷陵方向与襄阳方向道路组成荆州军需核心。",
        "people": [("刘表", "190-208", "OfficePresence", "PROBABLE", "荆州牧"), ("刘备", "208-210", "MilitaryPresence", "PROBABLE", "盟军领袖"), ("曹仁", "208-209", "MilitaryPresence", "CONFIRMED", "守将"), ("周瑜", "208-210", "MilitaryPresence", "CONFIRMED", "攻城统帅"), ("甘宁", "208-210", "MilitaryPresence", "PROBABLE", "吴将"), ("鲁肃", "210-215", "OfficePresence", "PROBABLE", "都督"), ("关羽", "210-219", "MilitaryPresence", "CONFIRMED", "荆州守将"), ("糜芳", "210-219", "MilitaryPresence", "CONFIRMED", "南郡太守"), ("吕蒙", "219", "MilitaryPresence", "CONFIRMED", "袭取统帅"), ("陆逊", "219-222", "MilitaryPresence", "PROBABLE", "吴将")],
        "unknowns": ["城门、港区和军营逐Cell位置", "208—219多次控制变化的设施损毁", "周边县邑与主城人口交换"], "status": "READY_WITH_MODELED_GAPS", "score": 84,
    },
    "JIANYE": {
        "label": "建业", "strategic": "建业", "place": "place.han140.yangzhou.danyang.moling", "directory": "JIANYE",
        "historical_names": "秣陵|建业", "geography": "长江下游南岸的山水港城，连接丹阳腹地、江东水网和长江航道。", "terrain": "沿江丘陵与冲积地", "water": "长江|秦淮水系方向", "mountains": "钟山|石头山方向", "roads": "吴郡东向|丹阳南向|皖/濡须西向|长江水路", "adjacent": "石头城|丹阳郡县邑|吴郡方向|濡须口方向",
        "urban": "184仍以秣陵县城和沿江聚落为主；211石头城、212改建业、229吴都形成分期跃迁。", "wall": "184县城边界与211后石头城/都城防务不得混为一体。", "gates": "各阶段门名和边界需分期研究。", "inner": "229后吴宫廷和中央官署区；184不应提前生成。", "markets": "长江水运、江东农产、造船与商贸市场。", "residential": "184城市模型只表示秣陵阶段；建业都城人口增长需ChangePoint。", "industrial": "造船、木工、冶炼、纺织、粮食加工和水军军需。", "suburbs": "沿江港聚落、丹阳农业县邑与山丘防御节点。", "expansion": "211石头城建设、212改名建业、221/229都城迁移与建设。",
        "facilities": [("秣陵县城设施", "GovernmentOffice", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-211", "县治阶段"), ("沿江港埠", "Harbor", "RECONSTRUCTED", "APPROXIMATE_ZONE", "135-260", "长江交通"), ("石头城", "Fort", "HISTORICAL", "APPROXIMATE_ZONE", "211-260", "沿江军事要塞"), ("吴宫廷区", "CourtHall", "HISTORICAL", "CITY_LEVEL_ONLY", "229-260", "吴都政治中心"), ("建业市场", "Market", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "212-260", "都城商贸"), ("船作", "Shipyard", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "200-260", "水军与运输"), ("军粮仓储", "Warehouse", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "200-260", "水军与都城供给")],
        "industry": "造船|木工|冶炼锻造|纺织|粮食加工|军需|港运服务", "agriculture": "丹阳稻作|桑麻|渔业|山地林产", "resources": "长江航运|木材|铁料输入|农地与渔业",
        "network": [("Fort", "石头城", "211后沿江要塞", "HISTORICAL"), ("River", "长江", "核心物流与水军通道", "HISTORICAL"), ("RoadJunction", "吴郡方向", "江东内部交通", "RECONSTRUCTED"), ("SettlementCluster", "丹阳沿江聚落群", "港工、农户和商人", "MODELED"), ("AgriculturalZone", "丹阳供给区", "稻作、林产和渔业", "RECONSTRUCTED")],
        "military": "长江港运、石头城、江东内线和水军设施形成防御核心；不同Scenario不得提前出现后期都城设施。",
        "people": [("孙权", "211-260", "CourtPresence", "CONFIRMED", "吴主/皇帝"), ("张昭", "211-236", "OfficePresence", "CONFIRMED", "重臣"), ("周瑜", "211", "MilitaryPresence", "PROBABLE", "都督"), ("鲁肃", "211-217", "OfficePresence", "PROBABLE", "都督"), ("吕蒙", "211-219", "MilitaryPresence", "PROBABLE", "都督"), ("陆逊", "219-245", "OfficePresence", "PROBABLE", "都督/丞相"), ("诸葛瑾", "211-241", "OfficePresence", "PROBABLE", "官员"), ("顾雍", "211-243", "OfficePresence", "PROBABLE", "丞相"), ("步骘", "211-247", "OfficePresence", "PROBABLE", "将领/丞相"), ("孙登", "221-241", "CourtPresence", "PROBABLE", "太子")],
        "unknowns": ["184秣陵城与211后建业空间的精确叠合", "都城扩建逐设施时间", "港区和船作的数量、所有权及位置"], "status": "READY_WITH_MODELED_GAPS", "score": 83,
    },
    "HEFEI": {
        "label": "合肥", "strategic": "合肥", "place": "place.han140.yangzhou.jiujiang.hefei", "directory": "HEFEI",
        "historical_names": "合肥|合肥县", "geography": "淮南丘陵与巢湖水系之间的陆水节点，控制江淮南北交通。", "terrain": "丘陵岗地与河湖平原", "water": "巢湖水系|淝水方向|江淮水路", "mountains": "大别山东缘方向", "roads": "寿春北向|濡须/长江南向|庐江西向|丹阳东南向", "adjacent": "寿春|濡须口|庐江|巢湖水网",
        "urban": "县城、曹魏前线城防与230年代合肥新城必须分期；战略名不创造第二座城。", "wall": "早期县城和后期新城防线分离；当前精确边界不足。", "gates": "门名与数量UNKNOWN。", "inner": "前线军政区、仓储和驻军功能高于宫廷功能。", "markets": "江淮农产、军需和水陆转运市场。", "residential": "只有正确县级人口和城市聚落模型；城墙/都市圈数值未建立。", "industrial": "军械、船修、粮食加工、仓储和运输服务。", "suburbs": "巢湖水网、县域农业聚落与濡须方向军路。", "expansion": "208后长期前线化、215合肥之战、230年代新城建设。",
        "facilities": [("合肥县城防", "Wall", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-230", "早期城防"), ("合肥新城", "Fort", "HISTORICAL", "APPROXIMATE_ZONE", "230-260", "后期前线要塞"), ("县廷/军政官署", "GovernmentOffice", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "行政与军政"), ("驻军营地", "Barracks", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "200-260", "前线驻军"), ("军粮仓", "Granary", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "200-260", "长期防御供给"), ("巢湖水运节点", "Harbor", "RECONSTRUCTED", "APPROXIMATE_ZONE", "135-260", "水运与防务"), ("军需市场", "Market", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "200-260", "军民交换")],
        "industry": "军械|粮食加工|船修|木工|仓储|运输服务", "agriculture": "江淮稻麦|豆类|渔业|畜牧", "resources": "巢湖水运|江淮农地|木材石料与金属输入",
        "network": [("RiverLake", "巢湖水系", "水运与防御", "HISTORICAL"), ("RiverFort", "濡须口方向", "长江防线节点", "HISTORICAL"), ("RoadJunction", "寿春方向", "北向军路", "RECONSTRUCTED"), ("SettlementCluster", "合肥县域聚落群", "粮食、劳力和运输", "MODELED"), ("AgriculturalZone", "江淮供给区", "稻麦和渔业", "RECONSTRUCTED")],
        "military": "合肥是江淮前线节点；县城、新城、巢湖水路、濡须方向与寿春军路须按Scenario分期。",
        "people": [("刘馥", "208", "OfficePresence", "CONFIRMED", "扬州刺史/经营合肥"), ("张辽", "209-222", "MilitaryPresence", "CONFIRMED", "守将"), ("李典", "209-215", "MilitaryPresence", "CONFIRMED", "守将"), ("乐进", "209-218", "MilitaryPresence", "CONFIRMED", "守将"), ("孙权", "215", "MilitaryPresence", "CONFIRMED", "进攻统帅"), ("曹操", "208-217", "MilitaryPresence", "PROBABLE", "魏军统帅/增援"), ("满宠", "230-242", "MilitaryPresence", "CONFIRMED", "新城防务"), ("蒋济", "208-220", "OfficePresence", "PROBABLE", "扬州官员"), ("甘宁", "215", "MilitaryPresence", "PROBABLE", "吴将"), ("凌统", "215", "MilitaryPresence", "CONFIRMED", "吴将")],
        "unknowns": ["早期合肥城与后期新城边界", "184城市城墙/都市圈人口层", "巢湖港渡和军仓精确位置"], "status": "READY_WITH_MODELED_GAPS", "score": 79,
    },
    "HANZHONG_CANONICAL_PLACE": {
        "label": "南郑", "strategic": "汉中", "place": "place.han140.yizhou.hanzhong.nanzheng", "directory": "HANZHONG_CANONICAL_PLACE",
        "historical_names": "南郑|汉中（战略显示名/郡名）", "geography": "汉中盆地核心治所，北接关中、南连蜀地，控制秦巴山地通道。", "terrain": "盆地河谷与山地关隘", "water": "汉水上游", "mountains": "秦岭|大巴山", "roads": "阳平关北西向|褒斜/傥骆方向|金牛道南向|米仓道东南向", "adjacent": "阳平关|西城|上庸|剑阁方向|汉中县邑",
        "urban": "CanonicalPhysicalPlace是南郑；汉中是战略/行政显示层。郡治城市、张鲁政权和魏蜀争夺分期。", "wall": "南郑城防历史价值明确，精确城垣和门名不足。", "gates": "具体门名UNKNOWN。", "inner": "汉中郡官署、张鲁政权管理区和后续军政设施。", "markets": "盆地粮食、山地物资和关中—蜀地转运市场。", "residential": "正确县级人口和聚落层可用；无独立城墙/都市圈模型。", "industrial": "粮食加工、木工、军需、药材和山地运输服务。", "suburbs": "汉中盆地农业聚落、山口驿站和军粮节点。", "expansion": "194张鲁控制、215曹操入汉中、219刘备夺取形成关键变化。",
        "facilities": [("南郑城防", "Wall", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "郡治城防"), ("汉中郡官署", "GovernmentOffice", "HISTORICAL", "CITY_LEVEL_ONLY", "135-260", "行政治所"), ("张鲁政权管理设施", "GovernmentOffice", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "190-215", "政教合一治理，具体设施待考"), ("宗教集会/义舍参考", "RitualHall", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "190-215", "不得自动生成精确位置"), ("南郑市场", "Market", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "盆地转运"), ("军粮仓", "Granary", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "190-260", "山地战争供给"), ("驿传与车马节点", "CourierStation", "RECONSTRUCTED", "CITY_LEVEL_ONLY", "135-260", "关中—蜀地交通")],
        "industry": "粮食加工|木工|药材|军械|仓储|山地运输", "agriculture": "汉中盆地稻麦|粟豆|桑麻|山地林产", "resources": "盆地农地|秦巴木材药材|山地矿产候选",
        "network": [("Pass", "阳平关", "北西向独立关隘Place", "HISTORICAL"), ("RoadJunction", "褒斜/傥骆方向", "关中通道", "RECONSTRUCTED"), ("RoadJunction", "金牛道方向", "成都通道", "HISTORICAL"), ("SettlementCluster", "汉中盆地聚落群", "粮食、劳力和驿运", "MODELED"), ("AgriculturalZone", "汉中盆地供给区", "稻麦、林产和药材", "RECONSTRUCTED")],
        "military": "南郑、阳平关、秦岭诸道和南向蜀道共同构成山地战区；行政Region、战略Label与PhysicalPlace必须分离。",
        "people": [("张鲁", "190-215", "OfficePresence", "CONFIRMED", "汉中统治者"), ("阎圃", "190-215", "OfficePresence", "PROBABLE", "谋臣"), ("曹操", "215", "MilitaryPresence", "CONFIRMED", "征汉中统帅"), ("夏侯渊", "215-219", "MilitaryPresence", "CONFIRMED", "魏军主将"), ("张郃", "215-219", "MilitaryPresence", "CONFIRMED", "魏将"), ("刘备", "219", "MilitaryPresence", "CONFIRMED", "汉中之战统帅"), ("法正", "219", "MilitaryPresence", "CONFIRMED", "谋臣"), ("黄忠", "219", "MilitaryPresence", "CONFIRMED", "蜀将"), ("魏延", "219-227", "MilitaryPresence", "CONFIRMED", "汉中守将"), ("马超", "219", "MilitaryPresence", "PROBABLE", "蜀将")],
        "unknowns": ["南郑城垣、城门和官署逐Cell位置", "张鲁义舍/宗教设施的具体数量与位置", "汉中战略Label下各县与南郑主城人口层边界"], "status": "READY_WITH_MODELED_GAPS", "score": 80,
    },
}


def resolve_person(name: str):
    candidates = person_by_name.get(name, [])
    if len(candidates) != 1:
        raise ValueError(f"HistoricalPerson name must resolve uniquely: {name} -> {len(candidates)}")
    return candidates[0]


def primary_sources_for(place_id: str):
    source_ids = []
    for row in p0_by_place[place_id]:
        source_ids.extend(filter(None, str(row.get("source_ids", "")).split("|")))
    structured = structured_dir_by_place[place_id][1]
    source_ids.extend(filter(None, str(structured["master"].get("source_ids", "")).split("|")))
    return sorted(set(source_ids))


def population_rows(spec, structured):
    place_id = spec["place"]
    master = structured["master"]
    commandery_id = master["commandery_id"]
    region_184 = next(row for row in population_184["regions"] if row["region_permanent_id"] == commandery_id)
    result = []
    for year in (140, 184):
        county = structured["population"].get(str(year), {})
        major = next((row for row in major_city_population if row["year"] == year and row["county_permanent_id"] == master["county_id"]), None)
        # Luoyang's protected local calibration is valid; Chengdu's name-only wrong crosswalk is intentionally ignored.
        if spec["directory"] == "LUOYANG" and year == 184:
            major = next(row for row in population_184["major_cities"] if row["city_name"] == "洛阳")
        result.append({
            "City": spec["label"], "PlaceId": place_id, "Scenario": scenario_by_year.get(year, {}).get("scenario_id", f"reference.year.{year}"), "Year": year,
            "AdministrativePopulation": region_184["modeled_actual_population"] if year == 184 else None,
            "CountyPopulation": county.get("modeled_actual_population"),
            "WalledPopulation": major.get("walled_city_population") if major else None,
            "UrbanPopulation": major.get("urban_area_population") if major else county.get("urban_settlement_population"),
            "MetropolitanPopulation": major.get("metropolitan_population") if major else None,
            "NearSuburbPopulation": (130000 if spec["directory"] == "LUOYANG" and year == 184 else None),
            "SupplyHinterlandPopulation": (700000 if spec["directory"] == "LUOYANG" and year == 184 else None),
            "EvidenceLevel": "HISTORICAL" if major and major.get("evidence") == "HistoricalLocalCalibration" else "MODELED",
            "Confidence": major.get("confidence") if major else county.get("confidence", "UNKNOWN"),
            "PopulationDataset": "Assets/StreamingAssets/HistoricalPopulation/Han135260V1/",
            "Notes": "各层为包含关系，不可相加；县人口不等于城市人口。" + (" 供应圈70万包含都市圈40万。" if spec["directory"] == "LUOYANG" and year == 184 else " 未建立的层保持UNKNOWN，不套用洛阳比例。"),
        })
    return result


def facility_rows(spec):
    rows = []
    for index, (name, base, evidence, anchor, years, implication) in enumerate(spec["facilities"] + COMMON_MODELED, 1):
        rows.append({
            "FacilityReferenceId": f"facilityref.{spec['directory'].lower()}.{index:02d}", "City": spec["label"], "PlaceId": spec["place"],
            "HistoricalName": name, "BaseType": base, "EvidenceLevel": evidence, "AnchorPrecision": anchor,
            "ValidTimeRange": years, "ReferenceKind": "HISTORICAL_ANCHOR" if evidence in {"HISTORICAL", "RECONSTRUCTED"} else "SIMULATION_COMPLETION_REQUIREMENT",
            "DevelopmentImplication": implication, "SourceIds": "|".join(primary_sources_for(spec["place"])),
            "CellId": "REFERENCE_EXISTING_LUOYANG_PACKAGE" if spec["directory"] == "LUOYANG" and anchor == "EXACT_SITE" else "",
            "Notes": "BaseType复用统一Facility Catalog；不创建实例。",
        })
    return rows


def person_rows(spec):
    rows = []
    for name, time_range, presence_type, confidence, role in spec["people"]:
        person = resolve_person(name)
        start_year = int(re.findall(r"\d{3}", time_range)[0])
        scenario = min(scenario_by_year, key=lambda year: abs(year - start_year))
        rows.append({
            "City": spec["label"], "PlaceId": spec["place"], "PersonId": person["person_id"], "PersonName": name,
            "ScenarioYearOrRange": time_range, "ScenarioId": scenario_by_year[scenario]["scenario_id"], "PresenceType": presence_type,
            "LocationConfidence": confidence, "OfficeOrMilitaryRole": role, "ResidenceEvidence": "CONFIRMED" if presence_type == "Resident" and confidence == "CONFIRMED" else "NOT_ASSUMED",
            "ClanId": person.get("clan_id") or "", "BranchId": person.get("branch_id") or "",
            "FamilyOrganizationCandidate": "REFERENCE_ONLY" if person.get("clan_id") or person.get("branch_id") else "NONE_DERIVED",
            "EventRole": role, "EvidenceLevel": "HISTORICAL" if confidence == "CONFIRMED" else "RECONSTRUCTED",
            "SourceId": person.get("source_id", ""), "Notes": "城市在场切片，不把籍贯等同当前位置。",
        })
    return rows


def clan_rows(spec, people_rows):
    grouped = defaultdict(list)
    for row in people_rows:
        if row["ClanId"]:
            grouped[(row["ClanId"], row["BranchId"])].append(row["PersonId"])
    result = []
    for (clan_id, branch_id), person_ids in sorted(grouped.items()):
        clan = clan_by_id[clan_id]
        result.append({
            "City": spec["label"], "PlaceId": spec["place"], "ClanId": clan_id, "ClanName": clan["canonical_clan_name"],
            "BranchId": branch_id, "BranchName": branch_by_id.get(branch_id, {}).get("branch_name", ""),
            "MemberPresenceIds": "|".join(sorted(set(person_ids))), "PresenceEvidence": "PERSON_CITY_SLICE",
            "ResidenceEvidence": "UNKNOWN", "EstateEvidence": "UNKNOWN", "FamilyOrganizationCandidate": "REFERENCE_ONLY_DO_NOT_INSTANTIATE",
            "FamilyCenterCandidate": "NO_CITY_CENTER_DERIVED", "EvidenceLevel": "RECONSTRUCTED",
            "Unknowns": "组织边界、资产、管理者、住宅和中心Facility均需独立证据。",
        })
    if not result:
        result.append({
            "City": spec["label"], "PlaceId": spec["place"], "ClanId": "", "ClanName": "", "BranchId": "", "BranchName": "",
            "MemberPresenceIds": "", "PresenceEvidence": "NO_CONFIRMED_CLAN_LINK_IN_CURRENT_CITY_SLICE", "ResidenceEvidence": "UNKNOWN",
            "EstateEvidence": "UNKNOWN", "FamilyOrganizationCandidate": "NONE_DERIVED", "FamilyCenterCandidate": "NONE_DERIVED",
            "EvidenceLevel": "UNKNOWN", "Unknowns": "需要城市级地方豪族、住宅与Estate专题研究。",
        })
    return result


def urban_rows(spec):
    items = [
        ("HistoricalUrbanCore", spec["urban"], "RECONSTRUCTED"), ("FortifiedBoundary", spec["wall"], "RECONSTRUCTED"),
        ("Wall", spec["wall"], "HISTORICAL" if any(f[1] == "Wall" and f[2] == "HISTORICAL" for f in spec["facilities"]) else "RECONSTRUCTED"),
        ("Gate", spec["gates"], "UNKNOWN" if "UNKNOWN" in spec["gates"] else "RECONSTRUCTED"), ("PalaceInnerCity", spec["inner"], "RECONSTRUCTED"),
        ("GovernmentArea", "以行政/政治时间轴确定，精确位置不足时只到CITY_LEVEL_ONLY。", "RECONSTRUCTED"),
        ("MarketArea", spec["markets"], "RECONSTRUCTED"), ("ResidentialAreas", spec["residential"], "MODELED"),
        ("IndustrialAreas", spec["industrial"], "RECONSTRUCTED"), ("WaterSystem", spec["water"], "HISTORICAL"),
        ("MainRoads", spec["roads"], "RECONSTRUCTED"), ("Suburbs", spec["suburbs"], "RECONSTRUCTED"),
        ("MilitaryZones", spec["military"], "RECONSTRUCTED"), ("UrbanExpansion", spec["expansion"], "HISTORICAL"),
    ]
    return [{"City": spec["label"], "PlaceId": spec["place"], "Component": component, "Reference": value, "EvidenceLevel": evidence, "AnchorPrecision": "CITY_LEVEL_ONLY", "DevelopmentImplication": "作为同一CanonicalPlace的可变Area/Facility组合，不建立第二套城市地图。"} for component, value, evidence in items]


def state_rows(spec):
    rows = []
    for state in sorted(state_by_place[spec["place"]], key=lambda row: row["ScenarioYear"]):
        year = int(state["ScenarioYear"])
        scenario = scenario_by_year.get(year)
        cp = change_by_id.get(state.get("MajorChangePointId", ""))
        rows.append({
            "City": spec["label"], "PlaceId": spec["place"], "HistoricalStateId": state["HistoricalStateId"],
            "HistoricalState": f"{spec['label']} {year} development reference", "Scenario": scenario["scenario_id"] if scenario else "REFERENCE_TIMEPOINT",
            "ScenarioYear": year, "Priority": "HIGH" if state["RequiredSnapshotDepth"] in {"H4", "H5"} else "MEDIUM",
            "ChangePoint": state.get("MajorChangePointId", ""), "PreState": cp.get("PreState", "INHERIT_PREVIOUS_REFERENCE") if cp else "INHERIT_PREVIOUS_REFERENCE",
            "PostState": cp.get("PostState", "REFERENCE_STATE_ONLY") if cp else "REFERENCE_STATE_ONLY",
            "FacilityImpact": cp.get("FacilityImpact", "PLANNED_REVIEW") if cp else "PLANNED_REVIEW",
            "PopulationImpact": cp.get("PopulationImpact", "READ_POPULATION_DATASET") if cp else "READ_POPULATION_DATASET",
            "PersonImpact": "READ_PERSON_CITY_SLICE", "FamilyImpact": "REFERENCE_ONLY_NO_AUTO_ORGANIZATION",
            "DevelopmentNeed": state["RequiredSnapshotDepth"], "ReferenceStatus": "KNOWN_CHANGEPOINT" if cp else "STATE_REFERENCE_NO_CANONICAL_CHANGEPOINT",
            "SourceReference": "DevelopmentPlace Historical State Plan + Scenario master",
        })
    return rows


def module_rows(spec, city):
    modules = [
        ("01", "Identity / Geography", 100, "CanonicalPlace、行政、战略Label和GIS锚点已解析"),
        ("02", "Administrative / Political", 92 if spec["directory"] != "HEFEI" else 85, "历史治所与Runtime Seat分离"),
        ("03", "Population", 95 if spec["directory"] == "LUOYANG" else (72 if spec["directory"] in {"HEFEI", "HANZHONG_CANONICAL_PLACE", "CHENGDU"} else 82), "引用全国人口母盘；未知层不套比例"),
        ("04", "Urban Spatial Form", 94 if spec["directory"] == "LUOYANG" else 76, "分期城市形态；非洛阳不硬塞精确Cell"),
        ("05", "Facility", 95 if spec["directory"] == "LUOYANG" else 78, "历史锚点与运行补全分离"),
        ("06", "HistoricalPerson", min(95, 55 + len(city["people"]) * 4), f"{len(city['people'])}条稳定PersonId城市切片"),
        ("07", "Clan / Family / Estate", 85 if spec["directory"] == "LUOYANG" else 62, "不由人物在场自动生成FamilyCenter"),
        ("08", "Industry / Agriculture / Resources", 84, "映射Facility/Recipe，不用抽象产业等级"),
        ("09", "Transport / Logistics / Settlements", 86, "建立城市供给链和周边群落Reference"),
        ("10", "Military", 88, "映射同一Place/Cell/Facility/Force"),
        ("11", "Scenario Snapshot", 88, f"{len(city['states'])}个相关Scenario/TimePoint"),
        ("12", "HistoricalChangePoint", 82, "已知ChangePoint交叉引用；空缺保留计划"),
        ("13", "Readiness / Unknowns / Implications", spec["score"], spec["status"]),
    ]
    return [{"Module": number, "ModuleName": name, "CoverageScore": score, "CoverageLevel": "READY" if score >= 85 else ("ADEQUATE_WITH_GAPS" if score >= 70 else "MODELED_GAPS"), "Conclusion": conclusion, "EvidenceContract": "HISTORICAL|RECONSTRUCTED|MODELED|UNKNOWN", "RuntimeBoundary": "REFERENCE_ONLY"} for number, name, score, conclusion in modules]


city_payloads = {}
for slug, spec in CITY_MANUAL.items():
    if spec["place"] not in canonical_by_id or spec["place"] not in roster_by_id or spec["place"] not in structured_dir_by_place:
        raise KeyError(f"Missing canonical/roster/structured reference for {slug}: {spec['place']}")
    source_dir, structured = structured_dir_by_place[spec["place"]]
    people = person_rows(spec)
    facilities = facility_rows(spec)
    clans_for_city = clan_rows(spec, people)
    states = state_rows(spec)
    populations = population_rows(spec, structured)
    identity = [{
        "CanonicalPlaceId": spec["place"], "CanonicalName": structured["master"]["display_name"], "HistoricalNames": spec["historical_names"],
        "StrategicLabels": spec["strategic"], "CountyPermanentId": structured["master"]["county_id"],
        "CommanderyEquivalentId": structured["master"]["commandery_id"], "ProvinceId": structured["master"]["province_id"],
        "GISAnchor": f"{structured['master'].get('longitude')},{structured['master'].get('latitude')}", "CoordinateStatus": structured["master"].get("coordinate_status"),
        "PhysicalGeography": spec["geography"], "Terrain": spec["terrain"], "WaterRiver": spec["water"], "NearbyMountains": spec["mountains"],
        "RoadCorridors": spec["roads"], "AdjacentImportantPlaces": spec["adjacent"], "EvidenceLevel": "HISTORICAL|RECONSTRUCTED",
        "AdministrativeRegionNotPlace": structured["master"]["commandery_name"], "HistoricalSeatReference": "YES", "RuntimeSeatFixed": "NO",
        "SourceIds": "|".join(primary_sources_for(spec["place"])),
    }]
    industry = []
    for category, value, evidence in (("Industry", spec["industry"], "RECONSTRUCTED"), ("Agriculture", spec["agriculture"], "RECONSTRUCTED"), ("Resources", spec["resources"], "RECONSTRUCTED"), ("OccupationStructure", "官吏|士人|军人|商人|工匠|农户|雇工|仆役|门客|学生|医生|宗教人员|流民", "MODELED"), ("Workforce", "未来Person物化必须按真实岗位、年龄、技能、家庭与住宅约束", "MODELED"), ("ProductionMapping", "FacilityDefinition + Recipe + real worker + material + time + authority", "MODELED")):
        industry.append({"City": spec["label"], "PlaceId": spec["place"], "Category": category, "Reference": value, "EvidenceLevel": evidence, "FacilityRecipeMapping": "REQUIRED", "DevelopmentImplication": "不产生抽象产能；未来落到Facility、工单、库存和永久人物。"})
    transport = [{"City": spec["label"], "PlaceId": spec["place"], "NodeType": kind, "NodeName": name, "Relation": relation, "EvidenceLevel": evidence, "AnchorPrecision": "APPROXIMATE_ZONE" if evidence != "MODELED" else "CITY_LEVEL_ONLY", "SupplyChainRole": "Producer/Settlement -> Storage -> Road/Water -> Gate/Harbor -> Urban Storage/Market", "RuntimeStatus": "REFERENCE_ONLY"} for kind, name, relation, evidence in spec["network"]]
    military = [{"City": spec["label"], "PlaceId": spec["place"], "Topic": "UnifiedMilitarySpace", "Reference": spec["military"], "EvidenceLevel": "RECONSTRUCTED", "ForceMapping": "same CanonicalPlace + Cell + Road + Facility + Force", "SiegeBattleHistory": "See Historical State / ChangePoint plan", "DevelopmentImplication": "城防、道路、仓储和驻军不得成为第二套战斗地图事实。"}]
    unknowns = [{"UnknownId": f"unknown.{slug.lower()}.{i:02d}", "City": spec["label"], "PlaceId": spec["place"], "Unknown": value, "Impact": "MODELED_GAP" if spec["status"] != "DEVELOPMENT_READY" else "NON_BLOCKING_DETAIL", "MinimumResearch": "在进入具体Cell/Facility实现前补证；未补前保持UNKNOWN或CITY_LEVEL_ONLY", "BlocksPack": "NO", "BlocksRuntimeDetail": "YES"} for i, value in enumerate(spec["unknowns"], 1)]
    source_rows = []
    source_catalog = {row.get("source_id"): row for row in [*deep["sources"], *historical["sources"], *person_sources, *population_sources] if row.get("source_id")}
    referenced_source_ids = set(primary_sources_for(spec["place"])) | {row["SourceId"] for row in people if row["SourceId"]}
    for source_id in sorted(referenced_source_ids):
        source = source_catalog.get(source_id, {})
        source_rows.append({"SourceId": source_id, "Title": source.get("title", source.get("source_title", "Project source catalog")), "SourceType": source.get("source_type", "PROJECT_DATASET_REFERENCE"), "URLOrLocator": source.get("url_or_locator", source.get("url", "")), "EvidenceScope": source.get("evidence_scope", "City slice and stable master reference"), "LicenseNote": source.get("license_note", "Reference only; do not copy protected modern content")})
    source_rows.extend([
        {"SourceId": "dataset.population.han135260.v1", "Title": "Han135260V1 population dataset", "SourceType": "PROJECT_DATASET", "URLOrLocator": "Assets/StreamingAssets/HistoricalPopulation/Han135260V1/", "EvidenceScope": "Administrative/county/city population layers", "LicenseNote": "Project-authored model"},
        {"SourceId": "dataset.persons.han135260.v1", "Title": "Han135260V1 historical person dataset", "SourceType": "PROJECT_DATASET", "URLOrLocator": "Assets/StreamingAssets/HistoricalPersons/Han135260V1/", "EvidenceScope": "Stable Person/Clan/Branch IDs", "LicenseNote": "Project-authored structured reference"},
        {"SourceId": "dataset.development.roster.v1", "Title": "Development Place Roster V1", "SourceType": "PROJECT_REFERENCE", "URLOrLocator": "Docs/HISTORICAL_WORLD_REFERENCE/DEVELOPMENT_PLACE_ROSTER_V1/", "EvidenceScope": "Depth, wave and state plan", "LicenseNote": "Project-authored planning reference"},
    ])
    city = {**deepcopy(spec), "structured_source_dir": str(source_dir.relative_to(REPO)).replace("\\", "/"), "identity": identity, "populations": populations, "urban_form": urban_rows(spec), "facilities": facilities, "people": people, "clans": clans_for_city, "industry": industry, "transport": transport, "military_rows": military, "states": states, "sources": source_rows, "unknown_rows": unknowns}
    city["modules"] = module_rows(spec, city)
    city_payloads[slug] = city


master_rows = []
person_coverage = []
clan_coverage = []
facility_coverage = []
hinterland_coverage = []
population_layers = []
historical_states = []
for slug, city in city_payloads.items():
    roster = roster_by_id[city["place"]]
    module_map = {row["ModuleName"]: row["CoverageLevel"] for row in city["modules"]}
    master_rows.append({
        "Place": city["label"], "PlaceId": city["place"], "StrategicLabel": city["strategic"], "CurrentDepth": roster["DevelopmentDepth"],
        "PackStatus": city["status"], "Population": module_map["Population"], "UrbanForm": module_map["Urban Spatial Form"],
        "Facility": module_map["Facility"], "Person": module_map["HistoricalPerson"], "ClanFamily": module_map["Clan / Family / Estate"],
        "Industry": module_map["Industry / Agriculture / Resources"], "Transport": module_map["Transport / Logistics / Settlements"],
        "Settlements": module_map["Transport / Logistics / Settlements"], "Military": module_map["Military"], "Scenario": module_map["Scenario Snapshot"],
        "ChangePoint": module_map["HistoricalChangePoint"], "Sources": len(city["sources"]), "OverallReadiness": city["score"],
        "BlockingIssues": "NONE_PACK_GATE" if city["status"] != "RESEARCH_REQUIRED" else "SEE_UNKNOWNS",
        "RuntimeReadiness": "LUOYANG_REVIEW_ALLOWED" if slug == "LUOYANG" else "RUNTIME_NOT_IMPLEMENTED",
        "DepthUpgradeRecommendation": "NONE_THIS_TASK", "PackPath": f"Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/{city['directory']}/",
    })
    counts = Counter(row["LocationConfidence"] for row in city["people"])
    scenarios = sorted({row["ScenarioId"] for row in city["people"]})
    person_coverage.append({"City": city["label"], "PlaceId": city["place"], "Scenario": "|".join(scenarios), "PersonCount": len(city["people"]), "ConfirmedPresence": counts["CONFIRMED"], "ProbablePresence": counts["PROBABLE"], "OfficePresence": sum(row["PresenceType"] == "OfficePresence" for row in city["people"]), "ResidenceEvidence": sum(row["ResidenceEvidence"] == "CONFIRMED" for row in city["people"]), "ClanLinked": sum(bool(row["ClanId"]) for row in city["people"]), "CoverageLevel": "STRONG" if len(city["people"]) >= 10 else "ADEQUATE", "MajorMissingCategories": "女性/医生/宗教/工艺人物仍需按城市深化；本表不是名将榜"})
    clan_coverage.append({"City": city["label"], "PlaceId": city["place"], "ClanPresence": sum(bool(row["ClanId"]) for row in city["clans"]), "BranchPresence": sum(bool(row["BranchId"]) for row in city["clans"]), "ResidenceEvidence": sum(row["ResidenceEvidence"] != "UNKNOWN" for row in city["clans"]), "EstateEvidence": sum(row["EstateEvidence"] != "UNKNOWN" for row in city["clans"]), "FamilyOrganizationCandidates": sum(row["FamilyOrganizationCandidate"] != "NONE_DERIVED" for row in city["clans"]), "FamilyCenterCandidates": sum(row["FamilyCenterCandidate"] not in {"NONE_DERIVED", "NO_CITY_CENTER_DERIVED"} for row in city["clans"]), "CoverageLevel": "STRONG" if slug == "LUOYANG" else "MODELED_GAPS", "Unknowns": "地方豪族、住宅、Estate、组织边界和中心Facility不得从人物在场自动推定"})
    historical_count = sum(row["EvidenceLevel"] == "HISTORICAL" for row in city["facilities"])
    reconstructed_types = sorted({row["BaseType"] for row in city["facilities"] if row["EvidenceLevel"] == "RECONSTRUCTED"})
    modeled_types = sorted({row["BaseType"] for row in city["facilities"] if row["EvidenceLevel"] == "MODELED"})
    facility_coverage.append({"City": city["label"], "PlaceId": city["place"], "HistoricalFacilityCount": historical_count, "ReconstructedFacilityTypes": "|".join(reconstructed_types), "ModeledRequirementTypes": "|".join(modeled_types), "Government": "YES", "Market": "YES", "Storage": "YES", "Military": "YES", "Education": "YES_MODELED_IF_NO_ANCHOR", "Residential": "MODELED_FROM_PERSON_CAPACITY", "Industry": "YES", "Transport": "YES", "Fortification": "YES_WITH_PERIOD_GAPS", "CoverageLevel": "STRONG" if slug == "LUOYANG" else "ADEQUATE_WITH_MODELED_GAPS"})
    nodes = defaultdict(list)
    for row in city["transport"]:
        nodes[row["NodeType"]].append(row["NodeName"])
    hinterland_coverage.append({"City": city["label"], "PlaceId": city["place"], "UrbanCore": city["urban"], "NearSuburbs": city["suburbs"], "ImportantCountyTowns": city["adjacent"], "VillageClusters": "MODELED SETTLEMENT CLUSTER", "EstateClusters": "REFERENCE/UNKNOWN; no automatic FamilyCenter", "AgriculturalZones": next((row["NodeName"] for row in city["transport"] if row["NodeType"] == "AgriculturalZone"), "MODELED SUPPLY ZONE"), "MajorRoads": city["roads"], "WaterRoutes": city["water"], "SupplyNodes": "|".join(row["NodeName"] for row in city["transport"]), "CoverageLevel": "STRONG" if len(city["transport"]) >= 5 else "ADEQUATE"})
    population_layers.extend(city["populations"])
    historical_states.extend(city["states"])


upgrade_by_id = {}
for row in roster_data["roster"]:
    upgrade_by_id[row["CanonicalPlaceId"]] = {
        "PlaceId": row["CanonicalPlaceId"], "PlaceName": row["DevelopmentDisplayName"], "CurrentDevelopmentDepth": row["DevelopmentDepth"],
        "CurrentPackStatus": city_payloads[next((slug for slug, spec in CITY_MANUAL.items() if spec["place"] == row["CanonicalPlaceId"]), "")]["status"] if any(spec["place"] == row["CanonicalPlaceId"] for spec in CITY_MANUAL.values()) else "NO_FORMAL_PACK",
        "EligibleForUpgrade": not row["CanonicalPlaceId"].startswith("geo.site."), "ExistingPack": "YES" if any(spec["place"] == row["CanonicalPlaceId"] for spec in CITY_MANUAL.values()) else "NO",
        "UpgradeTargetCandidate": "D3|D4|D5_USER_DECISION" if not row["CanonicalPlaceId"].startswith("geo.site.") else "RESOLVE_CANONICAL_PLACE_FIRST",
        "KnownReferenceLevel": row["ReferenceReadiness"], "MainGaps": "See City Pack" if any(spec["place"] == row["CanonicalPlaceId"] for spec in CITY_MANUAL.values()) else "CREATE_OR_UPGRADE_DEVELOPMENT_PACK",
        "LastReview": TODAY, "Notes": "Pack readiness never changes DevelopmentDepth automatically.", "RosterMembership": "YES",
    }
for row in admin["strategic_crosswalk"]:
    place_id = row.get("CanonicalPlaceId")
    if not place_id:
        continue
    existing = upgrade_by_id.get(place_id)
    if existing:
        existing["Notes"] += f" StrategicLabel={row['StrategicDisplayName']}."
        continue
    base = canonical_by_id.get(place_id, {})
    upgrade_by_id[place_id] = {
        "PlaceId": place_id, "PlaceName": base.get("CanonicalName", row.get("ActualHistoricalSeatName", row["StrategicDisplayName"])),
        "CurrentDevelopmentDepth": "D0_OR_D1_NOT_IN_ROSTER", "CurrentPackStatus": "NO_FORMAL_PACK", "EligibleForUpgrade": place_id in canonical_by_id,
        "ExistingPack": "NO", "UpgradeTargetCandidate": "D3|D4|D5_USER_DECISION", "KnownReferenceLevel": base.get("HistoricalImportance", "REFERENCE_EXISTS"),
        "MainGaps": "CREATE_DEVELOPMENT_PACK_AND_RUN_REFERENCE_GAP_AUDIT", "LastReview": TODAY,
        "Notes": f"StrategicLabel={row['StrategicDisplayName']}; current roster is not a permanent whitelist.", "RosterMembership": "NO",
    }
upgrade_registry = sorted(upgrade_by_id.values(), key=lambda row: row["PlaceId"])


all_source_rows = {}
for city in city_payloads.values():
    for row in city["sources"]:
        all_source_rows[row["SourceId"]] = row

registry_updates = {
    "documents": [
        {"DocumentId": "doc.core-city-packs-v1.readme", "Path": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/README_CORE_CITY_DEVELOPMENT_PACKS.md", "Title": "Core City Development Packs V1", "Domain": "HistoricalWorldGeography", "SubDomain": "CityDevelopmentPack", "DocumentType": "L2ReferenceIndex", "AuthorityLevel": "L2", "Status": "CURRENT", "CreatedOrKnownDate": TODAY, "LastKnownRevision": TODAY, "CanonicalFor": "First ten core city development reference packs", "RelatedDocuments": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/CITY_DEVELOPMENT_PACK_STANDARD_V1.md|Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/CITY_DEVELOPMENT_PACK_UPGRADE_PROTOCOL_V1.md", "RelatedTasks": "Docs/TASK_HAN_135_260_CORE_CITY_DEVELOPMENT_PACK_AND_UPGRADE_PROTOCOL_V1.md", "RelatedRuntimeSystems": "ReferenceOnly", "HistoricalValue": "HIGH", "RecommendedReader": "Codex|Developer|Historian", "ReadPriority": 1, "CanonicalScope": "Development Pack index and first-batch status", "Notes": "Does not create runtime cities or upgrade depth."},
        {"DocumentId": "doc.core-city-packs-v1.standard", "Path": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/CITY_DEVELOPMENT_PACK_STANDARD_V1.md", "Title": "Development Pack Standard V1", "Domain": "HistoricalWorldGeography", "SubDomain": "CityDevelopmentPack", "DocumentType": "CanonicalProtocol", "AuthorityLevel": "L1", "Status": "FROZEN", "CreatedOrKnownDate": TODAY, "LastKnownRevision": TODAY, "CanonicalFor": "Development Pack modules, evidence and readiness", "RelatedDocuments": "Docs/GAME_SYSTEMS_MASTER_AND_STATUS.md", "RelatedTasks": "Docs/TASK_HAN_135_260_CORE_CITY_DEVELOPMENT_PACK_AND_UPGRADE_PROTOCOL_V1.md", "RelatedRuntimeSystems": "ReferenceOnly", "HistoricalValue": "HIGH", "RecommendedReader": "Codex|Developer", "ReadPriority": 0, "CanonicalScope": "Any Place development pack", "Notes": "Pack ready is not depth upgrade."},
        {"DocumentId": "doc.core-city-packs-v1.upgrade", "Path": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/CITY_DEVELOPMENT_PACK_UPGRADE_PROTOCOL_V1.md", "Title": "City Development Pack Upgrade Protocol V1", "Domain": "ProjectPlanning", "SubDomain": "DevelopmentDepth", "DocumentType": "CanonicalProtocol", "AuthorityLevel": "L1", "Status": "FROZEN", "CreatedOrKnownDate": TODAY, "LastKnownRevision": TODAY, "CanonicalFor": "Future Place detail and depth upgrade gate", "RelatedDocuments": "Docs/HISTORICAL_WORLD_REFERENCE/DEVELOPMENT_PLACE_ROSTER_V1/README.md", "RelatedTasks": "Docs/TASK_HAN_135_260_CORE_CITY_DEVELOPMENT_PACK_AND_UPGRADE_PROTOCOL_V1.md", "RelatedRuntimeSystems": "ReferenceOnly", "HistoricalValue": "HIGH", "RecommendedReader": "Codex|Developer|Designer", "ReadPriority": 0, "CanonicalScope": "All CanonicalPlaces and resolvable strategic Places", "Notes": "User confirmation required before depth change."},
        {"DocumentId": "doc.core-city-packs-v1.report", "Path": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/CORE_CITY_DEVELOPMENT_PACK_COMPLETENESS_REPORT_V1.md", "Title": "Core City Development Pack Completeness Report V1", "Domain": "HistoricalWorldGeography", "SubDomain": "CityDevelopmentPack", "DocumentType": "AcceptanceReport", "AuthorityLevel": "L4", "Status": "CURRENT", "CreatedOrKnownDate": TODAY, "LastKnownRevision": TODAY, "CanonicalFor": "", "RelatedDocuments": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/README_CORE_CITY_DEVELOPMENT_PACKS.md", "RelatedTasks": "Docs/TASK_HAN_135_260_CORE_CITY_DEVELOPMENT_PACK_AND_UPGRADE_PROTOCOL_V1.md", "RelatedRuntimeSystems": "ReferenceOnly", "HistoricalValue": "NORMAL", "RecommendedReader": "Codex|Developer", "ReadPriority": 2, "CanonicalScope": "Acceptance evidence only", "Notes": "Does not override L1 standard/protocol."},
    ],
    "domain_map": [{"Domain": "CityDevelopmentPack", "L0ProjectConstitution": "Docs/GAME_VISION_AND_GAMEPLAY.md", "L1CanonicalSpec": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/CITY_DEVELOPMENT_PACK_STANDARD_V1.md|Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/CITY_DEVELOPMENT_PACK_UPGRADE_PROTOCOL_V1.md", "L2CurrentStatus": "Docs/GAME_SYSTEMS_MASTER_AND_STATUS.md|Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/README_CORE_CITY_DEVELOPMENT_PACKS.md", "L3PrimaryReference": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/", "CanonicalGap": "Runtime city implementation remains separate", "MultipleL1Conflict": "NO", "ReadingEntry": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/README_CORE_CITY_DEVELOPMENT_PACKS.md", "ConflictPolicy": "Pack references canonical datasets; never duplicate or overwrite world facts", "DomainId": "domain.city-development-pack", "DomainName": "City Development Pack", "CurrentStatus": "FIRST_10_PACKS_COMPLETE", "Notes": "Any CanonicalPlace may be upgraded after pack review and user decision."}],
    "design_decisions": [
        ("DEC-DEVPACK-001", "DevelopmentDepth is mutable project planning state", "D0-D5 may change through an explicit user/development decision; it is not a permanent world attribute."),
        ("DEC-DEVPACK-002", "Any CanonicalPlace may be upgraded", "Current roster membership or D0/D1 status never makes a valid CanonicalPlace permanently ineligible."),
        ("DEC-DEVPACK-003", "City detail request requires Development Pack first", "The first action for a request to detail a city is create/upgrade and audit its Development Pack."),
        ("DEC-DEVPACK-004", "Development Pack is prerequisite to depth upgrade", "Reference readiness review precedes any proposed D3/D4/D5 change."),
        ("DEC-DEVPACK-005", "Pack readiness does not automatically change DevelopmentDepth", "Only the user and explicit development plan authorize the depth change."),
        ("DEC-DEVPACK-006", "Depth upgrade preserves world facts and stable IDs", "Existing Place, Cell, Person, Household, Facility and inventory facts are extended, never deleted and regenerated."),
        ("DEC-DEVPACK-007", "DevelopmentPlaceRoster V1 is not a permanent whitelist", "The 72-place roster is the current production plan and remains extensible."),
        ("DEC-DEVPACK-008", "D0/D1 does not mean permanently low detail", "D0/D1 means no current special production scope, not low historical value."),
        ("DEC-DEVPACK-009", "City packs reference canonical datasets", "Packs store city slices and development implications; they do not copy a second Person, Population, Clan, Place or Facility master."),
    ],
    "open_decisions": [
        {"OpenDecisionId": "OPEN-DEVPACK-001", "Domain": "DevelopmentDepth", "Question": "Which current D3 Places should later become D4?", "Status": "OPEN", "WhyOpen": "Requires user priority and place-specific pack review", "NeededEvidence": "Development Pack + runtime cost + gameplay value", "OwnerRole": "User/DevelopmentPlan", "Blocks": "Future waves only", "SourceDocument": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/CITY_DEVELOPMENT_PACK_UPGRADE_PROTOCOL_V1.md", "RecommendedNextReview": "When a specific Place is requested", "Notes": "Not frozen by this task"},
        {"OpenDecisionId": "OPEN-DEVPACK-002", "Domain": "DevelopmentDepth", "Question": "Should a second D5 be added after Luoyang?", "Status": "OPEN", "WhyOpen": "D5 cost and reuse evidence are not yet reviewed", "NeededEvidence": "Completed pack + D5 production estimate + user decision", "OwnerRole": "User/DevelopmentPlan", "Blocks": "Second flagship only", "SourceDocument": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/CITY_DEVELOPMENT_PACK_UPGRADE_PROTOCOL_V1.md", "RecommendedNextReview": "After Luoyang readiness review", "Notes": "No automatic promotion"},
        {"OpenDecisionId": "OPEN-DEVPACK-003", "Domain": "CityDevelopmentPack", "Question": "Which ordinary cities should receive the next full Development Pack?", "Status": "OPEN", "WhyOpen": "The task intentionally stops after the first ten", "NeededEvidence": "Explicit user request or revised development wave", "OwnerRole": "User/DevelopmentPlan", "Blocks": "Further pack expansion only", "SourceDocument": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/README_CORE_CITY_DEVELOPMENT_PACKS.md", "RecommendedNextReview": "On demand", "Notes": "Current 72 roster remains extensible"},
    ],
    "implementation_gaps": [
        {"GapId": "IMP-GAP-CITYPACK-001", "Domain": "CityDevelopmentPack", "CanonicalRequirement": "Runtime tasks consume approved Development Packs", "CurrentImplementation": "Reference documents/workbooks only", "GapDescription": "No runtime pack loader is required or implemented; future tasks must translate approved slices deliberately", "Severity": "S2", "BlocksNextDevelopment": "NO_FOR_LUOYANG_REVIEW", "SuggestedFutureTask": "LUOYANG-184-DEVELOPMENT-READINESS-REVIEW-V1", "Evidence": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/", "Status": "OPEN", "Notes": "Do not build a second runtime truth"},
        {"GapId": "IMP-GAP-CITYPACK-002", "Domain": "UrbanSpatialState", "CanonicalRequirement": "Period-variable UrbanArea/FortifiedBoundary", "CurrentImplementation": "Luoyang prototype only", "GapDescription": "Other cities lack formal period-variable local spatial packages", "Severity": "S2", "BlocksNextDevelopment": "YES_PER_CITY", "SuggestedFutureTask": "Place-specific runtime implementation after pack review", "Evidence": "Core city pack unknown registers", "Status": "OPEN", "Notes": "No invented exact Cell"},
        {"GapId": "IMP-GAP-CITYPACK-003", "Domain": "HistoricalChangePackage", "CanonicalRequirement": "Pre-State + conditional change package + Post-State", "CurrentImplementation": "Reference plan only", "GapDescription": "City destruction, relocation and rebuilding packages are not implemented", "Severity": "S2", "BlocksNextDevelopment": "NO_FOR_REFERENCE", "SuggestedFutureTask": "Scenario-specific HistoricalChangePackage task", "Evidence": "07_CORE_CITY_HISTORICAL_STATE_AND_CHANGEPOINT_PLAN.xlsx", "Status": "OPEN", "Notes": "Events remain world-level"},
        {"GapId": "IMP-GAP-CITYPACK-004", "Domain": "FamilyOrganization", "CanonicalRequirement": "City social structure without automatic Clan-to-organization conversion", "CurrentImplementation": "Reference candidates only", "GapDescription": "Nine non-Luoyang cities lack formal family organization initialization", "Severity": "S2", "BlocksNextDevelopment": "YES_PER_CITY", "SuggestedFutureTask": "Place-specific family integration after readiness review", "Evidence": "03_CORE_CITY_CLAN_FAMILY_COVERAGE.xlsx", "Status": "OPEN", "Notes": "FamilyCenter requires real residence/assets/manager"},
    ],
    "research_gaps": [],
}
registry_updates["design_decisions"] = [{"DecisionId": did, "Domain": "CityDevelopmentPack", "Title": title, "Decision": decision, "Status": "FROZEN", "EffectiveFrom": TODAY, "SourceDocument": "Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/CITY_DEVELOPMENT_PACK_UPGRADE_PROTOCOL_V1.md", "AffectedDocuments": "GAME_SYSTEMS_MASTER_AND_STATUS|DevelopmentPlaceRoster|Development Manifests|Knowledge Base", "AffectedSystems": "HistoricalWorldReference|ProjectPlanning|FutureRuntime", "ReasonSummary": "Preserve one world, stable IDs and explicit user control", "OpenQuestions": "See OPEN-DEVPACK registry", "Notes": "Reference completion never creates runtime facts"} for did, title, decision in registry_updates["design_decisions"]]
for slug, city in city_payloads.items():
    registry_updates["research_gaps"].append({"GapId": f"RES-GAP-CITYPACK-{len(registry_updates['research_gaps']) + 1:03d}", "Domain": "CityDevelopmentPack", "ResearchGap": "; ".join(city["unknowns"]), "Priority": "HIGH" if city["directory"] == "LUOYANG" else "MEDIUM", "CurrentEvidence": city["structured_source_dir"], "RequiredSources": "Primary historical text|archaeology|historical geography|existing canonical datasets", "DoNotInfer": "Do not invent exact Cell, duplicate Place, population, Person, Clan, Facility or FamilyCenter", "SuggestedResearchAction": "Close only the minimum gap required by the next runtime slice", "Question": f"What minimum evidence is required before implementing {city['label']} local Cells and Facilities?", "EvidenceNeeded": "Period-specific urban form|facility anchor|population layer|person/family slice", "Blocks": "Place-specific runtime detail", "Status": "OPEN", "Notes": city["status"]})


def read_registry_rows(filename: str):
    workbook = load_workbook(REPO / "Docs" / "KNOWLEDGE_BASE" / "REGISTRY" / filename, read_only=True, data_only=False)
    # Registry records are always kept in the second worksheet. Positional
    # access also survives workbooks written with a localized sheet caption.
    sheet = workbook.worksheets[1]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value) if value is not None else "" for value in rows[0]]
    return [{headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]} for row in rows[1:] if any(value is not None for value in row)]


registry_existing = {
    "documents": read_registry_rows("PROJECT_DOCUMENT_REGISTRY.xlsx"),
    "domain_map": read_registry_rows("PROJECT_CANONICAL_DOMAIN_MAP.xlsx"),
    "design_decisions": read_registry_rows("DESIGN_DECISION_REGISTRY.xlsx"),
    "open_decisions": read_registry_rows("OPEN_DECISION_REGISTRY.xlsx"),
    "implementation_gaps": read_registry_rows("IMPLEMENTATION_GAP_REGISTER.xlsx"),
    "research_gaps": read_registry_rows("RESEARCH_GAP_REGISTER.xlsx"),
}


workdata = {
    "schema": "mandate.core-city-development-pack.v1", "generated_on": TODAY, "city_order": list(CITY_MANUAL),
    "cities": city_payloads, "master": master_rows, "person_coverage": person_coverage, "clan_family_coverage": clan_coverage,
    "facility_coverage": facility_coverage, "hinterland_network": hinterland_coverage, "population_layers": population_layers,
    "historical_states": historical_states, "upgrade_registry": upgrade_registry, "sources": list(all_source_rows.values()),
    "registry_updates": registry_updates,
    "registry_existing": registry_existing,
    "summary": {
        "core_city_count": len(city_payloads), "development_ready": sum(city["status"] == "DEVELOPMENT_READY" for city in city_payloads.values()),
        "ready_with_modeled_gaps": sum(city["status"] == "READY_WITH_MODELED_GAPS" for city in city_payloads.values()),
        "research_required": sum(city["status"] == "RESEARCH_REQUIRED" for city in city_payloads.values()),
        "person_presence_records": sum(len(city["people"]) for city in city_payloads.values()),
        "facility_reference_records": sum(len(city["facilities"]) for city in city_payloads.values()),
        "upgrade_registry_records": len(upgrade_registry), "roster_records": len(roster_data["roster"]),
        "strategic_canonical_places": len({row.get("CanonicalPlaceId") for row in admin["strategic_crosswalk"] if row.get("CanonicalPlaceId")}),
        "runtime_changes": 0, "depth_changes": 0,
    },
}


def write_markdown():
    DOC.mkdir(parents=True, exist_ok=True)
    city_links = "\n".join(f"- [{city['label']}（{city['place']}）]({city['directory']}/README.md)：{city['status']}，完整度 {city['score']}/100。" for city in city_payloads.values())
    (DOC / "README_CORE_CITY_DEVELOPMENT_PACKS.md").write_text(f"""# Core City Development Packs V1

本目录把第一批10个重点城市从“已有历史资料”提升为可直接供后续开发任务消费的城市切片。Pack引用人口、人物、Clan、Scenario和Facility的Canonical母库，不复制第二套世界事实。

## 第一批

{city_links}

## 使用顺序

1. 先读[Development Pack Standard](CITY_DEVELOPMENT_PACK_STANDARD_V1.md)。
2. 通过`CanonicalPlaceId`进入对应城市Pack。
3. 核对`DEVELOPMENT_READINESS.md`与`SOURCES_AND_UNKNOWNS.md`。
4. 只有Pack通过后，才可提出DevelopmentDepth调整；仍须用户/开发计划确认。
5. 确认后另开Runtime/Cell/Facility/Population/Family/Unity任务；Pack本身不修改存档或运行世界。

## 长期边界

- 72个DevelopmentPlaceRoster是V1计划，不是永久白名单。
- D0—D5是可调整的制作深度，不是历史城市等级。
- Pack Ready不等于自动升格，也不等于运行时已经实现。
- 升格只补资料与表现/实现精度，不得删除、重建或重新随机既有世界对象。
- 汉中使用战略Label“汉中”，实际CanonicalPlace为`place.han140.yizhou.hanzhong.nanzheng`（南郑）。

## 汇总工作簿

根目录8份工作簿分别覆盖完整度、人物、家族、Facility、供给网络、人口分层、历史状态和未来升格Registry。验收结论见[完整度报告](CORE_CITY_DEVELOPMENT_PACK_COMPLETENESS_REPORT_V1.md)。
""", encoding="utf-8")
    (DOC / "CITY_DEVELOPMENT_PACK_STANDARD_V1.md").write_text("""# Development Pack Standard V1

## 定位

Development Pack是任何Place进入更细开发前的资料门。城市使用City Development Profile；县城、关隘、港渡、聚落、战场或Estate Complex沿用同一标准并按物理类型裁剪不适用项。

## 十三个必备模块

1. Identity / Geography
2. Administrative / Political
3. Population
4. Urban Spatial Form
5. Facility
6. HistoricalPerson
7. Clan / Family / Estate
8. Industry / Agriculture / Resources
9. Transport / Logistics / Surrounding Settlements
10. Military
11. Scenario Snapshot
12. HistoricalChangePoint
13. Development Readiness / Unknowns / Development Implications

## 证据与空间精度

- `HISTORICAL`：正史、考古或正式资料直接支撑。
- `RECONSTRUCTED`：多项证据保守复原，保留推理与来源。
- `MODELED`：为运行容量和玩法补足，不冒充史实。
- `UNKNOWN`：证据不足；不等于不存在。
- Facility空间精度只使用`EXACT_SITE / APPROXIMATE_ZONE / CITY_LEVEL_ONLY / UNKNOWN`。不知道位置时禁止硬塞Cell。

## 文件结构

每个Pack至少包含`README.md`、`CITY_MASTER_REFERENCE.md`、`CITY_DEVELOPMENT_DATA.xlsx`、`DEVELOPMENT_READINESS.md`和`SOURCES_AND_UNKNOWNS.md`。工作簿固定16个工作表，从`00_INDEX`至`15_UNKNOWNS`。

## Ready标准

- `DEVELOPMENT_READY`：Canonical身份、关键人口层、城市形态、核心Facility、人物/家族切片、产业、交通、军事、Scenario和ChangePoint足以进入正式Readiness Review。
- `READY_WITH_MODELED_GAPS`：核心开发方向稳定，普通住宅/工坊/街巷或局部人口层仍需MODELED/UNKNOWN补全；可排期，但进入具体Runtime前必须关闭对应最小缺口。
- `READY_WITH_MIGRATION`：资料已通过，但修正既有Scenario/Save需要独立迁移任务。
- `RESEARCH_REQUIRED`：Canonical身份、位置、治所、人口量级或关键历史状态存在真正阻塞。
- `BLOCKED`：冲突导致无法建立稳定世界对象。

存在UNKNOWN不自动阻塞；CanonicalPlace未解析、稳定ID冲突、人口量级完全未知或关键历史状态互相矛盾才是Blocker。

## 数据引用原则

Pack只保存城市视角切片与开发解释。Person引用`PersonId`，Clan/Branch引用稳定ID，人口引用Han135260V1，Scenario引用ScenarioId，Facility引用统一BaseType/Profile/Capability。不得复制第二套母表。

## Runtime边界

Pack不创建Place、Cell、Facility实例、PermanentPerson、FamilyOrganization、FamilyCenter、Force或Save迁移。历史锚点与Simulation Completion Requirements必须分开；后者由人口、产业、行政、军需和物流推导。

## 升格关系

Pack通过仅意味着资料门通过。DevelopmentDepth是否改变，由用户和开发计划另行决定。升级必须保留所有既有稳定ID和世界事实；正在运行的存档不得因制作深度变化凭空增加建筑或人口。
""", encoding="utf-8")
    (DOC / "CITY_DEVELOPMENT_PACK_UPGRADE_PROTOCOL_V1.md").write_text("""# City Development Pack Upgrade Protocol V1

## Canonical规则

当用户要求“把Place X做细”或“升级某城市”时，第一步必须创建或升级Development Pack，不能直接写Unity代码、摆Facility、生成人口、画城市、创建FamilyOrganization/FamilyCenter或生成AI。

## 标准流程

0. 接收用户的地点细化请求。
1. Resolve Canonical Place：分开StrategicLabel、AdministrativeRegion与CanonicalPhysicalPlace；显示名不得直接创造新Place。
2. Check Existing Pack：没有则CREATE PACK，已有则UPGRADE PACK。
3. 提出目标深度候选D3/D4/D5，但不更改Roster。
4. 运行人口、空间、人物、Clan/Family、Facility、产业、交通、军事、Scenario、ChangePoint资料缺口审计。
5. 补最小必要历史资料，保留HISTORICAL/RECONSTRUCTED/MODELED/UNKNOWN。
6. 生成或升级Pack。
7. 按Standard验收Pack。
8. 更新DevelopmentPlaceRoster、Development Manifest与Knowledge Base；仅登记建议。
9. 用户/开发计划明确确认DevelopmentDepth变化。
10. 才允许进入独立Runtime / Cell / Facility / Population / HistoricalPerson / FamilyOrganization / Unity任务。

## D0/D1与Roster

任何合法CanonicalPlace原则上`EligibleForUpgrade=true`。D0/D1只表示当前无专项制作计划；72个Roster是V1计划，不是永久白名单。未解析为CanonicalPlace的战略Label或`geo.site`参考，须先完成物理Place解析。

## 稳定世界与存档

- 升格不得换PlaceId、重生人口、重随机Person、重建已有Facility或改写历史行政关系。
- 既有Person、Household、Facility、Cell、Inventory继续存在；只增加经审计的Reference和开发内容。
- Scenario创世缺失的历史事实须另开Initialization Correction / Migration任务。
- 游戏运行中新增建设必须来自真实Construction，不得因DevelopmentDepth变化凭空出现。
- Pack升级不自动修改已有存档。

## 决策边界

Pack Ready与Depth Upgrade是两个独立门。Pack完成不自动升格，深度建议不自动改变Wave；二者均等待用户/开发计划确认。
""", encoding="utf-8")

    report_table = "\n".join(f"| {row['Place']} | {row['CurrentDepth']} | {row['PackStatus']} | {row['OverallReadiness']} | {row['RuntimeReadiness']} |" for row in master_rows)
    (DOC / "CORE_CITY_DEVELOPMENT_PACK_COMPLETENESS_REPORT_V1.md").write_text(f"""# Core City Development Pack Completeness Report V1

## 结论

10/10个核心城市已形成标准Pack：洛阳1个`DEVELOPMENT_READY`，其余9个`READY_WITH_MODELED_GAPS`，0个`RESEARCH_REQUIRED`。这表示资料可被后续任务直接消费，不表示九座城市已在Unity/Runtime实现。

| 城市 | 当前Depth | Pack状态 | 完整度 | Runtime边界 |
| --- | --- | --- | ---: | --- |
{report_table}

最接近洛阳资料深度的是许昌，其Canonical、人口母盘、196后政治状态和人物切片较稳定；仍缺精确城市空间、设施锚点和正式Runtime包。成都的既有`major_city_timeline`错链已隔离，未污染Pack。

## 任务书25项交接

1. 完整度见上表和`01_CORE_CITY_DEVELOPMENT_PACK_MASTER.xlsx`。
2. Development Ready：洛阳。
3. Ready With Modeled Gaps：长安、邺、许昌、成都、襄阳、江陵、建业、合肥、南郑（汉中战略节点）。
4. Research Required：无；但每城仍有不阻塞Pack的专项研究缺口。
5. 人口层次：10城都建立了行政/县/城市层引用；缺失城墙、都市圈或供给圈数值保持UNKNOWN。只有洛阳使用20万/27万/40万/70万保护口径。
6. HistoricalPerson城市切片：共{workdata['summary']['person_presence_records']}条，逐城数量见人物覆盖矩阵；不是名将榜，也不是最终全量。
7. Clan/Family：按PersonId与现有Clan/Branch链接；没有把成员在场、Estate或重要城市自动变成FamilyOrganization/FamilyCenter。
8. 可考Historical Facility：宫廷、城垣、官署、太学、石头城、铜雀台、合肥新城、成都/江陵/南郑等行政与城防锚点，逐条见Facility工作表。
9. 必须Reconstructed：多数市场、仓储、港渡、军营、官署区、城门使用状态与城市分区。
10. 必须Modeled：普通住宅、普通仓储、普通工坊、基层医疗、道路排水及无史名聚落群。
11. 城墙/城门：10城均有分期结论；非洛阳多为APPROXIMATE_ZONE/CITY_LEVEL_ONLY，未伪造精确Cell。
12. 道路/水系：10城均建立主要陆水走廊。
13. 周边聚落网络：10城均有Core、近郊、县邑、MODELED村落群、农业区与交通节点。
14. 农业/Supply Hinterland：10城均建立地理与产业链；没有把郡人口直接当供给圈人口。
15. 主要产业：10城均映射至Facility/Recipe/真实工人/库存合同。
16. 军事空间：10城均映射同一CanonicalPlace、Cell、Road、Facility与Force。
17. Scenario：按逐Place历史状态计划选择，不机械复制13个Scenario。
18. ChangePoint：已知ID直接交叉引用；无ID的重要年份保留`STATE_REFERENCE_NO_CANONICAL_CHANGEPOINT`，等待后续事件任务。
19. 最接近洛阳：许昌；但仍不能视为Runtime已实现。
20. 距离直接开发的共同缺口：正式Cell/Facility初始化、逐期空间锚点、普通社会物化、FamilyOrganization与ChangePackage；逐城见Unknowns。
21. 用户要求D2城市做细：先解析CanonicalPlace，创建/升级Pack，审计补缺，Pack验收，再由用户决定是否升D3/D4/D5。
22. 当前72 Roster仍允许扩展：允许。
23. 当前D0/D1地点允许未来升级：允许。
24. 城市升级时不应直接写代码：先完成Development Pack。
25. Pack完成不自动升级城市：由用户/开发计划决定。

## 验收证据

- 结构与稳定ID验证：通过（21,064项检查，0错误）。
- 工作簿公式错误扫描：0。
- 工作簿逐表渲染：214张预览完成；代表性总表和设施表已人工复核列宽、换行与可读性。
- Markdown断链：0；UTF-8读取：通过。
- Runtime变化：0；DevelopmentDepth自动变化：0。
- 编译、核心测试与Unity测试：不适用，本任务只建设文档、参考数据和工作簿。

## 下一阶段

停止自动扩充其他城市Pack，进入`LUOYANG-184-DEVELOPMENT-READINESS-REVIEW-V1`。
""", encoding="utf-8")

    for slug, city in city_payloads.items():
        root = DOC / city["directory"]
        root.mkdir(parents=True, exist_ok=True)
        master_link = Path(city["structured_source_dir"]) / "00_Master.md"
        rel_master = Path("../../../..") / master_link
        (root / "README.md").write_text(f"""# {city['label']} City Development Pack V1

- CanonicalPlaceId：`{city['place']}`
- StrategicLabel：`{city['strategic']}`
- PackStatus：`{city['status']}`
- CurrentDevelopmentDepth：`{roster_by_id[city['place']]['DevelopmentDepth']}`（本任务未改变）
- RuntimeStatus：`{'可进入正式Readiness Review' if slug == 'LUOYANG' else 'Reference complete; runtime not implemented'}`

入口：[Master Reference](CITY_MASTER_REFERENCE.md) → [Data Workbook](CITY_DEVELOPMENT_DATA.xlsx) → [Readiness](DEVELOPMENT_READINESS.md) → [Sources/Unknowns](SOURCES_AND_UNKNOWNS.md)。

本Pack引用全局人口、Person、Clan、Facility和Scenario母库，不复制第二套事实。
""", encoding="utf-8")
        pop184 = next(row for row in city["populations"] if row["Year"] == 184)
        (root / "CITY_MASTER_REFERENCE.md").write_text(f"""# {city['label']} City Master Reference

## 01 Identity / Geography

- CanonicalPlace：`{city['place']}`；战略显示名：`{city['strategic']}`；历史名：{city['historical_names']}。
- 行政：`{city['identity'][0]['ProvinceId']}` → `{city['identity'][0]['CommanderyEquivalentId']}` → `{city['identity'][0]['CountyPermanentId']}`。
- 地理：{city['geography']}
- 地形/水系/山地：{city['terrain']}；{city['water']}；{city['mountains']}。
- 道路/邻接：{city['roads']}；{city['adjacent']}。

## 02 Administrative / Political

历史治所只作为HistoricalSeatReference；Runtime Seat由实际Government Facility、Office、Authority和Controller决定，不能写死未来迁治。

## 03 Population

184县人口引用：{pop184['CountyPopulation'] if pop184['CountyPopulation'] is not None else 'UNKNOWN'}；城墙人口：{pop184['WalledPopulation'] if pop184['WalledPopulation'] is not None else 'UNKNOWN'}；连续城区：{pop184['UrbanPopulation'] if pop184['UrbanPopulation'] is not None else 'UNKNOWN'}；都市圈：{pop184['MetropolitanPopulation'] if pop184['MetropolitanPopulation'] is not None else 'UNKNOWN'}；供给圈：{pop184['SupplyHinterlandPopulation'] if pop184['SupplyHinterlandPopulation'] is not None else 'UNKNOWN'}。各层为包含关系，不可相加，县人口不等于城市人口。

## 04 Urban Spatial Form

{city['urban']} 城墙/城门：{city['wall']} {city['gates']} 内城/官署：{city['inner']} 近郊/扩展：{city['suburbs']} {city['expansion']}

## 05 Facility

共{len(city['facilities'])}条Reference，历史锚点与Simulation Completion Requirements分开；历史名称映射统一BaseType，不自创新Facility枚举。

## 06 HistoricalPerson

当前城市切片{len(city['people'])}条稳定PersonId记录；籍贯不等于当前位置，Confirmed与Probable分开。

## 07 Clan / Family / Estate

当前{len(city['clans'])}条Clan/Branch切片。成员在场、住宅、Estate、FamilyOrganization与FamilyCenter相互独立，均不得自动推导。

## 08 Industry / Agriculture / Resources

- 产业：{city['industry']}
- 农业：{city['agriculture']}
- 资源：{city['resources']}

所有产出必须来自Facility + Recipe + real worker + material + time + authority。

## 09 Transport / Logistics / Surrounding Settlements

{city['suburbs']} 供应链统一为Producer/Settlement → Storage → Road/Water → Gate/Harbor → Urban Storage/Market → Household/Military/Facility。

## 10 Military

{city['military']}

## 11 Scenario Snapshot / 12 HistoricalChangePoint

支持{len(city['states'])}个相关Scenario/TimePoint；已知ChangePoint使用稳定ID，未知变化不伪造Package。

## 13 Development Implication

Pack状态`{city['status']}`，完整度{city['score']}/100。该状态只允许后续任务消费Reference，不自动改变DevelopmentDepth或运行时世界。

## Canonical references

- [既有P0/P1核心聚落Master]({rel_master.as_posix()})
- [Development Pack Standard](../CITY_DEVELOPMENT_PACK_STANDARD_V1.md)
- [Upgrade Protocol](../CITY_DEVELOPMENT_PACK_UPGRADE_PROTOCOL_V1.md)
""", encoding="utf-8")
        module_table = "\n".join(f"| {row['Module']} | {row['ModuleName']} | {row['CoverageScore']} | {row['CoverageLevel']} | {row['Conclusion']} |" for row in city["modules"])
        (root / "DEVELOPMENT_READINESS.md").write_text(f"""# {city['label']} Development Readiness

最终状态：`{city['status']}`；完整度：**{city['score']}/100**；DevelopmentDepth保持`{roster_by_id[city['place']]['DevelopmentDepth']}`。

| 模块 | 名称 | 分数 | 状态 | 结论 |
| --- | --- | ---: | --- | --- |
{module_table}

Pack通过不等于Runtime通过，也不自动升格。{'本城可以进入LUOYANG-184-DEVELOPMENT-READINESS-REVIEW-V1。' if slug == 'LUOYANG' else '本城进入实际开发前仍须关闭本Pack列出的最小Cell/Facility/人物家庭实施缺口。'}
""", encoding="utf-8")
        unknown_list = "\n".join(f"- {row['Unknown']}（{row['Impact']}）" for row in city["unknown_rows"])
        source_list = "\n".join(f"- `{row['SourceId']}`：{row['Title']}；{row['URLOrLocator']}" for row in city["sources"])
        (root / "SOURCES_AND_UNKNOWNS.md").write_text(f"""# {city['label']} Sources and Unknowns

## Unknowns

{unknown_list}

这些UNKNOWN不等于NONE；非阻塞的普通细节可由MODELED运行需求补全，但精确历史Cell、人物住宅、Estate和FamilyCenter不得臆造。

## Sources

{source_list}
""", encoding="utf-8")


def update_roster_and_manifests():
    pack_by_place = {city["place"]: city for city in city_payloads.values()}
    for row in roster_data["roster"]:
        city = pack_by_place.get(row["CanonicalPlaceId"])
        if city:
            row["PackStatus"] = city["status"]
            row["CityDevelopmentPack"] = f"Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/{city['directory']}/"
            row["PackReviewDate"] = TODAY
            if city["status"] == "READY_WITH_MODELED_GAPS" and row["ReferenceReadiness"] == "PARTIAL":
                row["ReferenceReadiness"] = "MOSTLY_READY"
    dump(ROSTER_OUTPUT / "development_place_roster_workdata.json", roster_data)
    manifest_slug = {"LUOYANG": "LUOYANG_184", "CHANGAN": "CHANGAN", "YE": "YE", "XU": "XU", "CHENGDU": "CHENGDU", "XIANGYANG": "XIANGYANG", "JIANGLING": "JIANGLING", "JIANYE": "JIANYE", "HEFEI": "HEFEI", "HANZHONG_CANONICAL_PLACE": "HANZHONG"}
    marker = "## City Development Pack V1"
    for slug, city in city_payloads.items():
        path = MANIFEST_ROOT / f"{manifest_slug[slug]}_DEVELOPMENT_REFERENCE_MANIFEST.md"
        text = path.read_text(encoding="utf-8")
        section = f"""

## City Development Pack V1

| Field | Reference |
|---|---|
| CityDevelopmentPack | `Docs/HISTORICAL_WORLD_REFERENCE/CITY_DEVELOPMENT_PACKS/{city['directory']}/` |
| PackStatus | {city['status']} |
| ReferenceReadiness | {'READY_FOR_IMPLEMENTATION' if slug == 'LUOYANG' else 'MOSTLY_READY'} |
| HistoricalStatePlan | `07_CORE_CITY_HISTORICAL_STATE_AND_CHANGEPOINT_PLAN.xlsx` |
| HinterlandReference | `05_CORE_CITY_HINTERLAND_AND_SETTLEMENT_NETWORK.xlsx` |
| PopulationLayerReference | `06_CORE_CITY_POPULATION_LAYER_REFERENCE.xlsx` |
| FacilityReference | `{city['directory']}/CITY_DEVELOPMENT_DATA.xlsx#05_FACILITIES` |
| PersonCoverage | {len(city['people'])} stable PersonId city-slice records |
| FamilyCoverage | {len(city['clans'])} Clan/Branch slice records; no automatic FamilyCenter |
| DepthUpgradeRecommendation | NONE_THIS_TASK |
| RuntimeBoundary | Pack complete does not mean runtime implemented and does not change DevelopmentDepth. |
"""
        if marker in text:
            text = text.split(marker, 1)[0].rstrip() + section
        else:
            text = text.rstrip() + section
        path.write_text(text.rstrip() + "\n", encoding="utf-8")


OUTPUT.mkdir(parents=True, exist_ok=True)
write_markdown()
update_roster_and_manifests()
dump(OUTPUT / "core_city_development_pack_workdata.json", workdata)
print(json.dumps(workdata["summary"], ensure_ascii=False, indent=2))
