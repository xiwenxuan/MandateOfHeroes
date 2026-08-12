#!/usr/bin/env python3
"""Validate the reference-only Core City Development Pack V1 delivery."""

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[2]
OUTPUT = REPO / "outputs" / "HAN_135_260_CORE_CITY_DEVELOPMENT_PACK_AND_UPGRADE_PROTOCOL_V1"
DOC = REPO / "Docs" / "HISTORICAL_WORLD_REFERENCE" / "CITY_DEVELOPMENT_PACKS"
WORKDATA = json.loads((OUTPUT / "core_city_development_pack_workdata.json").read_text(encoding="utf-8"))

EXPECTED_CITIES = {
    "LUOYANG": "place.han140.sili.henan.luoyang",
    "CHANGAN": "place.han140.sili.jingzhao.changan",
    "YE": "place.han140.jizhou.wei.ye",
    "XU": "place.han140.yuzhou.yingchuan.xu",
    "CHENGDU": "place.han140.yizhou.shu.chengdu",
    "XIANGYANG": "place.han140.jingzhou.nan.xiangyang",
    "JIANGLING": "place.han140.jingzhou.nan.jiangling",
    "JIANYE": "place.han140.yangzhou.danyang.moling",
    "HEFEI": "place.han140.yangzhou.jiujiang.hefei",
    "HANZHONG_CANONICAL_PLACE": "place.han140.yizhou.hanzhong.nanzheng",
}
EXPECTED_CITY_SHEETS = [
    "00_INDEX", "01_IDENTITY_ADMIN", "02_HISTORICAL_STATES", "03_POPULATION",
    "04_URBAN_FORM", "05_FACILITIES", "06_HISTORICAL_PERSONS", "07_CLAN_FAMILY_ESTATE",
    "08_INDUSTRY_AGRICULTURE", "09_TRANSPORT_SETTLEMENTS", "10_MILITARY",
    "11_SCENARIO_SNAPSHOTS", "12_CHANGEPOINTS", "13_DEVELOPMENT_MAPPING", "14_SOURCES", "15_UNKNOWNS",
]
SUMMARY_BOOKS = [
    "01_CORE_CITY_DEVELOPMENT_PACK_MASTER.xlsx",
    "02_CORE_CITY_HISTORICAL_PERSON_COVERAGE.xlsx",
    "03_CORE_CITY_CLAN_FAMILY_COVERAGE.xlsx",
    "04_CORE_CITY_FACILITY_REFERENCE_COVERAGE.xlsx",
    "05_CORE_CITY_HINTERLAND_AND_SETTLEMENT_NETWORK.xlsx",
    "06_CORE_CITY_POPULATION_LAYER_REFERENCE.xlsx",
    "07_CORE_CITY_HISTORICAL_STATE_AND_CHANGEPOINT_PLAN.xlsx",
    "08_CITY_DEVELOPMENT_PACK_UPGRADE_REGISTRY.xlsx",
]
FACILITY_TYPES = {
    "Residence", "Farmland", "Plantation", "HerbField", "Pasture", "Forestry", "Mine", "Quarry",
    "Mill", "Brewery", "Smelter", "Smithy", "Carpentry", "SilkwormHouse", "SilkReelingWorkshop",
    "WeavingWorkshop", "DyeWorkshop", "MedicineWorkshop", "Shipyard", "Kitchen", "Warehouse",
    "Granary", "Stable", "CarriageYard", "CourierStation", "Harbor", "Market", "Shop", "Inn",
    "Clinic", "GuildHall", "MerchantHall", "GovernmentOffice", "CourtHall", "School", "Academy",
    "Library", "RitualHall", "Observatory", "TrainingHall", "Barracks", "TrainingGround",
    "FieldHospital", "Wall", "Gate", "Moat", "Fort", "BeaconTower", "Road", "Bridge", "Canal",
    "Well", "WaterIntake", "Drainage", "Dike", "Garden", "Plaza", "Courtyard",
}
EVIDENCE_LEVELS = {"HISTORICAL", "RECONSTRUCTED", "MODELED", "UNKNOWN"}
FORMULA_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}


errors: list[str] = []
checks: list[str] = []


def require(condition: bool, message: str) -> None:
    if condition:
        checks.append(message)
    else:
        errors.append(message)


def load_records(relative: str, keys: tuple[str, ...]) -> list[dict]:
    payload = json.loads((REPO / relative).read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return payload
    for key in keys:
        if isinstance(payload.get(key), list):
            return payload[key]
    raise ValueError(relative)


persons = load_records("Assets/StreamingAssets/HistoricalPersons/Han135260V1/persons.json", ("persons", "records"))
clans = load_records("Assets/StreamingAssets/HistoricalPersons/Han135260V1/clans.json", ("clans", "records"))
branches = load_records("Assets/StreamingAssets/HistoricalPersons/Han135260V1/branches.json", ("branches", "records"))
person_ids = {row["person_id"] for row in persons}
clan_ids = {row["clan_id"] for row in clans}
branch_ids = {row["branch_id"] for row in branches}
scenario_ids = {row["scenario_id"] for row in json.loads((REPO / "outputs/HAN_135_260_HISTORICAL_WORLD_REFERENCE_DEEPENING_V1/deepening_workdata.json").read_text(encoding="utf-8"))["scenarios"]}
source_ids = {row["SourceId"] for row in WORKDATA["sources"] if row.get("SourceId")}

require(set(WORKDATA["cities"]) == set(EXPECTED_CITIES), "10 Core City Pack coverage")
require(WORKDATA["summary"]["core_city_count"] == 10, "core city summary count = 10")
require(WORKDATA["summary"]["runtime_changes"] == 0, "runtime changes = 0")
require(WORKDATA["summary"]["depth_changes"] == 0, "automatic depth changes = 0")
require(WORKDATA["summary"]["development_ready"] == 1, "one Development Ready pack")
require(WORKDATA["summary"]["ready_with_modeled_gaps"] == 9, "nine Ready With Modeled Gaps packs")

presence_keys = set()
facility_ids = set()
for slug, expected_place in EXPECTED_CITIES.items():
    city = WORKDATA["cities"].get(slug, {})
    require(city.get("place") == expected_place, f"{slug} CanonicalPlace resolution")
    require(len(city.get("modules", [])) == 13, f"{slug} has 13 pack modules")
    require(len(city.get("people", [])) == 10, f"{slug} person slice has 10 records")
    for row in city.get("people", []):
        key = (slug, row.get("PersonId"), row.get("ScenarioYearOrRange"))
        require(key not in presence_keys, f"unique person presence {key}")
        presence_keys.add(key)
        require(row.get("PersonId") in person_ids, f"PersonId integrity {row.get('PersonId')}")
        require(row.get("ScenarioId") in scenario_ids, f"ScenarioId integrity {row.get('ScenarioId')}")
        require(row.get("PresenceType") != "Residence", f"canonical Resident presence type {slug}/{row.get('PersonId')}")
        if row.get("ClanId"):
            require(row["ClanId"] in clan_ids, f"ClanId integrity {row['ClanId']}")
        if row.get("BranchId"):
            require(row["BranchId"] in branch_ids, f"BranchId integrity {row['BranchId']}")
    for row in city.get("facilities", []):
        fid = row.get("FacilityReferenceId")
        require(fid not in facility_ids, f"unique facility reference {fid}")
        facility_ids.add(fid)
        require(row.get("BaseType") in FACILITY_TYPES, f"Facility BaseType reference {fid}")
        require(row.get("EvidenceLevel") in EVIDENCE_LEVELS, f"facility evidence level {fid}")
        require(row.get("PlaceId") == expected_place, f"facility place integrity {fid}")
        if slug != "LUOYANG" and row.get("AnchorPrecision") == "EXACT_SITE":
            require(not row.get("CellId"), f"no invented exact Cell outside Luoyang {fid}")
        for source_id in str(row.get("SourceIds") or "").split("|"):
            if source_id:
                require(source_id in source_ids, f"SourceId integrity {source_id}")
    directory = DOC / city.get("directory", slug)
    for filename in ("README.md", "CITY_MASTER_REFERENCE.md", "CITY_DEVELOPMENT_DATA.xlsx", "DEVELOPMENT_READINESS.md", "SOURCES_AND_UNKNOWNS.md"):
        require((directory / filename).is_file(), f"{slug} output exists: {filename}")
    workbook = load_workbook(directory / "CITY_DEVELOPMENT_DATA.xlsx", read_only=True, data_only=False)
    require(workbook.sheetnames == EXPECTED_CITY_SHEETS, f"{slug} workbook has exact 16-sheet contract")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                require(cell.value not in FORMULA_ERRORS, f"{slug}/{sheet.title} formula scan")
    workbook.close()

require(WORKDATA["summary"]["person_presence_records"] == 100, "100 person presence records")
require(WORKDATA["summary"]["facility_reference_records"] == 123, "123 facility reference records")
wrong_chengdu = "admin.han140.jingzhou.nanyang.chengdu"
active_place_values = [
    city.get("place") for city in WORKDATA["cities"].values()
] + [
    row.get("PlaceId") for city in WORKDATA["cities"].values() for row in city.get("facilities", [])
]
require(wrong_chengdu not in active_place_values, "wrong Chengdu crosswalk isolated from active city/facility references")
require(WORKDATA["cities"]["HANZHONG_CANONICAL_PLACE"]["strategic"] == "汉中", "Hanzhong strategic label preserved")

upgrade_rows = WORKDATA["upgrade_registry"]
require(len(upgrade_rows) == 89, "89-row upgrade registry")
upgrade_place_ids = {row.get("PlaceId") for row in upgrade_rows if row.get("PlaceId")}
roster_payload = json.loads((REPO / "outputs/DEVELOPMENT_PLACE_ROSTER_AND_REFERENCE_READINESS_V1/development_place_roster_workdata.json").read_text(encoding="utf-8"))
roster = roster_payload["roster"]
require(len(roster) == 72, "72 roster records preserved")
require({row["CanonicalPlaceId"] for row in roster}.issubset(upgrade_place_ids), "upgrade registry covers all roster Places")
depth_counts = Counter(row["DevelopmentDepth"] for row in roster)
require(depth_counts == Counter({"D5": 1, "D4": 15, "D3": 33, "D2": 23}), "roster depth counts unchanged")
require(sum(bool(row.get("CityDevelopmentPack")) for row in roster) == 10, "ten roster rows link City Development Packs")

for filename in SUMMARY_BOOKS:
    path_ = DOC / filename
    require(path_.is_file(), f"summary workbook exists: {filename}")
    workbook = load_workbook(path_, read_only=True, data_only=False)
    require(workbook.sheetnames == ["说明", "数据", "来源"], f"summary workbook sheets: {filename}")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                require(cell.value not in FORMULA_ERRORS, f"summary formula scan {filename}/{sheet.title}")
    workbook.close()

manifest_files = {
    "LUOYANG_184", "CHANGAN", "YE", "XU", "CHENGDU", "XIANGYANG", "JIANGLING", "JIANYE", "HEFEI", "HANZHONG"
}
for name in manifest_files:
    path_ = REPO / "Docs/KNOWLEDGE_BASE/DEVELOPMENT_MANIFESTS" / f"{name}_DEVELOPMENT_REFERENCE_MANIFEST.md"
    text = path_.read_text(encoding="utf-8")
    require("## City Development Pack V1" in text, f"manifest upgraded: {name}")
    for field in ("CityDevelopmentPack", "PackStatus", "ReferenceReadiness", "HistoricalStatePlan", "HinterlandReference", "PopulationLayerReference", "FacilityReference", "PersonCoverage", "FamilyCoverage"):
        require(field in text, f"manifest field {name}/{field}")

registry_expectations = {
    "PROJECT_DOCUMENT_REGISTRY.xlsx": "doc.core-city-packs-v1.readme",
    "PROJECT_CANONICAL_DOMAIN_MAP.xlsx": "CityDevelopmentPack",
    "DESIGN_DECISION_REGISTRY.xlsx": "DEC-DEVPACK-001",
    "OPEN_DECISION_REGISTRY.xlsx": "OPEN-DEVPACK-001",
    "IMPLEMENTATION_GAP_REGISTER.xlsx": "IMP-GAP-CITYPACK-001",
    "RESEARCH_GAP_REGISTER.xlsx": "RES-GAP-CITYPACK-001",
}
for filename, token in registry_expectations.items():
    path_ = REPO / "Docs/KNOWLEDGE_BASE/REGISTRY" / filename
    workbook = load_workbook(path_, read_only=True, data_only=False)
    values = "\n".join(str(cell.value) for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row if cell.value is not None)
    require(token in values, f"knowledge registry update: {filename}")
    workbook.close()

markdown_roots = [DOC]
markdown_files = [path_ for root in markdown_roots for path_ in root.rglob("*.md")]
markdown_files += [
    REPO / "Docs/GAME_SYSTEMS_MASTER_AND_STATUS.md",
    REPO / "Docs/HISTORICAL_WORLD_REFERENCE/README_历史世界开发参考资料索引.md",
    REPO / "Docs/HISTORICAL_WORLD_REFERENCE/DEVELOPMENT_PLACE_ROSTER_V1/README.md",
    REPO / "Docs/KNOWLEDGE_BASE/README_PROJECT_KNOWLEDGE_BASE.md",
    REPO / "Docs/TASK_HAN_135_260_CORE_CITY_DEVELOPMENT_PACK_AND_UPGRADE_PROTOCOL_V1.md",
]
broken_links = []
for md in markdown_files:
    text = md.read_text(encoding="utf-8")
    for target in re.findall(r"\[[^\]]+\]\(([^)]+)\)", text):
        target = target.split("#", 1)[0].strip()
        if not target or "://" in target or target.startswith("mailto:"):
            continue
        resolved = (md.parent / target).resolve()
        if not resolved.exists():
            broken_links.append(f"{md.relative_to(REPO)} -> {target}")
require(not broken_links, "Markdown broken link = 0" if not broken_links else "broken links: " + " | ".join(broken_links[:10]))

preview_files = list((OUTPUT / "previews").rglob("*.png"))
require(len(preview_files) >= 200, "all workbook sheets rendered for visual inspection")
report = json.loads((OUTPUT / "workbook_build_report.json").read_text(encoding="utf-8"))
require(report.get("formulaErrors") == 0, "artifact-tool formula error scan = 0")
require(len(report.get("workbooks", [])) >= 25, "workbook build report covers deliverables and registries")

summary = {
    "schema": "mandate.core-city-development-pack-validation.v1",
    "status": "PASS" if not errors else "FAIL",
    "checks_passed": len(checks),
    "errors": errors,
    "core_city_count": WORKDATA["summary"]["core_city_count"],
    "person_presence_records": WORKDATA["summary"]["person_presence_records"],
    "facility_reference_records": WORKDATA["summary"]["facility_reference_records"],
    "upgrade_registry_records": len(upgrade_rows),
    "roster_records": len(roster),
    "rendered_previews": len(preview_files),
    "runtime_changes": 0,
    "depth_changes": 0,
}
(OUTPUT / "validation_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print(json.dumps(summary, ensure_ascii=False, indent=2))
if errors:
    raise SystemExit(1)
