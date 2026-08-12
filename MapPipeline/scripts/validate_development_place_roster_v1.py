#!/usr/bin/env python3
"""Validate Development Place Roster V1 reference and planning artifacts."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[2]
OUTPUT = REPO / "outputs" / "DEVELOPMENT_PLACE_ROSTER_AND_REFERENCE_READINESS_V1"
DOC = REPO / "Docs" / "HISTORICAL_WORLD_REFERENCE" / "DEVELOPMENT_PLACE_ROSTER_V1"
MANIFEST = REPO / "Docs" / "KNOWLEDGE_BASE" / "DEVELOPMENT_MANIFESTS"
SOURCE = REPO / "outputs" / "HAN_135_260_ADMINISTRATIVE_SEAT_CANONICAL_PLACE_AND_HISTORICAL_WORLD_STATE_V1"
DEEP = REPO / "outputs" / "HAN_135_260_HISTORICAL_WORLD_REFERENCE_DEEPENING_V1"


def read_json(path):
    return json.loads(path.read_text(encoding="utf-8"))


data = read_json(OUTPUT / "development_place_roster_workdata.json")
source = read_json(SOURCE / "administrative_seat_world_state_workdata.json")
deep = read_json(DEEP / "deepening_workdata.json")
failures = []
checks = []


def check(name, condition, detail=""):
    checks.append({"name": name, "passed": bool(condition), "detail": detail})
    if not condition:
        failures.append(f"{name}: {detail}")


roster = data["roster"]
ids = [row["CanonicalPlaceId"] for row in roster]
source_place_ids = {row["CanonicalPlaceId"] for row in source["canonical_places"]}
site_ids = {row["transport_id"] for row in deep["transport_nodes"] if row["transport_id"].startswith("geo.site.")}
check("Roster has records", bool(roster), str(len(roster)))
check("Roster count frozen", len(roster) == 72, str(len(roster)))
check("No duplicate Place", len(ids) == len(set(ids)), str(len(ids) - len(set(ids))))
check("CanonicalPlace ID validation", all(pid in source_place_ids or pid in site_ids for pid in ids), "every roster ID reuses a source Place/site reference")
check("No AdministrativeRegion as fake Place", not any(pid.startswith("admin.") for pid in ids), "no admin.* IDs")
check("77 Strategic Label source coverage", len(source["strategic_crosswalk"]) == 77, str(len(source["strategic_crosswalk"])))
check("133 Core Settlement source coverage", len(source["canonical_places"]) == 133, str(len(source["canonical_places"])))
check("250 Priority County source coverage", len(source["priority_places"]) == 250, str(len(source["priority_places"])))
check("Strategic labels are not roster", len(roster) != len(source["strategic_crosswalk"]), f"roster={len(roster)} labels=77")
check("Core settlements are not roster", len(roster) != len(source["canonical_places"]), f"roster={len(roster)} core=133")

valid_depths = {f"D{i}" for i in range(6)}
check("D0-D5 enum validation", all(row["DevelopmentDepth"] in valid_depths for row in roster), "all frozen roster depths valid")
check("D5 rare flagship", sum(row["DevelopmentDepth"] == "D5" for row in roster) == 1, "exactly Luoyang")
check("D1 not inflated", sum(row["DevelopmentDepth"] == "D1" for row in roster) == 0, "ordinary simulation places remain outside special roster")
check("Non-urban D4 included", any(row["DevelopmentDepth"] == "D4" and row["CanonicalPlaceId"].startswith("geo.site.") for row in roster), "non-urban eligibility proven")

d4d5 = data["d4_d5"]
slug = {
    "洛阳": "LUOYANG_184", "长安": "CHANGAN", "邺": "YE", "许昌": "XU", "成都": "CHENGDU",
    "襄阳": "XIANGYANG", "江陵": "JIANGLING", "建业": "JIANYE", "合肥": "HEFEI", "汉中": "HANZHONG",
    "虎牢": "HULAO", "樊城": "FANCHENG", "夏口": "XIAKOU", "阳平关": "YANGPING_PASS",
    "剑阁": "JIANGE", "濡须口": "RUXUKOU",
}
missing_manifests = []
for row in d4d5:
    path = MANIFEST / f"{slug[row['DevelopmentDisplayName']]}_DEVELOPMENT_REFERENCE_MANIFEST.md"
    if not path.exists() or "## Development Place Roster V1" not in path.read_text(encoding="utf-8"):
        missing_manifests.append(str(path.relative_to(REPO)))
check("D4/D5 Manifest coverage", not missing_manifests and len(d4d5) == 16, "|".join(missing_manifests) or "16/16")

state_rows = data["historical_state_plan"]
state_keys = [(row["PlaceId"], row["ScenarioYear"], row["HistoricalStateId"]) for row in state_rows]
check("HistoricalState Plan integrity", len(state_keys) == len(set(state_keys)), "unique place/time/state rows")
check("HistoricalState Place links", all(row["PlaceId"] in set(ids) for row in state_rows), "all states point to roster")
valid_support = {f"S{i}" for i in range(5)}
valid_snapshot = {f"H{i}" for i in range(6)}
check("SupportLevel validation", all(row["SupportLevel"] in valid_support for row in state_rows), "S0-S4")
check("Historical depth validation", all(row["RequiredSnapshotDepth"] in valid_snapshot for row in state_rows), "H0-H5")
formal_years = set(data["summary"]["formal_scenario_years"])
check("Scenario validity", all((row["ScenarioYear"] in formal_years) == (row["TimePointType"] == "FORMAL_SCENARIO") for row in state_rows), "non-scenario years are explicit reference/change years")
change_ids = {row["ChangePointId"] for row in source["change_points"]}
check("MajorChangePoint link validity", all(not row["MajorChangePointId"] or row["MajorChangePointId"] in change_ids for row in state_rows), "all non-empty links exist")

required_readiness = {"Geography", "Seat", "Population", "Urban", "Facility", "Transport", "Industry", "Person", "Clan", "Family", "Military", "Scenario", "ChangePoint", "Cell", "Art", "Runtime", "OverallReadiness"}
check("Readiness matrix completeness", all(required_readiness.issubset(row) and all(row[key] for key in required_readiness) for row in data["readiness"]), "all 72 rows cover required dimensions")
check("Readiness row count", len(data["readiness"]) == len(roster), str(len(data["readiness"])))
check("Luoyang ready gate", [row["PlaceName"] for row in data["readiness"] if row["OverallReadiness"] == "READY_FOR_IMPLEMENTATION"] == ["洛阳"], "only Luoyang enters formal review")

blocker_types = {"HISTORICAL_RESEARCH_BLOCKER", "DATA_MAPPING_BLOCKER", "DESIGN_BLOCKER", "IMPLEMENTATION_BLOCKER"}
check("Blocker classification", all(row["BlockerType"] in blocker_types for row in data["blockers"]), "four formal types")
check("Blocker Place links", all(row["PlaceId"] in set(ids) for row in data["blockers"]), "all blockers target roster")

wave_order = {"WAVE_0": 0, "WAVE_1": 1, "WAVE_2": 2, "WAVE_3": 3, "WAVE_4": 4, "RESERVE": 5}
check("DevelopmentWave values", all(row["RecommendedWave"] in wave_order for row in roster), "known wave vocabulary")
check("Wave 0 fixed", {row["DevelopmentDisplayName"] for row in roster if row["RecommendedWave"] == "WAVE_0"} == {"洛阳", "虎牢", "函谷关"}, "洛阳D5 + 虎牢D4 + 函谷D3")
check("DevelopmentWave dependency cycle", [row["Wave"] for row in data["wave_plan"]] == ["WAVE_0", "WAVE_1", "WAVE_2", "WAVE_3", "WAVE_4", "RESERVE"], "strict forward sequence")
check("DevelopmentRegion is not world entity", all(row["IsWorldEntity"] == "NO_DEVELOPMENT_WORK_PACKAGE_ONLY" for row in data["region_slices"]), "all 8 slices are planning packages")

workbook_names = [
    "01_DEVELOPMENT_PLACE_ROSTER.xlsx", "02_DEVELOPMENT_PLACE_HISTORICAL_STATE_PLAN.xlsx",
    "03_DEVELOPMENT_PLACE_REFERENCE_READINESS_MATRIX.xlsx", "04_DEVELOPMENT_PLACE_BLOCKER_REGISTER.xlsx",
    "05_DEVELOPMENT_REGION_SLICE_CANDIDATES.xlsx", "06_DEVELOPMENT_WAVE_PLAN_V1.xlsx",
    "07_D4_D5_PLACE_MASTER.xlsx", "08_D2_D3_ACCESSIBLE_PLACE_MASTER.xlsx",
    "09_NON_URBAN_STRATEGIC_PLACE_MASTER.xlsx", "10_DEVELOPMENT_PLACE_REFERENCE_GAP_PRIORITY.xlsx",
]
workbook_errors = []
for name in workbook_names:
    path = DOC / name
    if not path.exists():
        workbook_errors.append(f"missing:{name}")
        continue
    wb = load_workbook(path, read_only=False, data_only=False)
    if set(wb.sheetnames) != {"说明", "数据", "来源"}:
        workbook_errors.append(f"sheets:{name}:{wb.sheetnames}")
    if wb["数据"].max_row < 2:
        workbook_errors.append(f"empty:{name}")
check("10 workbook package", not workbook_errors, "|".join(workbook_errors) or "10/10")

build_report = read_json(OUTPUT / "workbook_build_report.json")
check("Workbook formula error = 0", build_report["formulaErrors"] == 0, str(build_report["formulaErrors"]))
check("Workbook visual preview coverage", build_report["previews"] == 42, str(build_report["previews"]))

required_links = [
    REPO / "Docs" / "GAME_SYSTEMS_MASTER_AND_STATUS.md",
    REPO / "Docs" / "KNOWLEDGE_BASE" / "README_PROJECT_KNOWLEDGE_BASE.md",
    REPO / "Docs" / "HISTORICAL_WORLD_REFERENCE" / "README_历史世界开发参考资料索引.md",
    REPO / "Docs" / "TASK_DEVELOPMENT_PLACE_ROSTER_AND_REFERENCE_READINESS_V1.md",
]
markers = ["DEVELOPMENT_PLACE_ROSTER_V1", "DEVELOPMENT_PLACE_ROSTER_V1", "DEVELOPMENT_PLACE_ROSTER_V1", "DEVELOPMENT-PLACE-ROSTER"]
check("Knowledge Base link validation", all(path.exists() and marker in path.read_text(encoding="utf-8") for path, marker in zip(required_links, markers)), "master, KB, historical index and task record linked")

markdown_paths = [DOC / "README.md", DOC / "DEVELOPMENT_PLACE_ROSTER_AND_REFERENCE_READINESS_V1_REPORT.md", *required_links]
broken = []
pattern = re.compile(r"\[[^\]]+\]\(([^)]+)\)")
for path in markdown_paths:
    text = path.read_text(encoding="utf-8")
    for target in pattern.findall(text):
        if target.startswith(("http://", "https://", "#")):
            continue
        target_path = (path.parent / target.split("#", 1)[0]).resolve()
        if not target_path.exists():
            broken.append(f"{path.relative_to(REPO)} -> {target}")
check("Markdown broken link = 0", not broken, "|".join(broken[:10]) or "0")

utf8_failures = []
for path in [*markdown_paths, *MANIFEST.glob("*_DEVELOPMENT_REFERENCE_MANIFEST.md")]:
    try:
        path.read_text(encoding="utf-8")
    except UnicodeDecodeError as exc:
        utf8_failures.append(f"{path}:{exc}")
check("UTF-8", not utf8_failures, "|".join(utf8_failures) or "all checked text is UTF-8")

summary = {"status": "PASS" if not failures else "FAIL", "check_count": len(checks), "failure_count": len(failures), "failures": failures, "checks": checks}
(OUTPUT / "validation_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print(json.dumps({k: summary[k] for k in ("status", "check_count", "failure_count", "failures")}, ensure_ascii=False, indent=2))
raise SystemExit(0 if not failures else 1)
