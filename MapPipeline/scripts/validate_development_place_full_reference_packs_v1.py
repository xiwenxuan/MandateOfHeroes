#!/usr/bin/env python3
"""Validate the HAN-135-260 Development Place FDRP V1 delivery."""

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[2]
OUT = REPO / "outputs" / "HAN_135_260_DEVELOPMENT_PLACE_FULL_REFERENCE_PACK_V1"
DOC = REPO / "Docs" / "HISTORICAL_WORLD_REFERENCE" / "PLACE_FULL_DEVELOPMENT_REFERENCE_PACKS"
PACKS = DOC / "PACKS"
ROSTER_PATH = REPO / "outputs" / "DEVELOPMENT_PLACE_ROSTER_AND_REFERENCE_READINESS_V1" / "development_place_roster_workdata.json"
WORKDATA_PATH = OUT / "full_reference_pack_workdata.json"
REPORT_PATH = OUT / "validation_report.json"


def load(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


old = load(ROSTER_PATH)
new = load(WORKDATA_PATH)
checks = []
errors = []


def check(name, condition, detail=""):
    checks.append({"name": name, "passed": bool(condition), "detail": detail})
    if not condition:
        errors.append(f"{name}: {detail}")


old_by_id = {x["CanonicalPlaceId"]: x for x in old["roster"]}
new_by_id = {x["PlaceId"]: x for x in new["master"]}
check("roster-count", len(old_by_id) == len(new_by_id) == 72, f"old={len(old_by_id)}, new={len(new_by_id)}")
check("roster-id-set-preserved", set(old_by_id) == set(new_by_id))
for pid in sorted(old_by_id):
    check(f"name-preserved:{pid}", old_by_id[pid]["CanonicalName"] == new_by_id[pid]["CanonicalName"])
    check(f"wave-preserved:{pid}", old_by_id[pid]["RecommendedWave"] == new_by_id[pid]["Wave"])
    check(f"depth-mapped:{pid}", {"D2": "T1", "D3": "T2", "D4": "T3", "D5": "T4"}[old_by_id[pid]["DevelopmentDepth"]] == new_by_id[pid]["DevelopmentTier"])

check("tier-counts", Counter(x["DevelopmentTier"] for x in new["master"]) == Counter({"T1": 23, "T2": 33, "T3": 15, "T4": 1}), str(Counter(x["DevelopmentTier"] for x in new["master"])))
check("wave-counts-preserved", Counter(x["Wave"] for x in new["master"]) == Counter(x["RecommendedWave"] for x in old["roster"]))
check("status-counts", Counter(x["FullPackStatus"] for x in new["master"]) == Counter({"FULL_READY": 1, "FULL_READY_WITH_MODELED_GAPS": 9, "FULL_READY_WITH_UNKNOWNS": 54, "RESEARCH_BLOCKED": 8}))
check("runtime-status-counts", Counter(x["RuntimeImplementationStatus"] for x in new["master"]) == Counter({"PARTIAL": 1, "NOT_STARTED": 71}))
check("no-t0", all(x["DevelopmentTier"] != "T0" for x in new["master"]))
check("orthogonal-master-fields", all(all(k in x for k in ("DevelopmentTier", "FullPackStatus", "RuntimeImplementationStatus")) for x in new["master"]))

expected_sheets = new["sheet_contract"]
check("sheet-contract-count", len(expected_sheets) == 25 and len(set(expected_sheets)) == 25)
pack_dirs = [x for x in PACKS.iterdir() if x.is_dir()]
check("pack-directory-count", len(pack_dirs) == 72, str(len(pack_dirs)))
for slug, pack in new["packs"].items():
    pdir = PACKS / slug
    check(f"pack-dir:{slug}", pdir.is_dir())
    for filename in ("README.md", "PLACE_DEVELOPMENT_REFERENCE.xlsx", "SOURCES_AND_UNKNOWNS.md"):
        check(f"pack-file:{slug}:{filename}", (pdir / filename).is_file())
    book = pdir / "PLACE_DEVELOPMENT_REFERENCE.xlsx"
    if book.exists():
        wb = load_workbook(book, read_only=True, data_only=False)
        check(f"pack-sheet-contract:{slug}", wb.sheetnames == expected_sheets, str(wb.sheetnames))
        for sheet in expected_sheets:
            ws = wb[sheet]
            check(f"pack-sheet-has-data:{slug}:{sheet}", ws["A1"].value not in (None, "") and ws["A3"].value not in (None, "") and ws["A4"].value not in (None, ""))
        wb.close()

summary_files = [
    "DEVELOPMENT_PLACE_MASTER.xlsx", "01_FULL_PACK_COMPLETENESS_MASTER.xlsx", "02_EVENT_DEPENDENT_SITE_MASTER.xlsx",
    "03_PLACE_HISTORICAL_PERSON_COVERAGE.xlsx", "04_PLACE_CLAN_FAMILY_ESTATE_COVERAGE.xlsx",
    "05_PLACE_FACILITY_REFERENCE_COVERAGE.xlsx", "06_PLACE_POPULATION_AND_SETTLEMENT_REFERENCE.xlsx",
    "07_PLACE_INDUSTRY_RESOURCE_SUPPLY_REFERENCE.xlsx", "08_PLACE_TRANSPORT_AND_HINTERLAND_REFERENCE.xlsx",
    "09_PLACE_MILITARY_AND_EVENT_STATE_REFERENCE.xlsx", "10_PLACE_DEVELOPMENT_PACK_UPGRADE_REGISTRY.xlsx",
]
for filename in summary_files:
    path = DOC / filename
    check(f"summary-exists:{filename}", path.is_file())
    if path.exists():
        wb = load_workbook(path, read_only=True, data_only=False)
        check(f"summary-sheets:{filename}", wb.sheetnames == ["说明", "数据", "来源"], str(wb.sheetnames))
        wb.close()

events = {x["CanonicalName"]: x for x in new["event_sites"]}
for name in ("官渡", "街亭", "五丈原", "赤壁", "祁山"):
    check(f"event-site:{name}", name in events)
    if name in events:
        check(f"battle-fame-not-settlement:{name}", events[name]["HistoricalBattleFameIsSettlement"] == "NO")
        check(f"event-trigger-policy:{name}", events[name]["EventPackagePolicy"] == "APPLY_ONLY_IF_EVENT_OCCURS")
check("qishan-unknown-not-invented", events.get("祁山", {}).get("EvidenceType") == "UNKNOWN" and "UNRESOLVED" in events.get("祁山", {}).get("SpatialExistenceMode", ""))

for slug, pack in new["packs"].items():
    modes = pack["identity"][0]["SpatialExistenceMode"]
    for row in pack["event_packages"]:
        evidence = row.get("EvidenceType")
        if evidence == "NOT_APPLICABLE":
            check(f"non-event-no-package:{slug}", "EVENT_DEPENDENT_COMPLEX" not in modes)
        else:
            required = ("PackageId", "EventTrigger", "FacilityType", "Reason", "Use", "ForceWorkers", "Materials", "Duration", "PostDisposition", "DoNotApplyBeforeTrigger")
            check(f"event-package-fields:{slug}", all(row.get(k) for k in required))
            check(f"event-package-not-preapplied:{slug}", row.get("DoNotApplyBeforeTrigger") == "YES")

check("upgrade-candidates-not-auto-admitted", all(x.get("AutomaticRosterAdmission") == "NO" for x in new["upgrade_registry"]))
check("current-roster-upgrade-count", sum(x["RosterStatus"] == "CURRENT_72" for x in new["upgrade_registry"]) == 72)

manifest_paths = [Path(x["ManifestPath"]) for x in new["master"] if x["ManifestPath"]]
check("manifest-count", len(manifest_paths) == 16, str(len(manifest_paths)))
for rel in manifest_paths:
    path = REPO / rel
    check(f"manifest-exists:{rel.name}", path.is_file())
    if path.exists():
        text = path.read_text(encoding="utf-8")
        check(f"manifest-fdrp-block:{rel.name}", "<!-- FDRP-V1:BEGIN -->" in text and "ReferencePackCompleteness" in text)

registry_files = {
    "PROJECT_DOCUMENT_REGISTRY.xlsx": "doc.han.place-full-development-reference-packs.v1",
    "PROJECT_CANONICAL_DOMAIN_MAP.xlsx": "domain.historical-world.development-place-full-pack",
    "DESIGN_DECISION_REGISTRY.xlsx": "decision.fdrp.same-standard-all-tiers",
    "OPEN_DECISION_REGISTRY.xlsx": "open.fdrp.second-t4",
    "IMPLEMENTATION_GAP_REGISTER.xlsx": "gap.fdrp.runtime-materialization",
    "RESEARCH_GAP_REGISTER.xlsx": "gap.fdrp.event-sites",
    "DOCUMENT_CONFLICT_REGISTER.xlsx": "conflict.fdrp.d-depth-vs-t-tier",
}
for filename, token in registry_files.items():
    path = REPO / "Docs" / "KNOWLEDGE_BASE" / "REGISTRY" / filename
    check(f"registry-exists:{filename}", path.is_file())
    if path.exists():
        wb = load_workbook(path, read_only=True, data_only=True)
        values = "\n".join(str(cell) for row in wb["数据"].iter_rows(values_only=True) for cell in row if cell is not None)
        check(f"registry-token:{filename}", token in values)
        wb.close()

required_docs = [
    DOC / "README.md", DOC / "PLACE_FULL_DEVELOPMENT_REFERENCE_PACK_STANDARD_V1.md",
    DOC / "DEVELOPMENT_TIER_TERMINOLOGY_V1.md", DOC / "PLACE_UPGRADE_PROTOCOL_V1.md",
    DOC / "FULL_PACK_COMPLETENESS_REPORT_V1.md", REPO / "Docs" / "TASK_HAN_135_260_DEVELOPMENT_PLACE_FULL_REFERENCE_PACK_V1.md",
]
for path in required_docs:
    check(f"document-exists:{path.name}", path.is_file())
    if path.exists():
        path.read_text(encoding="utf-8")

markdown_files = list(DOC.rglob("*.md")) + [
    REPO / "Docs" / "GAME_SYSTEMS_MASTER_AND_STATUS.md",
    REPO / "Docs" / "HISTORICAL_WORLD_REFERENCE" / "README_历史世界开发参考资料索引.md",
    REPO / "Docs" / "HISTORICAL_WORLD_REFERENCE" / "DEVELOPMENT_PLACE_ROSTER_V1" / "README.md",
    REPO / "Docs" / "HISTORICAL_WORLD_REFERENCE" / "CITY_DEVELOPMENT_PACKS" / "README_CORE_CITY_DEVELOPMENT_PACKS.md",
    REPO / "Docs" / "KNOWLEDGE_BASE" / "README_PROJECT_KNOWLEDGE_BASE.md",
    REPO / "Docs" / "TASK_HAN_135_260_DEVELOPMENT_PLACE_FULL_REFERENCE_PACK_V1.md",
]
broken_links = []
for md_path in markdown_files:
    text = md_path.read_text(encoding="utf-8")
    for target in re.findall(r"\[[^\]]+\]\(([^)]+)\)", text):
        target = target.split("#", 1)[0].strip()
        if not target or "://" in target or target.startswith("mailto:"):
            continue
        if not (md_path.parent / target).resolve().exists():
            broken_links.append(f"{md_path.relative_to(REPO)} -> {target}")
check("markdown-broken-links-zero", not broken_links, " | ".join(broken_links[:10]))

build_report = load(OUT / "workbook_build_report.json")
expected_workbooks = 72 + 11 + 7
check("workbook-report-count", len(build_report["workbooks"]) == expected_workbooks, f"actual={len(build_report['workbooks'])}, expected={expected_workbooks}")
check("preview-count", build_report["previewCount"] >= 72 * 25 + 11 * 3 + 7 * 4, str(build_report["previewCount"]))
check("formula-errors", build_report["formulaErrors"] == 0)
preview_pngs = list((OUT / "previews").rglob("*.png"))
check("preview-files", len(preview_pngs) >= 72 * 25 + 11 * 3 + 7 * 4, str(len(preview_pngs)))

for forbidden in ("Assets", "ProjectSettings", "Packages"):
    check(f"reference-output-not-under:{forbidden}", not any(str(Path(x)).replace("\\", "/").startswith(str(REPO / forbidden).replace("\\", "/")) for x in build_report["workbooks"]))

report = {
    "schema": "mandate.han135260.development-place-full-reference-pack.validation.v1",
    "passed": not errors,
    "check_count": len(checks), "passed_count": sum(x["passed"] for x in checks), "failed_count": len(errors),
    "errors": errors, "checks": checks,
    "summary": new["summary"], "workbook_count": len(build_report["workbooks"]), "preview_count": len(preview_pngs),
}
REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print(json.dumps({k: report[k] for k in ("passed", "check_count", "passed_count", "failed_count", "workbook_count", "preview_count")}, ensure_ascii=False, indent=2))
if errors:
    for error in errors[:30]:
        print("ERROR", error)
    raise SystemExit(1)
