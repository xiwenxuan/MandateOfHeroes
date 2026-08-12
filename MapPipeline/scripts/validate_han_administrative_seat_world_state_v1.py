from __future__ import annotations

import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[2]
TASK = REPO / "Docs" / "HISTORICAL_WORLD_REFERENCE" / "ADMINISTRATIVE_SEAT_AND_WORLD_STATE_V1"
OUT = REPO / "outputs" / "HAN_135_260_ADMINISTRATIVE_SEAT_CANONICAL_PLACE_AND_HISTORICAL_WORLD_STATE_V1"
WORKDATA = OUT / "administrative_seat_world_state_workdata.json"
BASE = REPO / "outputs" / "HAN_135_260_HISTORICAL_WORLD_REFERENCE_V1" / "historical_world_reference_workdata.json"
REGISTRY = REPO / "Docs" / "KNOWLEDGE_BASE" / "REGISTRY"

EXPECTED_MAIN = [
    "01_135-260行政单位与重要历史治所总表.xlsx",
    "02_135-260_CanonicalPhysicalPlace_Master.xlsx",
    "03_77战略名称与CanonicalPlace关系表.xlsx",
    "04_133CoreSettlement_SeatRole_Crosswalk.xlsx",
    "05_250PriorityCounty_ImportantPlace_And_SeatReference.xlsx",
    "06_13Scenario_ImportantPlace_WorldStateSnapshot_Index.xlsx",
    "07_HistoricalMajorChangePoint_Master.xlsx",
    "08_HistoricalChangePackage_Reference.xlsx",
    "09_三国志系列重要地点名称交叉参考.xlsx",
    "10_DevelopmentRelevantPlaceCandidateMaster.xlsx",
]
EXPECTED_LUOYANG = [
    "LuoyangHistoricalStateTimeline.xlsx",
    "LuoyangMajorChangePoints.xlsx",
    "Luoyang190PrePostReference.xlsx",
    "LuoyangFacilityLifecycleReference.xlsx",
    "LuoyangPopulationMigrationReference.xlsx",
    "LuoyangPersonFamilyMovementReference.xlsx",
]
REGISTRIES = [
    "PROJECT_DOCUMENT_REGISTRY.xlsx",
    "PROJECT_CANONICAL_DOMAIN_MAP.xlsx",
    "DESIGN_DECISION_REGISTRY.xlsx",
    "OPEN_DECISION_REGISTRY.xlsx",
    "DOCUMENT_CONFLICT_REGISTER.xlsx",
    "IMPLEMENTATION_GAP_REGISTER.xlsx",
    "RESEARCH_GAP_REGISTER.xlsx",
]
ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")


data = json.loads(WORKDATA.read_text(encoding="utf-8"))
base = json.loads(BASE.read_text(encoding="utf-8"))
checks = []
failures = []


def check(name: str, condition: bool, detail: str):
    checks.append({"name": name, "passed": bool(condition), "detail": detail})
    if not condition:
        failures.append(f"{name}: {detail}")


summary = data["summary"]
check("province_count", summary["province_count"] == 13, str(summary["province_count"]))
check("province_scenario_coverage", summary["province_scenario_records"] == 13 * 13, str(summary["province_scenario_records"]))
check("commandery_equivalent_coverage", summary["commandery_equivalent_count"] == 105, str(summary["commandery_equivalent_count"]))
check("county_id_integrity", summary["county_count"] == 1182 and len({row["county_id"] for row in base["counties"]}) == 1182, str(summary["county_count"]))
check("core_settlement_coverage", summary["core_settlement_count"] == 133, str(summary["core_settlement_count"]))
check("priority_county_coverage", summary["priority_county_count"] == 250, str(summary["priority_county_count"]))
check("strategic_label_coverage", summary["strategic_label_count"] == 77, str(summary["strategic_label_count"]))
check("scenario_year_coverage", summary["scenario_year_count"] == 13, str(summary["scenario_year_count"]))
check("scenario_snapshot_coverage", summary["scenario_snapshot_records"] == 133 * 13, str(summary["scenario_snapshot_records"]))
check("major_change_points_present", summary["major_change_point_count"] >= 25, str(summary["major_change_point_count"]))
check("change_packages_present", summary["change_package_record_count"] > summary["major_change_point_count"], str(summary["change_package_record_count"]))
check("candidate_master_present", summary["development_candidate_count"] > 133, str(summary["development_candidate_count"]))
check("runtime_boundary", summary["runtime_code_changed"] is False and summary["save_schema_changed"] is False, json.dumps(summary, ensure_ascii=False))

place_ids = [row["CanonicalPlaceId"] for row in data["canonical_places"]]
check("canonical_place_ids_unique", len(place_ids) == len(set(place_ids)) == 133, f"rows={len(place_ids)} unique={len(set(place_ids))}")
strategic_ids = [row["StrategicLabelId"] for row in data["strategic_crosswalk"]]
check("strategic_ids_unique", len(strategic_ids) == len(set(strategic_ids)) == 77, f"rows={len(strategic_ids)} unique={len(set(strategic_ids))}")
check("strategic_place_refs_valid", all(row["CanonicalPlaceId"] in set(place_ids) for row in data["strategic_crosswalk"]), "all strategic labels resolve to existing Place IDs")
relation_counts = Counter(row["RelationType"] for row in data["strategic_crosswalk"])
required_relations = {"PLACE_NAME_DIRECT", "ADMIN_REGION_AS_STRATEGIC_LABEL", "PLACE_RENAME_TIMELINE", "MOVING_SEAT_REGION_LABEL", "STRATEGIC_SETTLEMENT_NOT_MAJOR_SEAT"}
check("five_relation_types", required_relations.issubset(relation_counts), json.dumps(relation_counts, ensure_ascii=False))
check("duplicate_physical_place_audit", summary["strategic_distinct_place_count"] == 75 and summary["strategic_open_conflicts"] >= 2, json.dumps(summary, ensure_ascii=False))
check("moving_seat_audit", summary["moving_seat_labels"] >= 2, str(summary["moving_seat_labels"]))
check("administrative_region_place_separation", all("HistoricalSeatReference" in row["RuntimePolicy"] or "HISTORICAL_REFERENCE" in row["RuntimePolicy"] for row in data["administrative_seats"]), "seat references remain initialization/reference only")

snapshot_keys = [(row["PlaceId"], row["ScenarioYear"]) for row in data["scenario_snapshots"]]
check("snapshot_keys_unique", len(snapshot_keys) == len(set(snapshot_keys)), f"rows={len(snapshot_keys)} unique={len(set(snapshot_keys))}")
check("no_duplicate_population_initialization", all(row["DirectScenarioOnly"] == "YES" and "RUNTIME_WORLD" in row["ContinuousPlayPolicy"] for row in data["scenario_snapshots"]), "snapshot is direct-start only")
check("same_physical_map_contract", all("SAME_CANONICAL_CELLS" in row["UrbanStateReference"] for row in data["scenario_snapshots"]), "same Cell world used")

change_ids = [row["ChangePointId"] for row in data["change_points"]]
package_ids = [row["PackageId"] for row in data["change_packages"]]
check("change_point_ids_unique", len(change_ids) == len(set(change_ids)), f"rows={len(change_ids)} unique={len(set(change_ids))}")
check("package_ids_unique", len(package_ids) == len(set(package_ids)), f"rows={len(package_ids)} unique={len(set(package_ids))}")
check("package_change_refs_valid", all(row["ChangePointId"] in set(change_ids) for row in data["change_packages"]), "all packages reference a change point")
facility_ids = {row["FacilityPermanentId"] for row in data["luoyang_facility_lifecycle"]}
target_valid = all(
    (row["TargetEntityType"] == "PLACE" and row["TargetEntityId"] in set(place_ids))
    or (row["TargetEntityType"] == "FACILITY" and row["TargetEntityId"] in facility_ids)
    for row in data["change_packages"]
)
check("change_package_targets_valid", target_valid, "PLACE and FACILITY targets use existing stable IDs")
check("canonical_post_not_forced", all("never force overwrite" in row["Notes"] for row in data["change_packages"]), "all packages preserve divergence")
check("series_legal_boundary", all(row["NeedsFurtherResearch"].startswith("YES") and "No commercial" in row["LegalBoundary"] for row in data["series_cross"]), "per-title flags remain unaudited without importing proprietary data")

source_ids = {row["source_id"] for row in data["sources"]}
missing_sources = set()
for key, rows in data.items():
    if not isinstance(rows, list):
        continue
    for row in rows:
        for field, value in row.items():
            if field in ("Source", "Sources", "SourceIds", "ProjectSource") or field.endswith("SourceIds"):
                for source_id in str(value or "").split("|"):
                    if source_id.startswith("source.") and source_id not in source_ids:
                        missing_sources.add(source_id)
check("source_references_valid", not missing_sources, "missing=" + ",".join(sorted(missing_sources)))

required_markdown = [
    TASK / "README.md",
    TASK / "HAN_135_260_ADMINISTRATIVE_SEAT_CANONICAL_PLACE_AND_HISTORICAL_WORLD_STATE_V1_REPORT.md",
    TASK / "11_LUOYANG_MAJOR_HISTORICAL_WORLD_STATES" / "LuoyangDevelopmentImplication.md",
    REPO / "Docs" / "TASK_HAN_135_260_ADMINISTRATIVE_SEAT_CANONICAL_PLACE_AND_HISTORICAL_WORLD_STATE_V1.md",
]
for name in ("CHANGAN", "YE", "XU", "CHENGDU", "XIANGYANG", "JIANGLING", "JIANYE"):
    required_markdown.append(TASK / "12_P0_PLACE_CHANGEPOINT_CANDIDATES" / f"{name}_MAJOR_CHANGEPOINT_CANDIDATES.md")
check("required_markdown_exists", all(path.exists() for path in required_markdown), f"files={len(required_markdown)}")

absolute_path_hits = []
for path in required_markdown:
    if not path.exists():
        continue
    text = path.read_text(encoding="utf-8")
    if re.search(r"\b[A-Za-z]:[\\/]", text):
        absolute_path_hits.append(str(path.relative_to(REPO)))
check("portable_markdown_paths", not absolute_path_hits, "absolute=" + ",".join(absolute_path_hits))

canonical_docs = {
    "GAME_SYSTEMS_MASTER_AND_STATUS": REPO / "Docs" / "GAME_SYSTEMS_MASTER_AND_STATUS.md",
    "DATA_AND_CONTENT_FOUNDATION": REPO / "Docs" / "DATA_AND_CONTENT_FOUNDATION.md",
    "DETERMINISTIC_SIMULATION_AND_SAVE": REPO / "Docs" / "DETERMINISTIC_SIMULATION_AND_SAVE.md",
    "GAME_VISION_AND_GAMEPLAY": REPO / "Docs" / "GAME_VISION_AND_GAMEPLAY.md",
    "MAP_ART_RESOURCE_PLAN": REPO / "Docs" / "MAP_ART_RESOURCE_PLAN.md",
    "HistoricalWorldReference": REPO / "Docs" / "HISTORICAL_WORLD_REFERENCE" / "README_历史世界开发参考资料索引.md",
    "KnowledgeBase": REPO / "Docs" / "KNOWLEDGE_BASE" / "README_PROJECT_KNOWLEDGE_BASE.md",
}
for name, path in canonical_docs.items():
    present = path.exists() and "ADMINISTRATIVE-SEAT-CANONICAL-PLACE-V1" in path.read_text(encoding="utf-8")
    check(f"canonical_doc_{name}", present, str(path.relative_to(REPO)))

expected_books = [TASK / name for name in EXPECTED_MAIN]
expected_books += [TASK / "11_LUOYANG_MAJOR_HISTORICAL_WORLD_STATES" / name for name in EXPECTED_LUOYANG]
expected_books += [REGISTRY / name for name in REGISTRIES]
check("workbooks_exist", all(path.exists() and path.stat().st_size > 0 for path in expected_books), f"count={len(expected_books)}")

formula_errors = []
structure_errors = []
for path in expected_books:
    if not path.exists():
        continue
    try:
        workbook = load_workbook(path, read_only=False, data_only=False)
        if "说明" not in workbook.sheetnames or "数据" not in workbook.sheetnames:
            structure_errors.append(f"{path.name}: missing 说明/数据")
        for sheet in workbook.worksheets:
            if (sheet.max_row or 0) < 1 or (sheet.max_column or 0) < 1:
                structure_errors.append(f"{path.name}:{sheet.title}: empty")
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and any(token in cell.value for token in ERROR_TOKENS):
                        formula_errors.append(f"{path.name}:{sheet.title}!{cell.coordinate}={cell.value}")
        workbook.close()
    except Exception as exc:
        structure_errors.append(f"{path.name}: {exc}")
check("workbook_structure", not structure_errors, "errors=" + "|".join(structure_errors[:10]))
check("formula_error_zero", not formula_errors, "errors=" + "|".join(formula_errors[:10]))

previews = list((OUT / "previews").glob("*.png")) if (OUT / "previews").exists() else []
before_previews = list((OUT / "previews_before_registry_update").glob("*.png")) if (OUT / "previews_before_registry_update").exists() else []
check("render_preview_all_workbooks", len(previews) >= 16 * 3 + 7 * 2, f"after={len(previews)}")
check("registry_pre_edit_visual_baseline", len(before_previews) == 7 * 2, f"before={len(before_previews)}")

registry_needles = {
    "PROJECT_DOCUMENT_REGISTRY.xlsx": "ADMINISTRATIVE_SEAT_AND_WORLD_STATE_V1",
    "PROJECT_CANONICAL_DOMAIN_MAP.xlsx": "HistoricalWorldGeography",
    "DESIGN_DECISION_REGISTRY.xlsx": "DEC-PLACE-001",
    "OPEN_DECISION_REGISTRY.xlsx": "OPEN-PLACE-001",
    "DOCUMENT_CONFLICT_REGISTER.xlsx": "DOC-CONFLICT-PLACE-001",
    "IMPLEMENTATION_GAP_REGISTER.xlsx": "IMP-GAP-HISTORY-001",
    "RESEARCH_GAP_REGISTER.xlsx": "RES-GAP-PLACE-001",
}
for filename, needle in registry_needles.items():
    path = REGISTRY / filename
    found = False
    if path.exists():
        wb = load_workbook(path, read_only=True, data_only=False)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                if any(needle in str(value or "") for value in row):
                    found = True
                    break
            if found:
                break
        wb.close()
    check(f"registry_{filename}", found, needle)

result = {
    "status": "PASS" if not failures else "FAIL",
    "check_count": len(checks),
    "failure_count": len(failures),
    "checks": checks,
    "failures": failures,
    "summary": summary,
}
OUT.mkdir(parents=True, exist_ok=True)
(OUT / "validation_summary.json").write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print(json.dumps(result, ensure_ascii=False, indent=2))
sys.exit(0 if not failures else 1)
